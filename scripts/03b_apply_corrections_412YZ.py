"""
apply_corrections.py
Read reviewer decisions from output/412YZ/qa_questions_412YZ.xlsx (QA Questions tab)
and apply them to output/412YZ/survey_data_412YZ.csv.

Actions:
  recode  — replace current_value with corrected_value
  clear   — blank the field (scalar) or remove the flagged token (array)
  exclude — same as clear
  accept  — no change (skipped)

apply_to rules:
  this_survey  — change only the specified survey_id row
  all_surveys  — change every row where field contains current_value
  (blank)      — treated as this_survey

For array (pipe-separated) fields, current_value is the specific token to
act on; other tokens in the same cell are left untouched.

Overwrites survey_data_412YZ.csv in place and prints a change log.

Usage:
    python scripts/03b_apply_corrections_412YZ.py
"""

import sys
from pathlib import Path

import pandas as pd
import openpyxl

BASE_DIR = Path(__file__).parent.parent

# Fields stored as pipe-separated arrays in the CSV
ARRAY_FIELDS = {
    "q7a_not_registered_reasons",
    "q10_job_barriers",
    "q11_left_job_reasons",
    "q11a_quit_reasons",
    "q13_sleeping_location",
    "q14_housing_instability_reasons",
    "q15a_visit_reasons",
    "q15b_visit_barriers",
    "q17_program_helped",
    "q24_money_methods",
    "q25_bank_account",
    "q26a_account_setup",
    "q26b_account_usage",
    "race_ethnicity",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def split_pipe(val) -> list[str]:
    if val is None or str(val).strip() == "":
        return []
    return [t.strip() for t in str(val).split("|") if t.strip()]


def join_pipe(tokens: list[str]) -> str:
    return " | ".join(tokens)


# ---------------------------------------------------------------------------
# Load corrections from workbook
# ---------------------------------------------------------------------------

def load_corrections(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["QA Questions"]
    headers = [cell.value for cell in ws[1]]

    corrections = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        rec = dict(zip(headers, row))
        action = str(rec.get("action") or "").strip().lower()
        if not action or action == "accept":
            continue
        corrections.append(rec)

    return corrections


# ---------------------------------------------------------------------------
# Apply corrections
# ---------------------------------------------------------------------------

def apply_all(df: pd.DataFrame, corrections: list[dict]) -> tuple[pd.DataFrame, list[str]]:
    log = []

    for rec in corrections:
        survey_id   = str(rec.get("survey_id")      or "").strip()
        field       = str(rec.get("field")           or "").strip()
        current_val = str(rec.get("current_value")   or "").strip()
        action      = str(rec.get("action")          or "").strip().lower()
        corrected   = str(rec.get("corrected_value") or "").strip()
        scope_raw   = str(rec.get("apply_to")        or "").strip().lower()
        apply_to    = "all_surveys" if scope_raw == "all_surveys" else "this_survey"

        if field not in df.columns:
            log.append(f"  [SKIP]   {survey_id} / {field}: column not in CSV")
            continue

        is_array = field in ARRAY_FIELDS

        # Row mask
        survey_mask = df["survey_id"] == survey_id
        if apply_to == "all_surveys":
            mask = pd.Series(True, index=df.index)
            # For coach_name_corrected, never overwrite rows that already have
            # a suggested name — only fill in the blanks.
            if field == "coach_name_corrected":
                mask = mask & (df["coach_name_corrected"].isna() | (df["coach_name_corrected"] == ""))
        else:
            mask = survey_mask

        changed = 0

        if is_array:
            # Operate on the specific token; leave other tokens intact
            def fix_token(cell_val, cur=current_val, act=action, corr=corrected):
                tokens = split_pipe(cell_val)
                if cur not in tokens:
                    return cell_val
                if act in ("clear", "exclude"):
                    tokens = [t for t in tokens if t != cur]
                elif act == "recode":
                    tokens = [corr if t == cur else t for t in tokens]
                return join_pipe(tokens)

            before = df.loc[mask, field].copy()
            df.loc[mask, field] = df.loc[mask, field].apply(fix_token)
            changed = (df.loc[mask, field] != before).sum()

        else:
            # Scalar field
            if action in ("clear", "exclude"):
                if apply_to == "all_surveys":
                    rows = mask & (df[field] == current_val)
                else:
                    rows = mask
                before = df.loc[rows, field].copy()
                df.loc[rows, field] = ""
                changed = (df.loc[rows, field] != before).sum()

            elif action == "recode":
                if apply_to == "all_surveys":
                    rows = mask & (df[field] == current_val)
                else:
                    rows = mask
                before = df.loc[rows, field].copy()
                df.loc[rows, field] = corrected
                changed = (df.loc[rows, field] != before).sum()

        scope_label = "all surveys" if apply_to == "all_surveys" else survey_id
        new_val_label = f"'{corrected}'" if action == "recode" else "blank"
        log.append(
            f"  [{action.upper():<7}] {field}: '{current_val}' -> {new_val_label}"
            f"  ({scope_label}, {changed} row(s))"
        )

    return df, log


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    csv_path  = BASE_DIR / "output" / "412YZ" / "survey_data_412YZ.csv"
    xlsx_path = BASE_DIR / "output" / "412YZ" / "qa_questions_412YZ.xlsx"

    if not csv_path.exists():
        print(f"CSV not found: {csv_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}")
        sys.exit(1)

    df = pd.read_csv(str(csv_path), encoding="utf-8-sig", dtype=str).fillna("")
    corrections = load_corrections(xlsx_path)

    print(f"Loaded {len(df)} surveys. Applying {len(corrections)} correction(s)...\n")

    df, log = apply_all(df, corrections)

    df.to_csv(str(csv_path), index=False, encoding="utf-8-sig")

    print("Change log:")
    for entry in log:
        print(entry)
    print(f"\nSaved: {csv_path}")


if __name__ == "__main__":
    main()

"""
04_analyze_412YZ.py
Compute descriptive statistics for output/412YZ/survey_data_412YZ.csv and write
output/412YZ/analysis_412YZ.xlsx with one sheet per reporting component (22 sheets).

Usage:
    python scripts/04_analyze_412YZ.py
"""

import argparse
import sys
from pathlib import Path
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
CSV_PATH = BASE_DIR / "output" / "412YZ" / "survey_data_412YZ.csv"
OUT_PATH = BASE_DIR / "output" / "412YZ" / "analysis_412YZ.xlsx"

# ---------------------------------------------------------------------------
# Display constants
# ---------------------------------------------------------------------------

AGE_MAP = {
    "16_17": "16-17 years old",
    "18_20": "18-20 years old",
    "21_23": "21-23 years old",
}
AGE_ORDER = ["16-17 years old", "18-20 years old", "21-23 years old", "Unknown"]
GENDER_ORDER = ["Female", "Male", "Trans, Non-binary"]

Q1_FIELDS = [
    ("q1_trustworthy",      "Is trustworthy"),
    ("q1_reliable",         "Is reliable"),
    ("q1_values_opinions",  "Values my opinions about my life"),
    ("q1_available",        "Is available to me when I need them"),
    ("q1_heard_understood", "Makes me feel heard and understood"),
]

Q20_FIELDS = [
    ("q20_people_care",      "People around me care about my success"),
    ("q20_no_judgment",      "I feel accepted without judgment"),
    ("q20_diversity_valued", "Diversity of backgrounds is valued"),
    ("q20_treated_fairly",   "I am treated fairly"),
    ("q20_safe_sharing",     "I feel safe sharing my thoughts"),
]

RACE_GROUP_ORDER = [
    "Black", "White", "Multiracial", "Other single racial identity",
    "Hispanic or Latinx", "Asian", "Native American or Native Hawaiian",
    "Prefer not to answer", "Other",
]

HOUSING_LABEL_MAP = {
    "stable":          "Safe and stable",
    "safe_not_90days": "Safe, cannot stay 90 days",
    "90days_not_safe": "Can stay, but not safe",
    "no_place":        "No place to stay",
}
HOUSING_ORDER = ["stable", "safe_not_90days", "90days_not_safe", "no_place"]

SLEEP_LABEL_MAP = {
    "friends_family":     "Home, w/ Family/Friends",
    "shelter":            "Homeless shelter",
    "couch_surfing":      "Couch surfing",
    "car":                "In a car",
    "outside":            "Outside",
    "abandoned_building": "Abandoned building",
    "other":              "Other",
}

Q14_LABEL_MAP = {
    "lost_informal_housing": "Family/friends no longer let me stay",
    "left_unsafe":           "Left - felt unsafe",
    "evicted_nonpayment":    "Evicted, didn't/couldn't pay rent",
    "evicted_other":         "Evicted, other reason",
    "other":                 "Other",
}

Q14_OTHER_TEXT_MAP = {
    # → left_unsafe
    "got a place on my ex and he kept trying to break in": "left_unsafe",
    "had to leave apt because bug infestation":             "left_unsafe",
    "the things i put money into at fixed never get fixed within": "left_unsafe",
    "dangerous family members":                            "left_unsafe",
    # → evicted_other
    "my landlord is terrible":                             "evicted_other",
    "kicked out / house phas-":                            "evicted_other",
    "lease end, rent increase":                            "evicted_other",
    "facing possible eviction":                            "evicted_other",
    "my lease ended and finding other housing was unattainable currently": "evicted_other",
}

Q8_LABEL_MAP = {
    "yes_full_time":        "Full time",
    "yes_part_time":        "Part time",
    "job_training_program": "Job training program",
    "no":                   "Not working",
}
Q8_ORDER = ["yes_full_time", "yes_part_time", "job_training_program", "no"]

Q8A_LABEL_MAP = {
    "less_3mo": "Less than 3 months",
    "3_6mo":    "3 to 6 months",
    "more_6mo": "More than 6 months",
}
Q8A_ORDER = ["less_3mo", "3_6mo", "more_6mo"]

Q10_LABEL_MAP = {
    "transportation":        "Transportation issues",
    "not_getting_called":    "Applying and not getting called",
    "mental_physical_health":"Mental or physical health",
    "limited_experience":    "Limited work experience",
    "childcare":             "Childcare/parenting challenges",
    "no_references":         "Weak references",
    "interview_skills":      "Interview skills",
    "no_diploma":            "Lack of HS diploma/GED",
    "criminal_background":   "Criminal background",
    "drugs_alcohol":         "Use of drugs or alcohol",
    "something_else":        "Something else",
}

Q10_OTHER_TEXT_MAP = {
    "noting on my resume": "limited_experience",
    "nothing on my resume": "limited_experience",
    "distance": "transportation",
    "they say they are hiring but are turning people away": "not_getting_called",
}

Q10_OTHER_TEXT_THEME_MAP = {
    "bad job market": "Bad job market",
    "bad job marketing": "Bad job market",
}

Q11_LABEL_MAP = {
    "quit":               "Quit",
    "found_better":       "Found a better job",
    "seasonal":           "Seasonal/temporary",
    "fired_attendance":   "Fired: Missed work",
    "pregnancy_parenting":"Pregnant or Parenting",
    "fired_performance":  "Fired: Poor performance",
    "illness":            "Illness",
    "other":              "Other",
}

Q11_OTHER_TEXT_MAP = {
    "got sick": "illness",
    "medical issue": "illness",
    "pregnant": "pregnancy_parenting",
    "child needs and depression": "pregnancy_parenting",
    "fired for calling off to take my son to the hospital": "fired_attendance",
    "the job was over": "seasonal",
    "school": "quit",
    "moved couldn't get to work": "transportation",
    "income": "low_pay_hours",
    "my client tried to fight me, hr let me go.": "poor_conditions",
}

Q11_OTHER_TEXT_THEME_MAP = {
    "medical issue": "Health-related issue",
    "got sick": "Health-related issue",
}

Q11A_LABEL_MAP = {
    "low_pay_hours":          "Low pay or not enough hours",
    "schedule_conflict":      "Schedule conflict",
    "lack_of_support":        "Lack of support or respect",
    "poor_conditions":        "Poor working conditions",
    "mental_emotional_health":"Mental/emotional health",
    "transportation":         "Transportation challenges",
    "not_good_fit":           "Not a good fit",
    "personal_family":        "Personal or family reasons",
    "other":                  "Other",
}

Q6_LABEL_MAP = {
    "yes":            "Yes",
    "learners_permit":"Learner's Permit",
    "no":             "No",
}
Q6_ORDER = ["yes", "learners_permit", "no"]

Q6A_LABEL_MAP = {
    "own_reliable":    "Own reliable vehicle",
    "own_unreliable":  "Own vehicle, but unreliable",
    "share_reliable":  "Share reliable vehicle",
    "share_unreliable":"Share vehicle, but unreliable",
    "borrow":          "Can borrow a car",
    "no_access":       "No reliable vehicle access",
}
Q6A_ORDER = ["own_reliable", "share_reliable", "borrow", "own_unreliable", "share_unreliable", "no_access"]

Q9_LABEL_MAP = {
    "public_transit":   "Bus or public transportation",
    "driving_self":     "Driving myself",
    "rides_from_others":"Getting rides from someone else",
    "rideshare":        "RideShare app (Lyft, Uber)",
    "active_transport": "Walking, biking, scooter, etc.",
    "other":            "Other",
}
Q9_ORDER = ["public_transit", "driving_self", "rides_from_others", "rideshare", "active_transport", "other"]
NO_CAR_SET = {"rideshare", "rides_from_others", "active_transport"}

Q7A_LABEL_MAP = {
    "vote_wont_matter": "My vote won't make a difference",
    "dont_understand":  "I do not understand politics",
    "dont_know_how":    "I do not know how to register",
    "not_old_enough":   "I am not old enough to vote",
    "other":            "Other",
}

Q7A_OTHER_TEXT_MAP = {
    "doesn't believe in process": "vote_wont_matter",
}

Q7A_OTHER_TEXT_THEME_MAP = {
    "i am not registered as a pa id yet,": "Has not completed registration process",
    "haven't registered yet": "Has not completed registration process",
    "just didn't register": "Has not completed registration process",
    "i just haven't gotten registered yet": "Has not completed registration process",
    "im a felon": "Eligibility barrier",
    "not a citizen": "Eligibility barrier",
    "i chose not to vote": "Not interested",
    "don't care for it": "Not interested",
    "because it's my choice": "Not interested",
    "im just not sure": "Unsure",
}

Q15_LABEL_MAP = {
    "every_week":               "Every week",
    "1_3_times_per_month":      "1-3 times per month",
    "less_than_once_per_month": "Less than once per month",
    "never":                    "Never",
}
Q15_ORDER = ["every_week", "1_3_times_per_month", "less_than_once_per_month", "never"]
FREQUENT_VISIT = {"every_week", "1_3_times_per_month"}
INFREQUENT_VISIT = {"less_than_once_per_month", "never"}

Q16_LABEL_MAP = {
    "agree": "Agree",
    "somewhat_agree": "Somewhat agree",
    "disagree": "Disagree",
    "unsure": "Unsure, I don't have clear goals right now",
}
Q16_ORDER = ["agree", "somewhat_agree", "disagree", "unsure"]

Q15A_LABEL_MAP = {
    "see_coach_staff":   "See my Youth Coach or other Zone Staff",
    "work_on_goals":     "Work toward my goals",
    "safe_place":        "Be in a safe place",
    "computers":         "Access to computers",
    "laundry_shower":    "Do laundry, shower, etc.",
    "food":              "Eat food",
    "escape_problems":   "Escape problems/issues",
    "health_counseling": "Health care and/or counseling",
    "learn_skills":      "Learn new things, gain skills",
    "service_providers": "Meet with service providers",
    "socialize":         "Socialize/see friends",
    "scheduled_activity":"Participate in a scheduled activity",
    "other":             "Other",
}

Q15B_LABEL_MAP = {
    "better_activities": "More activities that interest me",
    "coach_invitation":  "Invitation from my Youth Coach",
    "more_info":         "Knowing more about activities",
    "other":             "Other",
}

Q15B_OTHER_TEXT_MAP = {
    "i don't know the location or how to get there i also don't know my 412 youth coach": "more_info",
    "activities with money": "better_activities",
}

Q15B_OTHER_TEXT_THEME_MAP = {
    "quest times": "Unclear or unspecified barrier",
    "less drama in da space": "Environment or social comfort",
    "buss passworks": "Transportation or distance",
    "i am just busy with work and school.": "Work or school schedule conflicts",
    "housing moving truck help": "Specific service or support need",
    "possibly help looking for a vehicle in the future": "Transportation or distance",
    "just being there and chillen and doing what they want me to do": "Activities or atmosphere",
    "idk": "Unclear or unspecified barrier",
    "being unsuspended": "Access restrictions",
    "if my schedule wasn't so busy with school. i also live in turtle creek which is far by bus.": "Work or school schedule conflicts",
    "if i didn't work so much": "Work or school schedule conflicts",
    "visit a doctor": "Specific service or support need",
    "better transportation": "Transportation or distance",
    "due to being in the military and not stationed in pittsburgh i would have to visit whenever i have leave days saved up to come.": "Transportation or distance",
    "not being around other people": "Environment or social comfort",
    "the hours of operation conflict with my job": "Work or school schedule conflicts",
    "food good food": "Food or amenities",
    "also it out of my way but i'm very busy": "Transportation or distance",
    "i don't know. but it's hard when i don't have a reliable ride": "Transportation or distance",
    "a more stable personal schedule": "Work or school schedule conflicts",
    "i live in erie": "Transportation or distance",
    "would love to come but the hours i work don't make it possible": "Work or school schedule conflicts",
    "i would have all the kids cuz my boyfriend works as well": "Childcare or family responsibilities",
    "s car ride there and back": "Transportation or distance",
    "nothing in particular, i'm just antisocial and don't like being in crowded confined spaces": "Environment or social comfort",
}

Q17_LABEL_MAP = {
    "education":              "Finish or further my education",
    "job":                    "Get or keep a job",
    "housing":                "Find or maintain housing",
    "health_counseling":      "Access health care/counseling",
    "positive_relationships": "Establish positive relationships",
    "handle_problems":        "Figure out how to handle problems",
    "drivers_license":        "Get my driver's license",
    "parenting":              "Improve parenting skills",
    "everyday_skills":        "Learn everyday skills",
    "decision_making":        "Develop decision-making skills",
    "vital_documents":        "Obtain vital documents",
    "future":                 "Think about my future",
    "something_else":         "Something else",
}

Q25_LABEL_MAP = {
    "checking":    "Checking account",
    "savings":     "Savings account",
    "had_in_past": "Had an account in the past",
    "never_had":   "Never had an account",
}
Q25_ORDER = ["checking", "savings", "had_in_past", "never_had"]

Q24_LABEL_MAP = {
    "digital_apps":  "Venmo, Zelle, CashApp, etc.",
    "bank_account":  "Bank account",
    "cash_at_home":  "Saving/storing cash at home",
    "paypal":        "PayPal",
    "check_cashing": "Check cashing service (ex. ACE)",
    "money_order":   "Money Order",
    "other":         "Other",
}

Q26B_LABEL_MAP = {
    "debit_card":       "Bank credit/debit card",
    "online_banking":   "Online banking/Bank app",
    "direct_deposit":   "Direct deposit",
    "saving":           "Saving money",
    "budgeting":        "Budgeting money",
    "atm":              "ATMs",
    "keep_safe":        "Keep my money safe",
    "transferring":     "Transferring money",
    "cashing_checks":   "Cashing checks",
    "paying_bills":     "Paying household bills",
    "writing_checks":   "Writing checks",
    "in_person_banking":"Banking in person",
    "none":             "None - I don't use my account",
    "other":            "Other",
}

ORIENT_ORDER = [
    "Heterosexual/Straight", "Bisexual", "Gay, Lesbian, or Same Gender Loving",
    "Asexual", "Pansexual", "Queer", "Demisexual",
    "I am not sure yet", "I don't understand the question", "No answer provided",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def split_pipe(val):
    if not str(val).strip():
        return []
    return [t.strip() for t in str(val).split("|") if t.strip()]


def pct_str(n, d, decimals=0):
    if d == 0:
        return ""
    val = round(100 * n / d, decimals)
    return f"{int(val)}%" if decimals == 0 else f"{val:.{decimals}f}%"


def age_label(code):
    return AGE_MAP.get(str(code).strip(), "Unknown")


def gender_group(g):
    g = str(g).strip()
    if g == "Female":
        return "Female"
    if g == "Male":
        return "Male"
    if not g:
        return "No answer"
    return "Trans, Non-binary"


def token_to_race_group(t):
    t = str(t).strip()
    if "Black" in t:
        return "Black"
    if "White" in t:
        return "White"
    if "Multi" in t:
        return "Multiracial"
    if "Hispanic" in t or "Latinx" in t:
        return "Hispanic or Latinx"
    if "Asian" in t:
        return "Asian"
    if "Native" in t:
        return "Native American or Native Hawaiian"
    if "Prefer not" in t:
        return "Prefer not to answer"
    return "Other"


def crosstab_age(df, field, label_map, age_col="_age", sort_by_total=True):
    """
    Count pipe-sep tokens in `field` by age group.
    Returns DataFrame: Label | 16-17... | 18-20... | 21-23... | Unknown | Total
    """
    rows = []
    for _, row in df.iterrows():
        for token in split_pipe(row[field]):
            rows.append({"_label": token, "_age": row[age_col]})
    if not rows:
        return pd.DataFrame()
    tmp = pd.DataFrame(rows)
    ct = tmp.groupby("_label")["_age"].value_counts().unstack(fill_value=0)
    for a in AGE_ORDER:
        if a not in ct.columns:
            ct[a] = 0
    ct = ct[AGE_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    ct = ct.reset_index()
    ct["Label"] = ct["_label"].map(label_map).fillna(ct["_label"])
    ct = ct.drop(columns=["_label"])
    ct = ct[["Label"] + AGE_ORDER + ["Total"]]
    if sort_by_total:
        ct = ct.sort_values("Total", ascending=False).reset_index(drop=True)
    return ct


def race_crosstab(df, multi_count=False):
    """
    Race x gender cross-tab.
    multi_count=False: each youth counted once (2+ tokens -> Multi-Racial).
    multi_count=True:  each token counted independently.
    """
    df2 = df[df["_gender"] != "No answer"].copy()

    # Groups collapsed into 'Other single racial identity' when counting each youth once
    _SMALL_GROUPS = {"Hispanic or Latinx", "Asian", "Native American or Native Hawaiian", "Other"}

    if not multi_count:
        def race_once(raw):
            tokens = split_pipe(raw)
            if not tokens:
                return None
            if len(tokens) >= 2:
                return "Multiracial"
            group = token_to_race_group(tokens[0])
            return "Other single racial identity" if group in _SMALL_GROUPS else group
        df2 = df2.copy()
        df2["_race"] = df2["race_ethnicity"].apply(race_once)
        df2 = df2.dropna(subset=["_race"])
        ct = pd.crosstab(df2["_race"], df2["_gender"])
        n_total = len(df2)
    else:
        race_df = df2[df2["race_ethnicity"].astype(str).str.strip() != ""].copy()
        rows = []
        for _, row in race_df.iterrows():
            for t in split_pipe(row["race_ethnicity"]):
                rows.append({"_race": token_to_race_group(t), "_gender": row["_gender"]})
        if not rows:
            return pd.DataFrame()
        tmp = pd.DataFrame(rows)
        ct = pd.crosstab(tmp["_race"], tmp["_gender"])
        n_total = len(race_df)

    for g in GENDER_ORDER:
        if g not in ct.columns:
            ct[g] = 0
    ct = ct[GENDER_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    ct["Percent"] = ct["Total"].apply(lambda n: pct_str(n, n_total))
    ct = ct.reset_index().rename(columns={"_race": "Race/Ethnicity"})

    present = [r for r in RACE_GROUP_ORDER if r in ct["Race/Ethnicity"].values]
    extra = [r for r in ct["Race/Ethnicity"].values if r not in RACE_GROUP_ORDER]
    ct = ct.set_index("Race/Ethnicity").reindex(present + extra).reset_index()

    total_row = {"Race/Ethnicity": "Total", "Percent": ""}
    for col in GENDER_ORDER + ["Total"]:
        total_row[col] = ct[col].sum()
    return pd.concat([ct, pd.DataFrame([total_row])], ignore_index=True)


# ---------------------------------------------------------------------------
# Section computations
# ---------------------------------------------------------------------------

def sec01_age(df):
    counts = df["_age"].value_counts().reindex(AGE_ORDER, fill_value=0)
    t = pd.DataFrame({"Age": counts.index, "Count": counts.values})
    return pd.concat([t, pd.DataFrame([{"Age": "Total", "Count": t["Count"].sum()}])],
                     ignore_index=True)


def sec02_gender_orient(df):
    df2 = df[df["_gender"] != "No answer"].copy()
    df2["_orient"] = df2["sexual_orientation"].apply(
        lambda x: x.strip() if x.strip() else "Unknown"
    )
    # Merge Gay or Lesbian and Same Gender Loving (matches prior report format)
    df2["_orient"] = df2["_orient"].replace(
        {"Gay or Lesbian": "Gay, Lesbian, or Same Gender Loving",
         "Same Gender Loving": "Gay, Lesbian, or Same Gender Loving"}
    )

    ct = pd.crosstab(df2["_orient"], df2["_gender"])
    for g in GENDER_ORDER:
        if g not in ct.columns:
            ct[g] = 0
    ct = ct[GENDER_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    ct = ct.reset_index().rename(columns={"_orient": "Sexual Orientation"})
    ct["__unknown_last"] = ct["Sexual Orientation"].eq("Unknown")
    ct = ct.sort_values(
        by=["__unknown_last", "Total", "Sexual Orientation"],
        ascending=[True, False, True],
    ).drop(columns=["__unknown_last"]).reset_index(drop=True)

    header = pd.DataFrame([{
        "Sexual Orientation": "Number of Youth",
        "Female": (df2["_gender"] == "Female").sum(),
        "Male": (df2["_gender"] == "Male").sum(),
        "Trans, Non-binary": (df2["_gender"] == "Trans, Non-binary").sum(),
        "Total": len(df2),
    }])
    return pd.concat([header, ct], ignore_index=True)


def sec03_race_once(df):
    return race_crosstab(df, multi_count=False)


def sec04_race_multi(df):
    return race_crosstab(df, multi_count=True)


def sec05_q1(df):
    TOP2 = {"4", "5"}
    rows = []
    for field, label in Q1_FIELDS:
        valid = df[field].astype(str)
        valid = valid[valid.str.strip() != ""]
        n_t2 = valid.isin(TOP2).sum()
        rows.append({
            "My Youth Coach...": label,
            "n": len(valid),
            "% Often or All the Time (4-5)": pct_str(n_t2, len(valid)),
        })
    return pd.DataFrame(rows)


def sec06_communication(df):
    Q2_LABEL_MAP = {
        "almost_every_day":       "Almost every day",
        "about_once_a_week":      "About once a week",
        "1_2_times_per_month":    "1-2 times per month",
        "less_than_once_a_month": "Less than once a month",
    }
    Q3_LABEL_MAP = {
        "not_enough": "Not enough",
        "good_amount": "Good amount",
        "too_much":    "Too much",
    }

    q3 = df["q3_communication_level"].replace("", pd.NA).dropna()
    total_q3 = len(q3)
    part_a = pd.DataFrame([{
        "Communication Level (Q3)": Q3_LABEL_MAP.get(k, k),
        "Count": v,
        "Percent": pct_str(v, total_q3),
    } for k, v in q3.value_counts().items()])

    not_enough = df[df["q3_communication_level"] == "not_enough"]
    q2 = not_enough["q2_communication_frequency"].replace("", pd.NA).dropna().value_counts()
    part_b = pd.DataFrame([{
        "Q2 Frequency (among 'Not Enough' responses)": Q2_LABEL_MAP.get(k, k),
        "Count": v,
    } for k, v in q2.items()])

    return {"Q3 Communication Level": part_a, "Q2 Frequency (Not Enough only)": part_b}


def sec07_housing(df):
    sleep_codes = []
    for v in df[df["q12_housing_stability"] != "stable"]["q13_sleeping_location"]:
        for t in split_pipe(v):
            if t not in sleep_codes:
                sleep_codes.append(t)
    ordered_sleep = [c for c in SLEEP_LABEL_MAP if c in sleep_codes]
    ordered_sleep += [c for c in sleep_codes if c not in SLEEP_LABEL_MAP]

    valid = df[df["q12_housing_stability"] != ""]
    total = len(valid)

    rows = []
    for code in HOUSING_ORDER:
        sub = valid[valid["q12_housing_stability"] == code]
        if sub.empty:
            continue
        row_data = {
            "Housing Status": HOUSING_LABEL_MAP.get(code, code),
            "Count": len(sub),
            "Percent": pct_str(len(sub), total),
        }
        for sc in ordered_sleep:
            cnt = sub["q13_sleeping_location"].apply(lambda v: sc in split_pipe(v)).sum()
            row_data[SLEEP_LABEL_MAP.get(sc, sc)] = cnt if cnt > 0 else ""
        rows.append(row_data)

    sleep_labels = [SLEEP_LABEL_MAP.get(sc, sc) for sc in ordered_sleep]
    total_row = {"Housing Status": "Total", "Count": total, "Percent": "100%"}
    for sl in sleep_labels:
        total_cnt = sum(int(r.get(sl, 0) or 0) for r in rows)
        total_row[sl] = total_cnt if total_cnt > 0 else ""
    rows.append(total_row)

    return pd.DataFrame(rows)


def sec08_housing_reasons(df):
    summary_df, _, _ = build_sec08_housing_reasons_outputs(df)
    return summary_df


def build_sec08_housing_reasons_outputs(df):
    UNSTABLE = {"safe_not_90days", "90days_not_safe", "no_place"}

    stable_row = {"Reason": "No unstable housing in past 6 months"}
    for age in AGE_ORDER:
        stable_row[age] = int(((df["_age"] == age) & ~df["q12_housing_stability"].isin(UNSTABLE)).sum())
    stable_row["Total"] = int((~df["q12_housing_stability"].isin(UNSTABLE)).sum())

    unstable_row = {"Reason": "Total youth reporting unstable housing"}
    for age in AGE_ORDER:
        unstable_row[age] = int(((df["_age"] == age) & df["q12_housing_stability"].isin(UNSTABLE)).sum())
    unstable_row["Total"] = int(df["q12_housing_stability"].isin(UNSTABLE).sum())

    header_df = pd.DataFrame([stable_row, unstable_row])
    header_df = header_df.reindex(columns=["Reason"] + AGE_ORDER + ["Total"])

    age_counters = {age: Counter() for age in AGE_ORDER}
    total_counter = Counter()
    mapping_rows = []
    residual_counter = Counter()
    residual_theme_texts = {}

    answered = df[df["q14_housing_instability_reasons"].str.strip() != ""]
    for _, row in answered.iterrows():
        age_bucket = row["_age"]
        base_codes = split_pipe(row["q14_housing_instability_reasons"])
        other_text = (row["q14_other_text"] or "").strip()
        normalized_other = " ".join(other_text.lower().split())
        mapped_code = Q14_OTHER_TEXT_MAP.get(normalized_other, "")
        has_mapping = bool(mapped_code)

        final_codes = [code for code in base_codes if not (code == "other" and has_mapping)]
        if mapped_code and mapped_code not in final_codes:
            final_codes.append(mapped_code)

        for code in final_codes:
            if age_bucket in age_counters:
                age_counters[age_bucket][code] += 1
            total_counter[code] += 1

        if "other" in base_codes:
            residual_reason = ""
            residual_theme = ""
            if not has_mapping:
                residual_reason = other_text if other_text else "No free-text provided"
                residual_theme = residual_reason
                residual_counter[residual_theme] += 1
                residual_theme_texts.setdefault(residual_theme, set()).add(residual_reason)
            mapping_rows.append({
                "survey_id": row["survey_id"],
                "age_range": age_bucket,
                "housing_stability": row["q12_housing_stability"],
                "original_q14_codes": " | ".join(base_codes),
                "q14_other_text": other_text,
                "mapped_existing_reason": Q14_LABEL_MAP.get(mapped_code, ""),
                "final_q14_codes": " | ".join(final_codes),
                "left_under_other": "Yes" if not has_mapping else "No",
                "residual_other_reason": residual_reason,
            })

    ct_rows = []
    order = list(Q14_LABEL_MAP.keys()) + [k for k in total_counter if k not in Q14_LABEL_MAP]
    for code in order:
        n = total_counter.get(code, 0)
        if n == 0:
            continue
        row_data = {"Reason": Q14_LABEL_MAP.get(code, code)}
        for age in AGE_ORDER:
            row_data[age] = age_counters[age].get(code, 0)
        row_data["Total"] = n
        ct_rows.append(row_data)
    ct_rows.sort(key=lambda r: (-r["Total"], r["Reason"]))

    if ct_rows:
        ct_df = pd.DataFrame(ct_rows, columns=["Reason"] + AGE_ORDER + ["Total"])
        summary_df = pd.concat([header_df, ct_df], ignore_index=True)
    else:
        summary_df = header_df

    mapping_df = pd.DataFrame(mapping_rows)
    repeated_residual_df = pd.DataFrame([
        {
            "Residual other reason": reason,
            "Total Youth": count,
            "Source texts": " | ".join(sorted(residual_theme_texts.get(reason, set()))),
        }
        for reason, count in residual_counter.items() if count > 1
    ])
    if not repeated_residual_df.empty:
        repeated_residual_df = repeated_residual_df.sort_values(
            by=["Total Youth", "Residual other reason"],
            ascending=[False, True],
        ).reset_index(drop=True)

    return summary_df, mapping_df, repeated_residual_df


def sec08_housing_reasons_reference(df):
    summary_df, mapping_df, repeated_residual_df = build_sec08_housing_reasons_outputs(df)
    return {
        "Updated Q14 housing reasons table": summary_df,
        "Q14 other mapping audit": mapping_df,
        "Repeated residual other reasons": repeated_residual_df,
    }


def sec09_education(df):
    total_resp = df[df["q5_school_status"] != ""].shape[0]
    in_school = df[(df["q5_school_status"] != "") & (df["q5_school_status"] != "not_in_school")]
    not_school = df[df["q5_school_status"] == "not_in_school"]

    rows = []
    rows.append({
        "Group": "Currently Enrolled in School",
        "Level": "Total Currently Enrolled",
        "Count": len(in_school),
        "Percent of All": pct_str(len(in_school), total_resp),
    })
    for code, label in [("graduate", "Graduate School"), ("college_career", "College/Vocational"),
                        ("ged", "GED Program"), ("high_school", "High School")]:
        n = (in_school["q5_school_status"] == code).sum()
        rows.append({
            "Group": "Currently Enrolled in School",
            "Level": label,
            "Count": n,
            "Percent of All": pct_str(n, total_resp),
        })

    rows.append({
        "Group": "Not in School",
        "Level": "Total Not in School",
        "Count": len(not_school),
        "Percent of All": pct_str(len(not_school), total_resp),
    })
    for code, label in [("college_degree", "College Degree or Certificate"),
                        ("some_college", "Some College"),
                        ("hs_diploma_ged", "HS Diploma or GED"),
                        ("some_hs", "Some High School")]:
        n = (not_school["q5a_highest_education"] == code).sum()
        if n > 0:
            rows.append({
                "Group": "Not in School",
                "Level": label,
                "Count": n,
                "Percent of All": pct_str(n, total_resp),
            })

    rows.append({"Group": "Total", "Level": "", "Count": total_resp, "Percent of All": "100%"})
    return pd.DataFrame(rows)


def sec10_employment(df):
    rows = []
    for code in Q8_ORDER:
        row_data = {"Employment Status": Q8_LABEL_MAP[code]}
        for age in AGE_ORDER:
            row_data[age] = int(((df["_age"] == age) & (df["q8_employment_status"] == code)).sum())
        row_data["Total"] = int((df["q8_employment_status"] == code).sum())
        rows.append(row_data)

        # Insert Seeking Employment sub-row after job training and after not working
        if code in ("job_training_program", "no"):
            seeking = (df["q8_employment_status"] == code) & (df["q8b_job_seeking"] == "yes")
            seek_row = {"Employment Status": "  Seeking Employment"}
            for age in AGE_ORDER:
                seek_row[age] = int(((df["_age"] == age) & seeking).sum())
            seek_row["Total"] = int(seeking.sum())
            rows.append(seek_row)

    total_row = {"Employment Status": "Total"}
    for age in AGE_ORDER:
        total_row[age] = int(((df["_age"] == age) & (df["q8_employment_status"] != "")).sum())
    total_row["Total"] = int((df["q8_employment_status"] != "").sum())
    rows.append(total_row)

    return pd.DataFrame(rows)


def sec11_job_tenure(df):
    employed = df[df["q8_employment_status"].isin(["yes_full_time", "yes_part_time"])]
    rows = []
    for code in Q8A_ORDER:
        ft = employed[(employed["q8a_job_tenure"] == code) & (employed["q8_employment_status"] == "yes_full_time")].shape[0]
        pt = employed[(employed["q8a_job_tenure"] == code) & (employed["q8_employment_status"] == "yes_part_time")].shape[0]
        rows.append({"Job Tenure": Q8A_LABEL_MAP[code], "Full time": ft, "Part time": pt, "Total": ft + pt})
    rows.append({
        "Job Tenure": "Total",
        "Full time": employed[employed["q8_employment_status"] == "yes_full_time"].shape[0],
        "Part time": employed[employed["q8_employment_status"] == "yes_part_time"].shape[0],
        "Total": len(employed),
    })
    return pd.DataFrame(rows)


def sec12_job_barriers(df):
    summary_df, _, _ = build_sec12_job_barriers_outputs(df)
    return summary_df


def build_sec12_job_barriers_outputs(df):
    answered = df[df["q10_job_barriers"].str.strip() != ""].copy()
    denom = int(len(answered))
    pct_col = f"Percent of Youth with Barriers (N = {denom})"

    counter = Counter()
    mapping_rows = []
    residual_counter = Counter()
    residual_theme_texts = {}

    for row in answered.itertuples(index=False):
        base_codes = split_pipe(row.q10_job_barriers)
        other_text = (row.q10_something_else_text or "").strip()
        normalized_other = " ".join(other_text.lower().split())
        mapped_code = Q10_OTHER_TEXT_MAP.get(normalized_other, "")

        final_codes = [code for code in base_codes if not (code == "something_else" and mapped_code)]
        if mapped_code and mapped_code not in final_codes:
            final_codes.append(mapped_code)
        for code in final_codes:
            counter[code] += 1

        if "something_else" in base_codes:
            residual_reason = ""
            residual_theme = ""
            if not mapped_code:
                residual_reason = other_text if other_text else "No free-text provided"
                residual_theme = Q10_OTHER_TEXT_THEME_MAP.get(normalized_other, residual_reason)
                residual_counter[residual_theme] += 1
                residual_theme_texts.setdefault(residual_theme, set()).add(residual_reason)
            mapping_rows.append({
                "survey_id": row.survey_id,
                "employment_status": row.q8_employment_status,
                "original_q10_codes": " | ".join(base_codes),
                "something_else_text": other_text,
                "mapped_existing_reason": Q10_LABEL_MAP.get(mapped_code, ""),
                "final_q10_codes": " | ".join(final_codes),
                "left_under_something_else": "Yes" if not mapped_code else "No",
                "residual_something_else_reason": residual_reason,
                "residual_something_else_theme": residual_theme,
            })

    rows = []
    order = list(Q10_LABEL_MAP.keys()) + [k for k in counter if k not in Q10_LABEL_MAP]
    for code in order:
        n = counter.get(code, 0)
        if n >= 2:
            rows.append({
                "Barrier": Q10_LABEL_MAP.get(code, code),
                "Total Youth": n,
                pct_col: pct_str(n, denom),
            })
    rows.sort(key=lambda r: (-r["Total Youth"], r["Barrier"]))

    mapping_df = pd.DataFrame(mapping_rows)
    repeated_residual_df = pd.DataFrame([
        {
            "Residual something else theme": reason,
            "Total Youth": count,
            "Source texts": " | ".join(sorted(residual_theme_texts.get(reason, set()))),
        }
        for reason, count in residual_counter.items() if count > 1
    ])
    if not repeated_residual_df.empty:
        repeated_residual_df = repeated_residual_df.sort_values(
            by=["Total Youth", "Residual something else theme"],
            ascending=[False, True],
        ).reset_index(drop=True)

    return pd.DataFrame(rows), mapping_df, repeated_residual_df


def sec12_job_barriers_reference(df):
    summary_df, mapping_df, repeated_residual_df = build_sec12_job_barriers_outputs(df)
    return {
        "Updated Q10 job barriers table": summary_df,
        "Q10 something_else mapping audit": mapping_df,
        "Repeated residual something_else reasons": repeated_residual_df,
    }


def sec13_left_job(df):
    summary_df, _, _ = build_sec13_left_job_outputs(df)
    return summary_df


def build_sec13_left_job_outputs(df):
    total_left = df[df["q11_left_job_reasons"] != ""].shape[0]
    pct_col = f"Percent Who Left a Job (N = {total_left})"

    counter = Counter()
    quit_counter = Counter()
    mapping_rows = []
    residual_counter = Counter()
    residual_theme_texts = {}

    for row in df.itertuples(index=False):
        base_codes = split_pipe(row.q11_left_job_reasons)
        base_quit_codes = split_pipe(row.q11a_quit_reasons)
        other_text = (row.q11_other_text or "").strip()
        normalized_other = " ".join(other_text.lower().split())
        mapped_code = Q11_OTHER_TEXT_MAP.get(normalized_other, "")
        mapped_parent_code = ""
        mapped_quit_code = ""
        if mapped_code in Q11_LABEL_MAP:
            mapped_parent_code = mapped_code
        elif mapped_code in Q11A_LABEL_MAP:
            mapped_parent_code = "quit"
            mapped_quit_code = mapped_code

        has_mapping = bool(mapped_parent_code or mapped_quit_code)
        final_codes = [code for code in base_codes if not (code == "other" and has_mapping)]
        if mapped_parent_code and mapped_parent_code not in final_codes:
            final_codes.append(mapped_parent_code)
        final_quit_codes = list(base_quit_codes)
        if mapped_quit_code and mapped_quit_code not in final_quit_codes:
            final_quit_codes.append(mapped_quit_code)
        for code in final_codes:
            counter[code] += 1
        for code in final_quit_codes:
            quit_counter[code] += 1

        if "other" in base_codes:
            residual_reason = ""
            residual_theme = ""
            if not has_mapping:
                residual_reason = other_text if other_text else "No free-text provided"
                residual_theme = Q11_OTHER_TEXT_THEME_MAP.get(normalized_other, residual_reason)
                residual_counter[residual_theme] += 1
                residual_theme_texts.setdefault(residual_theme, set()).add(residual_reason)
            mapping_rows.append({
                "survey_id": row.survey_id,
                "employment_status": row.q8_employment_status,
                "original_q11_codes": " | ".join(base_codes),
                "q11_other_text": other_text,
                "q11a_quit_reasons": row.q11a_quit_reasons,
                "mapped_existing_reason": (
                    Q11_LABEL_MAP.get(mapped_parent_code, "") if mapped_parent_code and not mapped_quit_code
                    else Q11A_LABEL_MAP.get(mapped_quit_code, "")
                ),
                "final_q11_codes": " | ".join(final_codes),
                "final_q11a_quit_reasons": " | ".join(final_quit_codes),
                "left_under_other": "Yes" if not has_mapping else "No",
                "residual_other_reason": residual_reason,
                "residual_other_theme": residual_theme,
            })

    rows = []

    order = list(Q11_LABEL_MAP.keys()) + [k for k in counter if k not in Q11_LABEL_MAP]
    parent_rows = []
    for code in order:
        n = counter.get(code, 0)
        if n >= 2:
            parent_rows.append({
                "Reason": Q11_LABEL_MAP.get(code, code),
                "Total Youth": n,
                pct_col: pct_str(n, total_left),
            })

    parent_rows.sort(key=lambda r: (-r["Total Youth"], r["Reason"]))

    quit_rows = []
    order_a = list(Q11A_LABEL_MAP.keys()) + [k for k in quit_counter if k not in Q11A_LABEL_MAP]
    for code in order_a:
        n = quit_counter.get(code, 0)
        if n >= 2:
            quit_rows.append({
                "Reason": "    " + Q11A_LABEL_MAP.get(code, code),
                "Total Youth": n,
                pct_col: pct_str(n, total_left),
            })
    quit_rows.sort(key=lambda r: (-r["Total Youth"], r["Reason"].strip()))

    for row in parent_rows:
        rows.append(row)
        if row["Reason"] == Q11_LABEL_MAP.get("quit", "Quit"):
            rows.extend(quit_rows)

    mapping_df = pd.DataFrame(mapping_rows)
    repeated_residual_df = pd.DataFrame([
        {
            "Residual other theme": reason,
            "Total Youth": count,
            "Source texts": " | ".join(sorted(residual_theme_texts.get(reason, set()))),
        }
        for reason, count in residual_counter.items() if count > 1
    ])
    if not repeated_residual_df.empty:
        repeated_residual_df = repeated_residual_df.sort_values(
            by=["Total Youth", "Residual other theme"],
            ascending=[False, True],
        ).reset_index(drop=True)

    return pd.DataFrame(rows), mapping_df, repeated_residual_df


def sec13_left_job_reference(df):
    summary_df, mapping_df, repeated_residual_df = build_sec13_left_job_outputs(df)
    return {
        "Updated Q11 left-job table": summary_df,
        "Q11 other mapping audit": mapping_df,
        "Repeated residual other reasons": repeated_residual_df,
    }


def sec14_transport(df):
    # Table A: Driver's license x age
    valid_q6 = df[df["q6_drivers_license"] != ""]
    n_per_age = {age: (valid_q6["_age"] == age).sum() for age in AGE_ORDER}
    n_total = len(valid_q6)

    rows_a = [{"Driver's License": "Number of Youth",
               **{age: n_per_age[age] for age in AGE_ORDER}, "Total": n_total}]
    for code in Q6_ORDER:
        sub = valid_q6[valid_q6["q6_drivers_license"] == code]
        row_data = {"Driver's License": Q6_LABEL_MAP[code]}
        for age in AGE_ORDER:
            row_data[age] = pct_str((sub["_age"] == age).sum(), n_per_age[age])
        row_data["Total"] = pct_str(len(sub), n_total)
        rows_a.append(row_data)
    rows_a.append({"Driver's License": "Total",
                   **{age: "100%" if n_per_age[age] > 0 else "" for age in AGE_ORDER},
                   "Total": "100%"})
    table_a = pd.DataFrame(rows_a)

    # Table B: Vehicle access x age (license holders)
    licensed = df[df["q6_drivers_license"] == "yes"]
    rows_b = []
    for code in Q6A_ORDER:
        sub = licensed[licensed["q6a_vehicle_access"] == code]
        if sub.empty:
            continue
        row_data = {"Vehicle Access (license holders)": Q6A_LABEL_MAP.get(code, code)}
        for age in AGE_ORDER:
            row_data[age] = (sub["_age"] == age).sum()
        row_data["Total"] = len(sub)
        rows_b.append(row_data)
    rows_b.append({"Vehicle Access (license holders)": "Total",
                   **{age: (licensed["_age"] == age).sum() for age in AGE_ORDER},
                   "Total": len(licensed)})
    table_b = pd.DataFrame(rows_b)

    # Table C: Primary transport
    counter = Counter(v.strip() for v in df["q9_primary_transport"] if v.strip())
    total_t = sum(counter.values())

    rows_c = []
    all_codes = Q9_ORDER + [k for k in counter if k not in Q9_ORDER]
    for code in all_codes:
        n = counter.get(code, 0)
        if n == 0:
            continue
        rows_c.append({
            "Primary Transport": Q9_LABEL_MAP.get(code, code),
            "Total": n,
            "Percent": pct_str(n, total_t),
        })
    rows_c.sort(key=lambda r: -r["Total"])
    rows_c.append({"Primary Transport": "Total", "Total": total_t, "Percent": "100%"})
    table_c = pd.DataFrame(rows_c)

    return {
        "Driver's License by Age": table_a,
        "Vehicle Access (licensed)": table_b,
        "Primary Transport": table_c,
    }


def build_sec15_voter_reg_outputs(df):
    voter_ages = ["18-20 years old", "21-23 years old"]
    vdf = df[df["_age"].isin(voter_ages)].copy()
    n_per_age = {age: (vdf[(vdf["_age"] == age) & (vdf["q7_registered_to_vote"] != "")]).shape[0]
                 for age in voter_ages}
    n_total = sum(n_per_age.values())

    rows_a = [{"Status": "n =", **{age: n_per_age[age] for age in voter_ages}, "Total": n_total}]
    for code, label in [("yes", "Registered to Vote"), ("no", "Not Registered")]:
        row_data = {"Status": label}
        for age in voter_ages:
            n = (vdf[(vdf["_age"] == age) & (vdf["q7_registered_to_vote"] == code)]).shape[0]
            row_data[age] = pct_str(n, n_per_age[age])
        total_n = (vdf["q7_registered_to_vote"] == code).sum()
        row_data["Total"] = pct_str(total_n, n_total)
        rows_a.append(row_data)
    table_a = pd.DataFrame(rows_a)

    not_reg = vdf[vdf["q7_registered_to_vote"] == "no"]
    counter_by_age = {age: Counter() for age in voter_ages}
    counter_all = Counter()
    mapping_rows = []
    residual_counter = Counter()
    residual_counter_by_age = {age: Counter() for age in voter_ages}
    residual_theme_texts = {}
    for _, row in not_reg.iterrows():
        age_bucket = row["_age"]
        base_codes = split_pipe(row["q7a_not_registered_reasons"])
        other_text = (row["q7a_other_text"] or "").strip()
        normalized_other = " ".join(other_text.lower().split())
        mapped_code = Q7A_OTHER_TEXT_MAP.get(normalized_other, "")
        has_mapping = bool(mapped_code)

        final_codes = [code for code in base_codes if not (code == "other" and has_mapping)]
        if mapped_code and mapped_code not in final_codes:
            final_codes.append(mapped_code)

        for code in final_codes:
            counter_by_age[age_bucket][code] += 1
            counter_all[code] += 1

        if "other" in base_codes:
            residual_reason = ""
            residual_theme = ""
            if not has_mapping:
                residual_reason = other_text if other_text else "No free-text provided"
                residual_theme = Q7A_OTHER_TEXT_THEME_MAP.get(normalized_other, residual_reason)
                residual_counter[residual_theme] += 1
                residual_counter_by_age[age_bucket][residual_theme] += 1
                residual_theme_texts.setdefault(residual_theme, set()).add(residual_reason)
            mapping_rows.append({
                "survey_id": row["survey_id"],
                "age_range": age_bucket,
                "original_q7a_codes": " | ".join(base_codes),
                "q7a_other_text": other_text,
                "mapped_existing_reason": Q7A_LABEL_MAP.get(mapped_code, ""),
                "final_q7a_codes": " | ".join(final_codes),
                "left_under_other": "Yes" if not has_mapping else "No",
                "residual_other_reason": residual_reason,
                "residual_other_theme": residual_theme,
            })

    rows_b = []
    order = list(Q7A_LABEL_MAP.keys()) + [k for k in counter_all if k not in Q7A_LABEL_MAP]
    base_rows = []
    for code in order:
        n = counter_all.get(code, 0)
        if n == 0:
            continue
        row_data = {"Reason Not Registered": Q7A_LABEL_MAP.get(code, code)}
        for age in voter_ages:
            row_data[age] = counter_by_age[age].get(code, 0)
        row_data["Total"] = n
        base_rows.append((code, row_data))
    base_rows.sort(key=lambda item: -item[1]["Total"])

    display_residual_themes = [
        theme for theme, count in residual_counter.items()
        if count > 0 and theme != "No free-text provided"
    ]
    display_residual_themes.sort(key=lambda theme: (-residual_counter[theme], theme))

    for code, row_data in base_rows:
        rows_b.append(row_data)
        if code == "other":
            for theme in display_residual_themes:
                child_row = {"Reason Not Registered": "    " + theme}
                for age in voter_ages:
                    child_row[age] = residual_counter_by_age[age].get(theme, 0)
                child_row["Total"] = residual_counter.get(theme, 0)
                rows_b.append(child_row)
    table_b = pd.DataFrame(rows_b)

    eligible = vdf[vdf["q7_registered_to_vote"].isin(["yes", "no"])].copy()
    trans_tokens = {
        "Non-binary", "Gender Nonconforming", "Transgender Male",
        "Transgender Female", "Genderqueer", "Two-Spirit",
    }
    eligible["_gender_bucket"] = eligible["gender"].astype(str).apply(
        lambda raw: (
            "Trans, Non-binary" if any(token in trans_tokens for token in split_pipe(raw)) else (
                "Female" if split_pipe(raw) == ["Female"] else (
                    "Male" if split_pipe(raw) == ["Male"] else ""
                )
            )
        )
    )
    lgbtq_values = {
        "Asexual", "Bisexual", "Demisexual", "Gay or Lesbian",
        "Same Gender Loving", "Mostly heterosexual", "Pansexual", "Queer",
    }
    eligible["_orient_bucket"] = eligible["sexual_orientation"].astype(str).apply(
        lambda raw: (
            "LGBTQ+" if any(token in lgbtq_values for token in split_pipe(raw)) else (
                "Heterosexual/Straight" if split_pipe(raw) in (["Heterosexual/Straight"], ["Heterosexual"]) else ""
            )
        )
    )

    gender_rows = []
    for gender in ["Female", "Male", "Trans, Non-binary"]:
        subset = eligible[eligible["_gender_bucket"] == gender]
        n_total_gender = len(subset)
        n_registered = int((subset["q7_registered_to_vote"] == "yes").sum())
        if n_total_gender == 0:
            continue
        gender_rows.append({
            "Gender": gender,
            "Registered": n_registered,
            "Total with Voting Response": n_total_gender,
            "Percent Registered": pct_str(n_registered, n_total_gender),
        })
    table_c = pd.DataFrame(gender_rows)

    orientation_rows = []
    for orient in ["Heterosexual/Straight", "LGBTQ+"]:
        subset = eligible[eligible["_orient_bucket"] == orient]
        n_total_orient = len(subset)
        if n_total_orient == 0:
            continue
        n_registered = int((subset["q7_registered_to_vote"] == "yes").sum())
        orientation_rows.append({
            "Sexual Orientation": orient,
            "Registered": n_registered,
            "Total with Voting Response": n_total_orient,
            "Percent Registered": pct_str(n_registered, n_total_orient),
        })
    table_e = pd.DataFrame(orientation_rows)

    reason_rows = eligible[eligible["q7a_not_registered_reasons"].str.strip() != ""].copy()
    dont_know_how_n = sum(
        "dont_know_how" in split_pipe(val) for val in reason_rows["q7a_not_registered_reasons"]
    )
    inconsistent_n = int(((eligible["q7_registered_to_vote"] == "yes") & (eligible["q7a_not_registered_reasons"].str.strip() != "")).sum())
    table_d = pd.DataFrame([
        {
            "Metric": "Youth providing a reason for not being registered",
            "Count": len(reason_rows),
            "Denominator": len(eligible),
            "Percent": pct_str(len(reason_rows), len(eligible)),
        },
        {
            "Metric": "Youth reporting they do not know how to register",
            "Count": dont_know_how_n,
            "Denominator": len(reason_rows),
            "Percent": pct_str(dont_know_how_n, len(reason_rows)),
        },
        {
            "Metric": "Youth marked registered and also selected a not-registered reason",
            "Count": inconsistent_n,
            "Denominator": len(eligible),
            "Percent": pct_str(inconsistent_n, len(eligible)),
        },
    ])

    mapping_df = pd.DataFrame(mapping_rows)
    repeated_residual_df = pd.DataFrame([
        {
            "Residual other theme": reason,
            "Total Youth": count,
            "Source texts": " | ".join(sorted(residual_theme_texts.get(reason, set()))),
        }
        for reason, count in residual_counter.items() if count > 1
    ])
    if not repeated_residual_df.empty:
        repeated_residual_df = repeated_residual_df.sort_values(
            by=["Total Youth", "Residual other theme"],
            ascending=[False, True],
        ).reset_index(drop=True)

    return {
        "Voter Registration by Age": table_a,
        "Not Registered Reasons by Age": table_b,
        "Registration by Gender (18-23)": table_c,
        "Registration by Sexual Orientation (18-23)": table_e,
        "Voting Narrative Support": table_d,
        "Q7a other mapping audit": mapping_df,
        "Repeated residual other reasons": repeated_residual_df,
    }


def sec15_voter_reg(df):
    return build_sec15_voter_reg_outputs(df)


def sec15_voter_reg_reference(df):
    outputs = build_sec15_voter_reg_outputs(df)
    return {
        "Q7a other mapping audit": outputs["Q7a other mapping audit"],
        "Repeated residual other reasons": outputs["Repeated residual other reasons"],
    }


def sec16_visit(df):
    rows_a = []
    for code in Q15_ORDER:
        row_data = {"Visit Frequency": Q15_LABEL_MAP[code]}
        for age in AGE_ORDER:
            row_data[age] = int(((df["_age"] == age) & (df["q15_visit_frequency"] == code)).sum())
        row_data["Total"] = int((df["q15_visit_frequency"] == code).sum())
        rows_a.append(row_data)
    total_row = {"Visit Frequency": "Total"}
    for age in AGE_ORDER:
        total_row[age] = int(((df["_age"] == age) & (df["q15_visit_frequency"] != "")).sum())
    total_row["Total"] = int((df["q15_visit_frequency"] != "").sum())
    rows_a.append(total_row)
    table_a = pd.DataFrame(rows_a)

    frequent = df[df["q15_visit_frequency"].isin(FREQUENT_VISIT)]
    ct_b = crosstab_age(frequent, "q15a_visit_reasons", Q15A_LABEL_MAP)
    if not ct_b.empty:
        ct_b = ct_b.rename(columns={"Label": "Visit Reason (frequent visitors)"})
    else:
        ct_b = None

    reason_combo_order = [
        "see_coach_staff",
        "food",
        "scheduled_activity",
        "work_on_goals",
        "learn_skills",
        "safe_place",
        "socialize",
        "escape_problems",
        "service_providers",
        "laundry_shower",
        "health_counseling",
        "computers",
    ]
    reason_combo_labels = {
        "see_coach_staff": "See my Youth Coach",
        "food": "Eat food",
        "scheduled_activity": "Scheduled activity",
        "work_on_goals": "Work toward my goals",
        "learn_skills": "Learn new things, skills",
        "safe_place": "Safe place",
        "socialize": "Socialize/see friends",
        "escape_problems": "Escape problems elsewhere",
        "service_providers": "Service providers",
        "laundry_shower": "Do laundry, shower, etc.",
        "health_counseling": "Health care, counseling",
        "computers": "Access to computers",
    }
    combo_visit_groups = [
        ("every_week", "Every week"),
        ("1_3_times_per_month", "1 to 3 times per month"),
    ]

    combo_rows = []
    combo_base = df[df["q15_visit_frequency"].isin([code for code, _ in combo_visit_groups])]
    combo_total_n = len(combo_base)
    for reason_code in reason_combo_order:
        row = {"Reason": reason_combo_labels[reason_code]}
        total_count = 0
        for visit_code, visit_label in combo_visit_groups:
            visit_df = combo_base[combo_base["q15_visit_frequency"] == visit_code]
            visit_total = len(visit_df)
            reason_count = int(sum(reason_code in split_pipe(val) for val in visit_df["q15a_visit_reasons"]))
            row[f"{visit_label} Count"] = reason_count
            row[f"{visit_label} Percent"] = pct_str(reason_count, visit_total)
            total_count += reason_count
        row["Total Count"] = total_count
        row["Total Percent"] = pct_str(total_count, combo_total_n)
        combo_rows.append(row)
    table_b2 = pd.DataFrame(combo_rows)
    if not table_b2.empty:
        table_b2 = table_b2.sort_values(by=["Total Count", "Reason"], ascending=[False, True]).reset_index(drop=True)

    infreq = df[df["q15_visit_frequency"].isin(INFREQUENT_VISIT)].copy()
    barrier_counter_by_age = {age: Counter() for age in AGE_ORDER}
    barrier_counter_all = Counter()
    barrier_mapping_rows = []
    residual_barrier_counter = Counter()
    residual_barrier_counter_by_age = {age: Counter() for age in AGE_ORDER}
    residual_barrier_texts = {}
    processed_barrier_codes = []

    for _, row in infreq.iterrows():
        age_bucket = row["_age"]
        base_codes = split_pipe(row["q15b_visit_barriers"])
        other_text = (row.get("q15b_other_text") or "").strip()
        normalized_other = " ".join(other_text.lower().split())
        mapped_code = Q15B_OTHER_TEXT_MAP.get(normalized_other, "")
        has_mapping = bool(mapped_code)

        final_codes = [code for code in base_codes if not (code == "other" and has_mapping)]
        if mapped_code and mapped_code not in final_codes:
            final_codes.append(mapped_code)

        for code in final_codes:
            barrier_counter_by_age[age_bucket][code] += 1
            barrier_counter_all[code] += 1
        processed_barrier_codes.append(" | ".join(final_codes))

        if "other" in base_codes:
            residual_reason = ""
            residual_theme = ""
            if not has_mapping:
                residual_reason = other_text if other_text else "No free-text provided"
                residual_theme = Q15B_OTHER_TEXT_THEME_MAP.get(normalized_other, residual_reason)
                residual_barrier_counter[residual_theme] += 1
                residual_barrier_counter_by_age[age_bucket][residual_theme] += 1
                residual_barrier_texts.setdefault(residual_theme, set()).add(residual_reason)
            barrier_mapping_rows.append({
                "survey_id": row["survey_id"],
                "age_range": age_bucket,
                "original_q15b_codes": " | ".join(base_codes),
                "q15b_other_text": other_text,
                "mapped_existing_barrier": Q15B_LABEL_MAP.get(mapped_code, ""),
                "final_q15b_codes": " | ".join(final_codes),
                "left_under_other": "Yes" if not has_mapping else "No",
                "residual_other_reason": residual_reason,
                "residual_other_theme": residual_theme,
            })

    infreq["_q15b_processed"] = processed_barrier_codes

    barrier_rows = []
    order = list(Q15B_LABEL_MAP.keys()) + [k for k in barrier_counter_all if k not in Q15B_LABEL_MAP]
    base_barrier_rows = []
    for code in order:
        total_count = barrier_counter_all.get(code, 0)
        if total_count == 0:
            continue
        row_data = {"Visit Barrier (infrequent/never visitors)": Q15B_LABEL_MAP.get(code, code)}
        for age in AGE_ORDER:
            row_data[age] = barrier_counter_by_age[age].get(code, 0)
        row_data["Total"] = total_count
        base_barrier_rows.append((code, row_data))
    base_barrier_rows.sort(key=lambda item: (-item[1]["Total"], item[1]["Visit Barrier (infrequent/never visitors)"]))

    display_residual_themes = [
        theme for theme, count in residual_barrier_counter.items()
        if count >= 2 and theme != "No free-text provided"
    ]
    display_residual_themes.sort(key=lambda theme: (-residual_barrier_counter[theme], theme))

    total_row = {"Visit Barrier (infrequent/never visitors)": "Total respondents"}
    for age in AGE_ORDER:
        total_row[age] = int((infreq["_age"] == age).sum())
    total_row["Total"] = len(infreq)
    barrier_rows.append(total_row)

    for code, row_data in base_barrier_rows:
        barrier_rows.append(row_data)
        if code == "other":
            for theme in display_residual_themes:
                child_row = {"Visit Barrier (infrequent/never visitors)": "    " + theme}
                for age in AGE_ORDER:
                    child_row[age] = residual_barrier_counter_by_age[age].get(theme, 0)
                child_row["Total"] = residual_barrier_counter.get(theme, 0)
                barrier_rows.append(child_row)
    ct_c = pd.DataFrame(barrier_rows) if barrier_rows else None

    total_visit_n = int((df["q15_visit_frequency"] != "").sum())
    weekly_n = int((df["q15_visit_frequency"] == "every_week").sum())
    monthly_n = int((df["q15_visit_frequency"] == "1_3_times_per_month").sum())
    less_monthly_n = int((df["q15_visit_frequency"] == "less_than_once_per_month").sum())
    never_n = int((df["q15_visit_frequency"] == "never").sum())
    at_least_monthly_n = weekly_n + monthly_n

    visit_support_rows = [
        {"Metric": "Attending at least monthly", "Count": at_least_monthly_n, "Denominator": total_visit_n, "Percent": pct_str(at_least_monthly_n, total_visit_n)},
        {"Metric": "Attending every week", "Count": weekly_n, "Denominator": total_visit_n, "Percent": pct_str(weekly_n, total_visit_n)},
        {"Metric": "Attending 1-3 times per month", "Count": monthly_n, "Denominator": total_visit_n, "Percent": pct_str(monthly_n, total_visit_n)},
        {"Metric": "Attending less than once per month", "Count": less_monthly_n, "Denominator": total_visit_n, "Percent": pct_str(less_monthly_n, total_visit_n)},
        {"Metric": "Never visiting downtown Zone", "Count": never_n, "Denominator": total_visit_n, "Percent": pct_str(never_n, total_visit_n)},
    ]
    for age in ["16-17 years old", "18-20 years old", "21-23 years old"]:
        age_total = int(((df["_age"] == age) & (df["q15_visit_frequency"] != "")).sum())
        age_weekly = int(((df["_age"] == age) & (df["q15_visit_frequency"] == "every_week")).sum())
        visit_support_rows.append({
            "Metric": f"{age} attending every week",
            "Count": age_weekly,
            "Denominator": age_total,
            "Percent": pct_str(age_weekly, age_total),
        })
    table_d = pd.DataFrame(visit_support_rows)

    frequent_total = len(frequent)
    infreq_total = len(infreq)
    reason_counter = Counter()
    barrier_counter = Counter()
    for _, row in frequent.iterrows():
        reason_counter.update(split_pipe(row["q15a_visit_reasons"]))
    for _, row in infreq.iterrows():
        barrier_counter.update(split_pipe(row["_q15b_processed"]))

    frequent_18_20 = frequent[frequent["_age"] == "18-20 years old"]
    frequent_21_23 = frequent[frequent["_age"] == "21-23 years old"]
    infreq_18_20 = infreq[infreq["_age"] == "18-20 years old"]

    def _count_multi(sub_df, field, code):
        return int(sum(code in split_pipe(val) for val in sub_df[field]))

    table_e = pd.DataFrame([
        {"Metric": "Frequent visitors coming to see coach/staff", "Count": reason_counter["see_coach_staff"], "Denominator": frequent_total, "Percent": pct_str(reason_counter["see_coach_staff"], frequent_total)},
        {"Metric": "Frequent visitors coming for food", "Count": reason_counter["food"], "Denominator": frequent_total, "Percent": pct_str(reason_counter["food"], frequent_total)},
        {"Metric": "Frequent visitors coming for scheduled activities", "Count": reason_counter["scheduled_activity"], "Denominator": frequent_total, "Percent": pct_str(reason_counter["scheduled_activity"], frequent_total)},
        {"Metric": "Frequent visitors coming to work toward goals", "Count": reason_counter["work_on_goals"], "Denominator": frequent_total, "Percent": pct_str(reason_counter["work_on_goals"], frequent_total)},
        {"Metric": "Frequent visitors 21-23 coming for a safe place", "Count": _count_multi(frequent_21_23, "q15a_visit_reasons", "safe_place"), "Denominator": len(frequent_21_23), "Percent": pct_str(_count_multi(frequent_21_23, "q15a_visit_reasons", "safe_place"), len(frequent_21_23))},
        {"Metric": "Frequent visitors 21-23 coming to escape problems/issues", "Count": _count_multi(frequent_21_23, "q15a_visit_reasons", "escape_problems"), "Denominator": len(frequent_21_23), "Percent": pct_str(_count_multi(frequent_21_23, "q15a_visit_reasons", "escape_problems"), len(frequent_21_23))},
        {"Metric": "Frequent visitors 18-20 coming for a safe place", "Count": _count_multi(frequent_18_20, "q15a_visit_reasons", "safe_place"), "Denominator": len(frequent_18_20), "Percent": pct_str(_count_multi(frequent_18_20, "q15a_visit_reasons", "safe_place"), len(frequent_18_20))},
        {"Metric": "Frequent visitors 18-20 coming to escape problems/issues", "Count": _count_multi(frequent_18_20, "q15a_visit_reasons", "escape_problems"), "Denominator": len(frequent_18_20), "Percent": pct_str(_count_multi(frequent_18_20, "q15a_visit_reasons", "escape_problems"), len(frequent_18_20))},
        {"Metric": "Infrequent/never visitors wanting more activities", "Count": barrier_counter["better_activities"], "Denominator": infreq_total, "Percent": pct_str(barrier_counter["better_activities"], infreq_total)},
        {"Metric": "Infrequent/never visitors wanting coach invitation", "Count": barrier_counter["coach_invitation"], "Denominator": infreq_total, "Percent": pct_str(barrier_counter["coach_invitation"], infreq_total)},
        {"Metric": "Infrequent/never visitors wanting more activity information", "Count": barrier_counter["more_info"], "Denominator": infreq_total, "Percent": pct_str(barrier_counter["more_info"], infreq_total)},
        {"Metric": "Infrequent/never visitors 18-20 wanting coach invitation", "Count": _count_multi(infreq_18_20, "q15b_visit_barriers", "coach_invitation"), "Denominator": len(infreq_18_20), "Percent": pct_str(_count_multi(infreq_18_20, "q15b_visit_barriers", "coach_invitation"), len(infreq_18_20))},
    ])

    valid_focus = df[(df["q15_visit_frequency"] != "") & (df["q16_stay_focused"] != "")]
    focus_rows = []
    for visit_code in Q15_ORDER:
        visit_df = valid_focus[valid_focus["q15_visit_frequency"] == visit_code]
        visit_total = len(visit_df)
        row = {"Visit Frequency": Q15_LABEL_MAP[visit_code], "Respondents": visit_total}
        for focus_code in Q16_ORDER:
            focus_count = int((visit_df["q16_stay_focused"] == focus_code).sum())
            row[f"{Q16_LABEL_MAP[focus_code]} Count"] = focus_count
            row[Q16_LABEL_MAP[focus_code]] = pct_str(focus_count, visit_total)
        focus_rows.append(row)
    table_f = pd.DataFrame(focus_rows)

    at_least_monthly = valid_focus[valid_focus["q15_visit_frequency"].isin(FREQUENT_VISIT)]
    less_than_monthly = valid_focus[valid_focus["q15_visit_frequency"].isin(INFREQUENT_VISIT)]
    total_q16 = len(valid_focus)
    agree_total = int((valid_focus["q16_stay_focused"] == "agree").sum())
    agree_some_at_least_monthly = int(at_least_monthly["q16_stay_focused"].isin(["agree", "somewhat_agree"]).sum())
    agree_some_less_than_monthly = int(less_than_monthly["q16_stay_focused"].isin(["agree", "somewhat_agree"]).sum())
    unsure_at_least_monthly = int((at_least_monthly["q16_stay_focused"] == "unsure").sum())
    unsure_less_than_monthly = int((less_than_monthly["q16_stay_focused"] == "unsure").sum())

    table_g = pd.DataFrame([
        {"Metric": "Overall fully agree Youth Zone helps stay focused", "Count": agree_total, "Denominator": total_q16, "Percent": pct_str(agree_total, total_q16)},
        {"Metric": "At least monthly visitors agree or somewhat agree Youth Zone helps stay focused", "Count": agree_some_at_least_monthly, "Denominator": len(at_least_monthly), "Percent": pct_str(agree_some_at_least_monthly, len(at_least_monthly))},
        {"Metric": "Less than monthly or never visitors agree or somewhat agree Youth Zone helps stay focused", "Count": agree_some_less_than_monthly, "Denominator": len(less_than_monthly), "Percent": pct_str(agree_some_less_than_monthly, len(less_than_monthly))},
        {"Metric": "At least monthly visitors unsure about goals", "Count": unsure_at_least_monthly, "Denominator": len(at_least_monthly), "Percent": pct_str(unsure_at_least_monthly, len(at_least_monthly))},
        {"Metric": "Less than monthly or never visitors unsure about goals", "Count": unsure_less_than_monthly, "Denominator": len(less_than_monthly), "Percent": pct_str(unsure_less_than_monthly, len(less_than_monthly))},
    ])

    barrier_mapping_df = pd.DataFrame(barrier_mapping_rows)
    repeated_residual_barriers_df = pd.DataFrame([
        {
            "Residual other barrier theme": theme,
            "Total Youth": count,
            "Source texts": " | ".join(sorted(residual_barrier_texts.get(theme, set()))),
        }
        for theme, count in residual_barrier_counter.items() if count > 1
    ])
    if not repeated_residual_barriers_df.empty:
        repeated_residual_barriers_df = repeated_residual_barriers_df.sort_values(
            by=["Total Youth", "Residual other barrier theme"],
            ascending=[False, True],
        ).reset_index(drop=True)

    return {
        "Visit Frequency by Age": table_a,
        "Visit Reasons (frequent)": ct_b,
        "Visit Reasons by Visit Frequency": table_b2,
        "Visit Barriers (infrequent)": ct_c,
        "Visit Frequency Narrative Support": table_d,
        "Visit Reason and Barrier Narrative Support": table_e,
        "Stay Focused by Visit Frequency (Q16)": table_f,
        "Stay Focused Narrative Support": table_g,
        "Q15b other mapping audit": barrier_mapping_df,
        "Repeated residual other barriers": repeated_residual_barriers_df,
    }


def sec17_program_impact(df):
    ct_a = crosstab_age(df, "q17_program_helped", Q17_LABEL_MAP)
    if not ct_a.empty:
        ct_a = ct_a.rename(columns={"Label": "Program Helped With"})
    else:
        ct_a = None

    q17_age_groups = ["16-17 years old", "18-20 years old", "21-23 years old"]
    q17_chart_label_map = {
        "future": "Think about my future",
        "decision_making": "Make good decisions",
        "handle_problems": "Figure out how to handle problems",
        "vital_documents": "Obtain vital documents",
        "positive_relationships": "Establish positive relationships",
        "drivers_license": "Get my driver's license",
        "health_counseling": "Access health care and/or counseling",
        "housing": "Find or maintain housing",
        "job": "Get or keep a job",
        "education": "Finish or further my education",
        "everyday_skills": "Learn everyday skills",
        "parenting": "Improve parenting skills",
        "something_else": "Something else",
    }
    q17_chart_order = [
        "future",
        "decision_making",
        "handle_problems",
        "vital_documents",
        "positive_relationships",
        "drivers_license",
        "health_counseling",
        "housing",
        "job",
        "education",
        "everyday_skills",
        "parenting",
        "something_else",
    ]
    q17_known_age = df[(df["q17_program_helped"].astype(str).str.strip() != "") & (df["_age"].isin(q17_age_groups))].copy()
    q17_n_by_age = {age: len(q17_known_age[q17_known_age["_age"] == age]) for age in q17_age_groups}
    q17_total_known = sum(q17_n_by_age.values())
    q17_chart_rows = []
    for code in q17_chart_order:
        row = {"Program Helped With": q17_chart_label_map[code]}
        total_count = 0
        for age in q17_age_groups:
            age_df = q17_known_age[q17_known_age["_age"] == age]
            count = int(sum(code in split_pipe(val) for val in age_df["q17_program_helped"]))
            row[f"{age} Count"] = count
            row[age] = pct_str(count, q17_n_by_age[age])
            total_count += count
        row["Total Count"] = total_count
        row["Total"] = pct_str(total_count, q17_total_known)
        q17_chart_rows.append(row)
    table_d = pd.DataFrame(q17_chart_rows)
    if not table_d.empty:
        table_d = table_d.sort_values(by=["Total Count", "Program Helped With"], ascending=[False, True]).reset_index(drop=True)

    q16 = df["q16_stay_focused"].replace("", pd.NA).dropna().value_counts()
    total_q16 = q16.sum()
    table_b = pd.DataFrame([{
        "Q16 - Coach/Zone helps stay focused": {**Q16_LABEL_MAP, "unsure": "Unsure"}.get(k, k),
        "Count": v, "Percent": pct_str(v, total_q16),
    } for k, v in q16.items()])

    Q21_LABEL_MAP = {"agree": "Agree", "somewhat": "Somewhat agree",
                     "disagree": "Disagree", "unsure": "Unsure"}
    q21 = df["q21_gained_independence"].replace("", pd.NA).dropna().value_counts()
    total_q21 = q21.sum()
    table_c = pd.DataFrame([{
        "Q21 - Gained independence": Q21_LABEL_MAP.get(k, k),
        "Count": v, "Percent": pct_str(v, total_q21),
    } for k, v in q21.items()])

    return {
        "Program Helped With (Q17) by Age": ct_a,
        "Program Helped With (Q17) Chart Reference": table_d,
        "Stay Focused (Q16)": table_b,
        "Gained Independence (Q21)": table_c,
    }


def sec18_respect(df):
    RESP_ORDER = ["never", "rarely", "sometimes", "often", "all_the_time"]
    RESP_LABELS = {"never": "Never", "rarely": "Rarely", "sometimes": "Sometimes",
                   "often": "Often", "all_the_time": "All the time"}
    TOP2 = {"often", "all_the_time"}

    rows = []
    for field, label in [
        ("q18_staff_respect", "Staff treat me with respect and acceptance"),
        ("q19_peer_respect",  "Peers treat me with respect and acceptance"),
    ]:
        col = df[field].replace("", pd.NA).dropna()
        n_t2 = col.isin(TOP2).sum()
        row_data = {
            "Statement": label,
            "n": len(col),
            "% Often or All the Time": pct_str(n_t2, len(col)),
        }
        for code in RESP_ORDER:
            row_data[RESP_LABELS[code]] = (col == code).sum()
        rows.append(row_data)
    table_a = pd.DataFrame(rows)

    rare_never_any = df[df["q18_staff_respect"].isin(["rarely", "never"]) | df["q19_peer_respect"].isin(["rarely", "never"])].copy()
    known_age_rare_never = rare_never_any[rare_never_any["_age"].isin(["16-17 years old", "18-20 years old", "21-23 years old"])].copy()

    demo_rows = []
    for age in ["16-17 years old", "18-20 years old", "21-23 years old", "Unknown"]:
        count = int((rare_never_any["_age"] == age).sum())
        demo_rows.append({
            "Age Group": age,
            "Count": count,
            "Percent of Rarely/Never Group": pct_str(count, len(rare_never_any)),
        })
    demo_rows.append({
        "Age Group": "Total",
        "Count": len(rare_never_any),
        "Percent of Rarely/Never Group": "100%" if len(rare_never_any) else "",
    })
    table_b = pd.DataFrame(demo_rows)

    majority_21_23 = int((known_age_rare_never["_age"] == "21-23 years old").sum())
    total_known = len(known_age_rare_never)
    under18 = int((known_age_rare_never["_age"] == "16-17 years old").sum())
    staff_rare_never = df[df["q18_staff_respect"].isin(["rarely", "never"])].copy()
    peer_rare_never = df[df["q19_peer_respect"].isin(["rarely", "never"])].copy()
    respect_response_den = int(((df["q18_staff_respect"] != "") | (df["q19_peer_respect"] != "")).sum())
    table_c = pd.DataFrame([
        {"Metric": "Youth reporting rarely/never by staff or peers", "Count": len(rare_never_any), "Denominator": respect_response_den, "Percent": pct_str(len(rare_never_any), respect_response_den)},
        {"Metric": "Known-age rarely/never group ages 21-23", "Count": majority_21_23, "Denominator": total_known, "Percent": pct_str(majority_21_23, total_known)},
        {"Metric": "Known-age rarely/never group younger than 18", "Count": under18, "Denominator": total_known, "Percent": pct_str(under18, total_known)},
        {"Metric": "Youth reporting rarely/never by staff", "Count": len(staff_rare_never), "Denominator": int((df["q18_staff_respect"] != "").sum()), "Percent": pct_str(len(staff_rare_never), int((df["q18_staff_respect"] != "").sum()))},
        {"Metric": "Youth reporting rarely/never by peers", "Count": len(peer_rare_never), "Denominator": int((df["q19_peer_respect"] != "").sum()), "Percent": pct_str(len(peer_rare_never), int((df["q19_peer_respect"] != "").sum()))},
    ])
    return {
        "Respect Summary": table_a,
        "Rarely/Never Age Breakdown": table_b,
        "Respect Narrative Support": table_c,
    }


def sec19_environment(df):
    TOP2 = {"4", "5"}
    rows = []
    for field, label in Q20_FIELDS:
        valid = df[field].astype(str)
        valid = valid[valid.str.strip() != ""]
        n_t2 = valid.isin(TOP2).sum()
        rows.append({
            "The Zone is a place where...": label,
            "n": len(valid),
            "% Top-2 Box (4-5)": pct_str(n_t2, len(valid)),
        })
    return pd.DataFrame(rows)


def sec20_banking(df):
    has_account = df["q25_bank_account"].apply(
        lambda v: any(t in split_pipe(v) for t in ["checking", "savings"])
    )

    # Table A: bank account status x age
    rows_a = [{"Account Status": "Number of Youth",
               **{age: int((df["_age"] == age).sum()) for age in AGE_ORDER},
               "Total": len(df), "Percent of Total": ""}]
    rows_a.append({
        "Account Status": "Currently have a bank account",
        **{age: int(has_account[df["_age"] == age].sum()) for age in AGE_ORDER},
        "Total": int(has_account.sum()),
        "Percent of Total": pct_str(has_account.sum(), len(df)),
    })
    for code in Q25_ORDER:
        sub = df["q25_bank_account"].apply(lambda v: code in split_pipe(v))
        n = int(sub.sum())
        if n == 0:
            continue
        rows_a.append({
            "Account Status": Q25_LABEL_MAP[code],
            **{age: int(sub[df["_age"] == age].sum()) for age in AGE_ORDER},
            "Total": n,
            "Percent of Total": pct_str(n, len(df)),
        })
    table_a = pd.DataFrame(rows_a)

    # Table B: money methods x age
    ct_b = crosstab_age(df, "q24_money_methods", Q24_LABEL_MAP)
    if not ct_b.empty:
        ct_b = ct_b.rename(columns={"Label": "Money Method (Q24)"})
        ct_b["Percent of All"] = ct_b["Total"].apply(lambda n: pct_str(n, len(df)))
    else:
        ct_b = None

    # Table C: account usage x age (account holders only)
    has_acc_df = df[has_account]
    ct_c = crosstab_age(has_acc_df, "q26b_account_usage", Q26B_LABEL_MAP)
    if not ct_c.empty:
        ct_c = ct_c.rename(columns={"Label": "Account Usage (Q26b)"})
        ct_c["% of Account Holders"] = ct_c["Total"].apply(
            lambda n: pct_str(n, len(has_acc_df))
        )
    else:
        ct_c = None

    use_bank_method = df["q24_money_methods"].apply(lambda v: "bank_account" in split_pipe(v))
    use_digital_apps = df["q24_money_methods"].apply(lambda v: "digital_apps" in split_pipe(v))
    has_acc_no_bank_method = has_acc_df[~use_bank_method[has_acc_df.index]].copy()
    bills_21_23 = int(sum(
        "paying_bills" in split_pipe(v)
        for v in has_acc_df.loc[has_acc_df["_age"] == "21-23 years old", "q26b_account_usage"]
    ))
    acc_21_23 = int((has_acc_df["_age"] == "21-23 years old").sum())
    digital_among_no_bank_method = int(use_digital_apps[has_acc_no_bank_method.index].sum())
    table_d = pd.DataFrame([
        {
            "Metric": "Account holders ages 21-23 using account to pay household bills",
            "Count": bills_21_23,
            "Denominator": acc_21_23,
            "Percent": pct_str(bills_21_23, acc_21_23),
        },
        {
            "Metric": "Account holders not reporting bank account as a money-management method",
            "Count": len(has_acc_no_bank_method),
            "Denominator": len(has_acc_df),
            "Percent": pct_str(len(has_acc_no_bank_method), len(has_acc_df)),
        },
        {
            "Metric": "Those account holders relying on digital apps instead",
            "Count": digital_among_no_bank_method,
            "Denominator": len(has_acc_no_bank_method),
            "Percent": pct_str(digital_among_no_bank_method, len(has_acc_no_bank_method)),
        },
    ])

    return {
        "Bank Account Status by Age": table_a,
        "Money Methods by Age (Q24)": ct_b,
        "Account Usage by Age (Q26b)": ct_c,
        "Banking Narrative Support": table_d,
    }


def sec21_nps(df):
    nps_raw = df["q22_nps"].replace("", pd.NA).dropna()
    try:
        nps_int = nps_raw.astype(float).astype(int)
    except Exception:
        nps_int = pd.Series([], dtype=int)

    total = len(nps_int)
    promoters = int((nps_int >= 9).sum())
    passives   = int(((nps_int >= 7) & (nps_int <= 8)).sum())
    detractors = int((nps_int <= 6).sum())
    nps_score = round(100 * promoters / total - 100 * detractors / total) if total > 0 else ""

    return pd.DataFrame([
        {"Category": "Promoters (9-10)", "Count": promoters, "Percent": pct_str(promoters, total)},
        {"Category": "Passives (7-8)",   "Count": passives,  "Percent": pct_str(passives, total)},
        {"Category": "Detractors (0-6)", "Count": detractors,"Percent": pct_str(detractors, total)},
        {"Category": "Total Responded",  "Count": total,     "Percent": ""},
        {"Category": "NPS Score",        "Count": nps_score, "Percent": ""},
    ])


def sec22_comments(df):
    comments = df[df["q23_other_comments"].str.strip() != ""][
        ["survey_id", "q23_other_comments"]
    ].copy()
    comments.columns = ["Survey ID", "Comment"]
    comments = comments.reset_index(drop=True)

    trivial_comments = {
        "no", "nope", "none.", "none", "n/a", "na", "not at the moment",
        "not at this time.", "no comment all good", "nah i'm good thxs", "a/k",
        "no.", "not at this time", "not at the moment.",
    }
    substantive_mask = ~comments["Comment"].str.strip().str.lower().isin(trivial_comments)
    substantive_n = int(substantive_mask.sum())
    summary = pd.DataFrame([
        {"Metric": "Total non-blank comments", "Count": len(comments)},
        {"Metric": "Substantive comments used for narrative review", "Count": substantive_n},
        {"Metric": "Brief no/none comments", "Count": int(len(comments) - substantive_n)},
    ])

    return {
        "Comment Summary": summary,
        "Comment Listing": comments,
    }


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

TITLE_FONT  = Font(bold=True, color="FFFFFF", size=12)
TITLE_FILL  = PatternFill("solid", fgColor="2F5496")
SUB_FONT    = Font(bold=True, italic=True)
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")


def write_section(ws, title, tables):
    """Write title then one or more DataFrames to worksheet."""
    if isinstance(tables, pd.DataFrame):
        tables = [("", tables)]
    elif isinstance(tables, dict):
        tables = list(tables.items())

    row = 1
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    row += 1

    for subtitle, df in tables:
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        if subtitle:
            ws.cell(row=row, column=1, value=subtitle).font = SUB_FONT
            row += 1
        for col_idx, col_name in enumerate(df.columns, 1):
            c = ws.cell(row=row, column=col_idx, value=str(col_name))
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        row += 1
        for _, data_row in df.iterrows():
            for col_idx, val in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=val)
            row += 1
        row += 1  # blank row between tables


def autofit_columns(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 55)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

SECTIONS = [
    ("01_age",            "1. Age Distribution",                         sec01_age),
    ("02_gender_orient",  "2. Gender x Sexual Orientation",              sec02_gender_orient),
    ("03_race_once",      "3. Race/Ethnicity (Counted Once)",            sec03_race_once),
    ("04_race_multi",     "4. Race/Ethnicity (Counted Multiple Times)",  sec04_race_multi),
    ("05_q1",             "5. Coach Satisfaction (Q1)",                  sec05_q1),
    ("06_communication",  "6. Communication Frequency & Level (Q2/Q3)", sec06_communication),
    ("07_housing",        "7. Housing Stability (Q12 x Q13)",            sec07_housing),
    ("08_housing_reasons",    "8. Housing Instability Reasons (Q14 x Age)",  sec08_housing_reasons),
    ("08_housing_reasons_ref", "8A. Housing Reasons Mapping Reference",       sec08_housing_reasons_reference),
    ("09_education",           "9. Education (Q5 + Q5a)",                    sec09_education),
    ("10_employment",     "10. Employment Status (Q8 x Age)",            sec10_employment),
    ("11_job_tenure",     "11. Job Tenure (Q8a x Full/Part-time)",       sec11_job_tenure),
    ("12_job_barriers",   "12. Job Barriers (Q10)",                      sec12_job_barriers),
    ("12_job_barriers_ref", "12A. Job Barriers Mapping Reference",       sec12_job_barriers_reference),
    ("13_left_job",       "13. Reasons Left Job (Q11 + Q11a)",           sec13_left_job),
    ("13_left_job_ref",   "13A. Left Job Mapping Reference",             sec13_left_job_reference),
    ("14_transport",      "14. Transportation (Q6 + Q6a + Q9)",          sec14_transport),
    ("15_voter_reg",      "15. Voter Registration (Q7 + Q7a)",           sec15_voter_reg),
    ("15_voter_reg_ref",  "15A. Voter Registration Mapping Reference",    sec15_voter_reg_reference),
    ("16_visit",          "16. Visit Frequency + Reasons + Barriers",    sec16_visit),
    ("17_impact",         "17. Program Impact (Q17 + Q16 + Q21)",        sec17_program_impact),
    ("18_respect",        "18. Staff & Peer Respect (Q18 + Q19)",        sec18_respect),
    ("19_environment",    "19. Program Environment (Q20)",               sec19_environment),
    ("20_banking",        "20. Banking (Q25 + Q24 + Q26b)",              sec20_banking),
    ("21_nps",            "21. NPS (Q22)",                               sec21_nps),
    ("22_comments",       "22. Additional Comments (Q23)",               sec22_comments),
]


def generate_charts(selected_charts=None):
    """Generate chart PNGs from analysis_412YZ.xlsx into output/412YZ/charts/."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np

    CHARTS_DIR = OUT_PATH.parent / "charts"
    CHARTS_DIR.mkdir(parents=True, exist_ok=True)

    PALETTE = ["#426FB3", "#C67A2A", "#8FB98A", "#B24A3D", "#D2A72C"]
    ACCENT_DARK = "#2E568F"
    # Shared 4-step analytical palette (low-risk → high-risk or good → bad).
    # Used for housing status, and as accent source for single-highlight charts.
    ANALYTICAL_PALETTE = ["#2E86AB", "#A8C5DA", "#F4A261", "#C1440E"]
    ACCENT_ORANGE = ANALYTICAL_PALETTE[2]   # soft orange — shared highlight color
    GRID_COLOR = "#D9DDE3"
    TEXT_COLOR = "#1F2933"
    AXIS_COLOR = "#9AA5B1"

    def _style_ticks(ax, x_size=9, y_size=10):
        ax.tick_params(axis="x", labelsize=x_size, colors=TEXT_COLOR)
        ax.tick_params(axis="y", labelsize=y_size, colors=TEXT_COLOR)

    def _apply_base_style(ax):
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color(AXIS_COLOR)
        ax.spines["bottom"].set_color(AXIS_COLOR)
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, color=GRID_COLOR, linewidth=0.8, zorder=0)
        ax.set_axisbelow(True)
        _style_ticks(ax)

    def _apply_hbar_style(ax):
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_visible(False)
        ax.spines["bottom"].set_color(AXIS_COLOR)
        ax.xaxis.grid(True, color=GRID_COLOR, linewidth=0.8, zorder=0)
        ax.yaxis.grid(False)
        ax.set_axisbelow(True)
        _style_ticks(ax)

    normalized_selected = None
    if selected_charts:
        normalized_selected = set()
        for name in selected_charts:
            clean = str(name).strip()
            if not clean:
                continue
            normalized_selected.add(clean)
            if clean.endswith(".png"):
                normalized_selected.add(clean[:-4])
            else:
                normalized_selected.add(f"{clean}.png")

    written_charts = []

    def _save(fig, name):
        if normalized_selected is not None and name not in normalized_selected and name[:-4] not in normalized_selected:
            plt.close(fig)
            return None
        fig.tight_layout()
        out = CHARTS_DIR / name
        fig.savefig(str(out), dpi=150, bbox_inches="tight")
        plt.close(fig)
        written_charts.append(out)
        return out

    def _load_sheet(sheet_name):
        """Load a sheet from analysis_412YZ.xlsx, skip title row (row 0), use row 1 as header."""
        import openpyxl as _opxl
        wb = _opxl.load_workbook(str(OUT_PATH), read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # row 0 = section title, row 1 = column headers, row 2+ = data
        if len(rows) < 2:
            return []
        headers = [str(h) if h is not None else "" for h in rows[1]]
        result = []
        for r in rows[2:]:
            if any(v not in (None, "") for v in r):
                obj = {headers[i]: (r[i] if r[i] is not None else "") for i in range(len(headers))}
                result.append(obj)
        return result

    # ------------------------------------------------------------------
    # Chart 1: Coach Satisfaction (horizontal bar, top-2 box %)
    # ------------------------------------------------------------------
    rows05 = _load_sheet("05_q1")
    coach_rows = []
    for r in rows05:
        raw_label = list(r.values())[0]
        raw_pct   = list(r.values())[2] if len(r) > 2 else ""
        if not raw_pct:
            continue
        full_label = str(raw_label).strip()
        label = (full_label[:35] + "…") if len(full_label) > 35 else full_label
        pct_val = float(str(raw_pct).replace("%", "").strip()) if str(raw_pct).replace("%", "").strip() else 0
        coach_rows.append((full_label, label, pct_val))

    if coach_rows:
        coach_rows = sorted(coach_rows, key=lambda item: (-item[2], item[0]))
        labels1 = [item[1] for item in coach_rows]
        vals1 = [item[2] for item in coach_rows]
        full_labels1 = [item[0] for item in coach_rows]
        colors1 = [ACCENT_ORANGE if label == "Is available to me when I need them" else PALETTE[0] for label in full_labels1]

        fig, ax = plt.subplots(figsize=(7.6, 5.0))
        fig.patch.set_edgecolor("white")
        fig.patch.set_linewidth(0)
        y_pos = range(len(labels1))
        bars = ax.barh(list(y_pos), vals1, color=colors1, zorder=3, height=0.8)
        ax.set_yticks(list(y_pos))
        ax.set_yticklabels(labels1, fontsize=10)
        ax.invert_yaxis()
        ax.set_xlim(0, 100)
        ax.xaxis.set_major_locator(mticker.MultipleLocator(20))
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x)}%"))
        ax.set_xlabel("Percent of respondents rating each item Often or All the Time", fontsize=9, color=TEXT_COLOR, labelpad=8)
        _apply_hbar_style(ax)
        fig.suptitle(
            "Coach Satisfaction Ratings, Mar-26 (% Often or All the Time)",
            x=0.5,
            y=1.0,
            ha="center",
            va="top",
            fontsize=12,
            fontweight="bold",
            color=TEXT_COLOR,
        )
        ax.set_title(
            "Current year only. See table above for trend across all survey years.",
            loc="left",
            fontsize=9,
            color=TEXT_COLOR,
            pad=10,
        )
        for bar, val in zip(bars, vals1):
            inside_bar = val >= 14
            x_pos = val - 1.8 if inside_bar else val + 1.2
            ax.text(
                x_pos,
                bar.get_y() + bar.get_height() / 2,
                f"{int(val)}%",
                va="center",
                ha="right" if inside_bar else "left",
                fontsize=9,
                fontweight="bold",
                color="white" if inside_bar else TEXT_COLOR,
            )
        _save(fig, "chart_01_coach_satisfaction.png")

    # ------------------------------------------------------------------
    # Chart 2: Housing Stability — 100% stacked horizontal bar by age
    # ------------------------------------------------------------------
    rows07 = _load_sheet("07_housing")

    # Build age × status cross-tab from the CSV directly
    csv_df = pd.read_csv(str(CSV_PATH), encoding="utf-8-sig", dtype=str).fillna("")
    csv_df["_age"] = csv_df["age_range"].apply(age_label)

    HOUSING_CHART_LABELS = {
        "stable":          "Safe and stable",
        "safe_not_90days": "Safe <90 days",
        "90days_not_safe": "Can stay, not safe",
        "no_place":        "No place to stay",
    }
    HOUSING_CHART_ORDER = ["stable", "safe_not_90days", "90days_not_safe", "no_place"]
    age_groups_h = ["16-17 years old", "18-20 years old", "21-23 years old", "Unknown"]
    age_display_h = ["16-17", "18-20", "21-23", "Unknown"]

    h_data = {}
    for age in age_groups_h:
        sub = csv_df[csv_df["_age"] == age]
        counts = {code: int((sub["q12_housing_stability"] == code).sum()) for code in HOUSING_CHART_ORDER}
        h_data[age] = counts

    # Only keep age groups that have at least 1 respondent
    age_groups_h2 = [a for a in age_groups_h if sum(h_data[a].values()) > 0]
    age_display_h2 = [age_display_h[age_groups_h.index(a)] for a in age_groups_h2]

    HOUSING_COLORS = ANALYTICAL_PALETTE  # steel blue → light blue → soft orange → brick red
    fig, ax = plt.subplots(figsize=(8, 4.5))
    lefts = np.zeros(len(age_groups_h2))
    for ci, code in enumerate(HOUSING_CHART_ORDER):
        seg_counts = np.array([h_data[a][code] for a in age_groups_h2], dtype=float)
        row_totals = np.array([sum(h_data[a].values()) for a in age_groups_h2], dtype=float)
        seg_pcts   = np.where(row_totals > 0, 100 * seg_counts / row_totals, 0)
        bars = ax.barh(range(len(age_groups_h2)), seg_pcts, left=lefts,
                       color=HOUSING_COLORS[ci], label=HOUSING_CHART_LABELS[code], zorder=3)
        label_color = TEXT_COLOR if ci in (1, 2) else "white"
        for i, (bar, pval) in enumerate(zip(bars, seg_pcts)):
            if pval >= 8:
                ax.text(lefts[i] + pval / 2, bar.get_y() + bar.get_height() / 2,
                        f"{int(round(pval))}%", va="center", ha="center", fontsize=9, color=label_color)
        lefts += seg_pcts
    ax.set_yticks(range(len(age_groups_h2)))
    ax.set_yticklabels(age_display_h2, fontsize=10)
    ax.set_xlim(0, 100)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x)}%"))
    ax.invert_yaxis()
    ax.set_title("Current Housing Status by Age", fontsize=11, pad=10)
    _apply_hbar_style(ax)
    ax.legend(loc="upper left", bbox_to_anchor=(1, 1), fontsize=9, frameon=False)
    _save(fig, "chart_02_housing_stability.png")

    # ------------------------------------------------------------------
    # Chart 3: Employment by Age — 100% stacked horizontal bar
    # ------------------------------------------------------------------
    EMP_CHART_MAP = {
        "yes_full_time":        "Full time",
        "yes_part_time":        "Part time",
        "job_training_program": "Job training",
        "no":                   "Not working",
    }
    EMP_CHART_ORDER = ["yes_full_time", "yes_part_time", "job_training_program", "no"]

    age_groups_e = ["16-17 years old", "18-20 years old", "21-23 years old", "Unknown"]
    age_display_e = ["16-17", "18-20", "21-23", "Unknown"]

    e_data = {}
    for age in age_groups_e:
        sub = csv_df[(csv_df["_age"] == age) & (csv_df["q8_employment_status"] != "")]
        counts = {code: int((sub["q8_employment_status"] == code).sum()) for code in EMP_CHART_ORDER}
        e_data[age] = counts

    age_groups_e2  = [a for a in age_groups_e if sum(e_data[a].values()) > 0]
    age_display_e2 = [age_display_e[age_groups_e.index(a)] for a in age_groups_e2]

    fig, ax = plt.subplots(figsize=(8, 4.0))
    lefts = np.zeros(len(age_groups_e2))
    y_positions = np.arange(len(age_groups_e2)) * 0.52
    EMP_CHART_COLORS = ANALYTICAL_PALETTE
    for ci, code in enumerate(EMP_CHART_ORDER):
        seg_counts = np.array([e_data[a][code] for a in age_groups_e2], dtype=float)
        row_totals = np.array([sum(e_data[a].values()) for a in age_groups_e2], dtype=float)
        seg_pcts   = np.where(row_totals > 0, 100 * seg_counts / row_totals, 0)
        bars = ax.barh(y_positions, seg_pcts, left=lefts,
                       color=EMP_CHART_COLORS[ci], label=EMP_CHART_MAP[code], zorder=3, height=0.4)
        for i, (bar, pval) in enumerate(zip(bars, seg_pcts)):
            if pval >= 8:
                ax.text(lefts[i] + pval / 2, bar.get_y() + bar.get_height() / 2,
                        f"{int(round(pval))}%", va="center", ha="center", fontsize=9, color="white")
        lefts += seg_pcts
    ax.set_yticks(y_positions)
    ax.set_yticklabels(age_display_e2, fontsize=10)
    ax.set_xlim(0, 100)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x)}%"))
    ax.invert_yaxis()
    ax.set_title("Employment Status by Age", fontsize=11, pad=10)
    _apply_hbar_style(ax)
    ax.legend(loc="upper left", bbox_to_anchor=(1, 1), fontsize=9, frameon=False)
    _save(fig, "chart_03_employment_by_age.png")

    # ------------------------------------------------------------------
    # Chart 3b: Employment by Self-Reported School Enrollment
    # Match prior-year category order and legend order.
    # ------------------------------------------------------------------
    SCHOOL_EMPLOYMENT_LABELS = {
        "not_in_school":   "Not in school",
        "high_school":     "High school",
        "ged":             "GED program",
        "college_career":  "College/Career or technical school",
        "graduate":        "Graduate school",
    }
    SCHOOL_EMPLOYMENT_ORDER = [
        "not_in_school",
        "high_school",
        "ged",
        "college_career",
        "graduate",
    ]
    SCHOOL_EMPLOYMENT_STATUS = [
        ("yes_full_time", "Full time", "#4F81BD"),
        ("yes_part_time", "Part time", "#9DC3E6"),
        ("job_training_program", "Job training program", "#F79646"),
        ("no", "Unemployed", "#C00000"),
    ]

    school_totals = {
        code: int((csv_df["q5_school_status"] == code).sum())
        for code in SCHOOL_EMPLOYMENT_ORDER
    }
    school_codes = [code for code in SCHOOL_EMPLOYMENT_ORDER if school_totals[code] > 0]
    school_labels = [SCHOOL_EMPLOYMENT_LABELS[code] for code in school_codes]
    school_n = sum(school_totals[code] for code in school_codes)

    if school_codes:
        x = np.arange(len(school_codes))
        width = 0.17
        offsets = np.linspace(-1.5 * width, 1.5 * width, len(SCHOOL_EMPLOYMENT_STATUS))

        fig, ax = plt.subplots(figsize=(8, 4.8))
        for offset, (status_code, status_label, color) in zip(offsets, SCHOOL_EMPLOYMENT_STATUS):
            vals = [
                int(((csv_df["q5_school_status"] == school_code) & (csv_df["q8_employment_status"] == status_code)).sum())
                for school_code in school_codes
            ]
            bars = ax.bar(x + offset, vals, width=width, label=status_label, color=color, zorder=3)
            for bar, val in zip(bars, vals):
                if val > 0:
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        val + 0.8,
                        f"{val}",
                        ha="center",
                        va="bottom",
                        fontsize=9,
                    )

        school_labels_wrapped = [
            "College/Career or\ntechnical school" if label == "College/Career or technical school" else label
            for label in school_labels
        ]
        ax.set_xticks(x)
        ax.set_xticklabels(school_labels_wrapped, fontsize=10)
        ax.set_ylim(0, max(1, max(
            int(((csv_df["q5_school_status"] == school_code) & (csv_df["q8_employment_status"] == status_code)).sum())
            for school_code in school_codes
            for status_code, _, _ in SCHOOL_EMPLOYMENT_STATUS
        )) + 10)
        ax.set_title(
            f"Youth Employment Status by Self-Reported School Enrollment\nMarch 2026, n={school_n}",
            fontsize=11,
            pad=12,
        )
        _apply_base_style(ax)
        ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.02), ncol=4, fontsize=9, frameon=False,
                  handlelength=0.6, handletextpad=0.3, columnspacing=1.4)
        _save(fig, "chart_09_employment_by_school.png")

    # ------------------------------------------------------------------
    # Chart 4: Visit Frequency by Age — grouped vertical bar
    # Build directly from csv_df (same logic as sec16_visit) to avoid
    # multi-sub-table sheet parsing problems.
    # ------------------------------------------------------------------
    VISIT_FREQ_LABELS_CHART = {
        "every_week":               "Every week",
        "1_3_times_per_month":      "1-3x/month",
        "less_than_once_per_month": "<1x/month",
        "never":                    "Never",
    }
    VISIT_AGE_COLS  = ["16-17 years old", "18-20 years old", "21-23 years old"]
    VISIT_DISP_COLS = ["16-17", "18-20", "21-23"]

    freq_labels_chart = [VISIT_FREQ_LABELS_CHART[c] for c in Q15_ORDER]
    freq_matrix = {d: [] for d in VISIT_DISP_COLS}
    for code in Q15_ORDER:
        for age_col, disp in zip(VISIT_AGE_COLS, VISIT_DISP_COLS):
            cnt = int(((csv_df["_age"] == age_col) & (csv_df["q15_visit_frequency"] == code)).sum())
            freq_matrix[disp].append(cnt)

    if freq_labels_chart:
        n_groups  = len(freq_labels_chart)
        n_bars    = len(VISIT_DISP_COLS)
        width     = 0.25
        x         = np.arange(n_groups)
        visit_chart_colors = [PALETTE[0], "#6EA0E0", "#A8C5DA"]
        total_visit_responses = int((csv_df["q15_visit_frequency"] != "").sum())

        fig, ax = plt.subplots(figsize=(8, 4.5))
        for bi, (disp, col) in enumerate(zip(VISIT_DISP_COLS, visit_chart_colors[:n_bars])):
            offsets = x + (bi - (n_bars - 1) / 2) * width
            bars = ax.bar(offsets, freq_matrix[disp], width=width, color=col, label=disp, zorder=3)
            for bar, val in zip(bars, freq_matrix[disp]):
                if val > 0:
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        val + 0.7,
                        f"{val}",
                        ha="center",
                        va="bottom",
                        fontsize=8.5,
                        color=TEXT_COLOR,
                    )
        ax.set_xticks(x)
        ax.set_xticklabels(freq_labels_chart, fontsize=10)
        ax.set_ylabel("Number of Youth", fontsize=10)
        ax.set_xlabel("Reported Frequency Visiting the Youth Zone", fontsize=10, color=TEXT_COLOR, labelpad=10)
        ax.set_title(
            f"Visit Frequency by Age\nMarch 2026, n={total_visit_responses}",
            fontsize=13,
            fontweight="bold",
            color=TEXT_COLOR,
            pad=14,
        )
        ax.set_ylim(0, max(max(vals) for vals in freq_matrix.values()) + 6)
        _apply_base_style(ax)
        ax.xaxis.grid(False)
        ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.03), ncol=3, fontsize=9, frameon=False)
        _save(fig, "chart_04_visit_frequency.png")

    # ------------------------------------------------------------------
    # Chart 10: Stay Focused on Goals by Visit Frequency — grouped % bars
    # ------------------------------------------------------------------
    focus_df = csv_df[(csv_df["q15_visit_frequency"] != "") & (csv_df["q16_stay_focused"] != "")].copy()
    focus_visit_labels = ["Every week", "1-3 times per month", "Less than once per\nmonth", "Never"]
    focus_resp_labels = [
        "Agree",
        "Somewhat agree",
        "Disagree",
        "Unsure, I don't have clear goals right now",
    ]
    focus_resp_colors = [PALETTE[0], "#6EA0E0", "#C1440E", "#CFCFCF"]
    focus_matrix = {label: [] for label in focus_resp_labels}

    for visit_code in Q15_ORDER:
        visit_df = focus_df[focus_df["q15_visit_frequency"] == visit_code]
        visit_total = len(visit_df)
        for focus_code, focus_label in zip(Q16_ORDER, focus_resp_labels):
            count = int((visit_df["q16_stay_focused"] == focus_code).sum())
            pct_value = (100 * count / visit_total) if visit_total else 0
            focus_matrix[focus_label].append(pct_value)

    if len(focus_df) > 0:
        n_groups = len(focus_visit_labels)
        n_bars = len(focus_resp_labels)
        width = 0.18
        x = np.arange(n_groups)

        fig, ax = plt.subplots(figsize=(9.6, 5.6))
        for bi, (resp_label, color) in enumerate(zip(focus_resp_labels, focus_resp_colors)):
            offsets = x + (bi - (n_bars - 1) / 2) * width
            vals = focus_matrix[resp_label]
            bars = ax.bar(offsets, vals, width=width, color=color, label=resp_label, zorder=3)
            for bar, val in zip(bars, vals):
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    val + 1.5,
                    f"{round(val):.0f}%",
                    ha="center",
                    va="bottom",
                    fontsize=8.5,
                    color=TEXT_COLOR,
                )

        ax.set_xticks(x)
        ax.set_xticklabels(focus_visit_labels, fontsize=10)
        ax.set_ylim(0, max(max(vals) for vals in focus_matrix.values()) + 16)
        ax.yaxis.set_major_locator(mticker.MultipleLocator(20))
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{int(y)}%"))
        ax.set_xlabel("Reported Frequency Visiting the Youth Zone", fontsize=10, color=TEXT_COLOR, labelpad=10)
        ax.set_title(
            f"The Youth Zone Helps Me Stay\nFocused On My Goals\nMarch 2026, n={len(focus_df)}",
            fontsize=11,
            fontweight="bold",
            color=TEXT_COLOR,
            pad=22,
        )
        _apply_base_style(ax)
        ax.xaxis.grid(False)
        ax.legend(
            loc="upper center",
            bbox_to_anchor=(0.5, 1.04),
            ncol=4,
            fontsize=9,
            frameon=False,
            handlelength=0.8,
            handletextpad=0.4,
            columnspacing=1.1,
        )
        _save(fig, "chart_10_stay_focused_visit_frequency.png")

    # ------------------------------------------------------------------
    # Chart 11: Visit reasons by frequency — integrated table/chart
    # ------------------------------------------------------------------
    reason_combo_order = [
        ("see_coach_staff", "See my Youth Coach"),
        ("food", "Eat food"),
        ("scheduled_activity", "Scheduled activity"),
        ("work_on_goals", "Work toward my goals"),
        ("learn_skills", "Learn new things, skills"),
        ("safe_place", "Safe place"),
        ("socialize", "Socialize/see friends"),
        ("escape_problems", "Escape problems elsewhere"),
        ("service_providers", "Service providers"),
        ("laundry_shower", "Do laundry, shower, etc."),
        ("health_counseling", "Health care, counseling"),
        ("computers", "Access to computers"),
    ]
    weekly_df = csv_df[csv_df["q15_visit_frequency"] == "every_week"]
    monthly_df = csv_df[csv_df["q15_visit_frequency"] == "1_3_times_per_month"]
    combo_df = csv_df[csv_df["q15_visit_frequency"].isin(["every_week", "1_3_times_per_month"])]

    def _multi_count(sub_df, field, code):
        return int(sum(code in split_pipe(val) for val in sub_df[field]))

    parsed_rows = []
    for code, label in reason_combo_order:
        weekly_count = _multi_count(weekly_df, "q15a_visit_reasons", code)
        monthly_count = _multi_count(monthly_df, "q15a_visit_reasons", code)
        total_count = weekly_count + monthly_count
        parsed_rows.append({
            "reason": label,
            "weekly_count": weekly_count,
            "monthly_count": monthly_count,
            "total_count": total_count,
            "weekly_pct": (100 * weekly_count / len(weekly_df)) if len(weekly_df) else 0,
            "monthly_pct": (100 * monthly_count / len(monthly_df)) if len(monthly_df) else 0,
            "total_pct": (100 * total_count / len(combo_df)) if len(combo_df) else 0,
        })
    parsed_rows = sorted(parsed_rows, key=lambda row: (-row["total_count"], row["reason"]))

    weekly_n = len(weekly_df)
    monthly_n = len(monthly_df)
    combo_total_n = len(combo_df)

    if parsed_rows and combo_total_n > 0:
        fig_h = max(5.2, 1.8 + 0.33 * len(parsed_rows))
        fig, ax = plt.subplots(figsize=(10.4, fig_h))
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")

        left = 0.03
        reason_w = 0.22
        week_w = 0.20
        month_w = 0.20
        total_w = 0.20
        cell_h = 0.052
        header_h = 0.064
        top = 0.88
        header_fill = "#DCE6F1"
        bar_fill = "#5B88C2"
        border = "#6B7280"

        ax.text(0.5, 0.988, "What Are the Main Reasons Youth Come to the Youth Zone?",
                ha="center", va="top", fontsize=13.5, fontweight="bold", color=TEXT_COLOR)
        ax.text(0.5, 0.944, "Among youth who reported visiting at least monthly, March 2026",
                ha="center", va="top", fontsize=9.5, color=TEXT_COLOR, style="italic")

        y = top - header_h
        x_positions = [left, left + reason_w, left + reason_w + week_w, left + reason_w + week_w + month_w]
        widths = [reason_w, week_w, month_w, total_w]
        headers = ["Frequency at the Zone", "Every week", "1 to 3 times per\nmonth", "Total"]

        for x0, width, label in zip(x_positions, widths, headers):
            fill = "white" if label == "Frequency at the Zone" else header_fill
            ax.add_patch(plt.Rectangle((x0, y), width, header_h, facecolor=fill, edgecolor=border, linewidth=0.7))
            ax.text(x0 + width / 2, y + header_h / 2, label, ha="center", va="center",
                    fontsize=9.4, color=TEXT_COLOR, fontweight="bold" if label != "Frequency at the Zone" else None,
                    style="italic" if label == "Frequency at the Zone" else None)

        y -= cell_h
        counts = ["Number of Youth", str(weekly_n), str(monthly_n), str(combo_total_n)]
        for x0, width, label in zip(x_positions, widths, counts):
            ax.add_patch(plt.Rectangle((x0, y), width, cell_h, facecolor="white", edgecolor=border, linewidth=0.7))
            ax.text(x0 + width / 2, y + cell_h / 2, label, ha="center", va="center",
                    fontsize=10, color=TEXT_COLOR, fontweight="bold" if label != "Number of Youth" else None,
                    style="italic" if label == "Number of Youth" else None)

        bar_cols = [
            (left + reason_w, week_w, "weekly_pct", "weekly_count"),
            (left + reason_w + week_w, month_w, "monthly_pct", "monthly_count"),
            (left + reason_w + week_w + month_w, total_w, "total_pct", "total_count"),
        ]

        for row in parsed_rows:
            y -= cell_h
            ax.add_patch(plt.Rectangle((left, y), reason_w, cell_h, facecolor="white", edgecolor=border, linewidth=0.7))
            ax.text(left + 0.006, y + cell_h / 2, row["reason"], ha="left", va="center", fontsize=9.5, color=TEXT_COLOR)

            for x0, width, pct_key, count_key in bar_cols:
                ax.add_patch(plt.Rectangle((x0, y), width, cell_h, facecolor="white", edgecolor=border, linewidth=0.7))
                pct_val = row[pct_key]
                if pct_val > 0:
                    bar_margin = 0.003
                    usable_w = width - 2 * bar_margin
                    bar_w = usable_w * min(pct_val / 100, 1)
                    ax.add_patch(plt.Rectangle((x0 + bar_margin, y + 0.006), bar_w, cell_h - 0.012,
                                               facecolor=bar_fill, edgecolor="none"))
                ax.text(x0 + width - 0.008, y + cell_h / 2, f"{int(round(pct_val))}%", ha="right", va="center",
                        fontsize=9.5, color=TEXT_COLOR)

        _save(fig, "chart_11_visit_reasons_combo.png")

    # ------------------------------------------------------------------
    # Chart 12: Program helped with by age — integrated table/chart
    # ------------------------------------------------------------------
    q17_age_groups = ["16-17 years old", "18-20 years old", "21-23 years old"]
    q17_age_headers = ["16-17 years old", "18-20 years old", "21-23 years old", "Total"]
    q17_chart_order = [
        ("future", "Think about my future"),
        ("decision_making", "Make good decisions"),
        ("handle_problems", "Figure out how to handle problems"),
        ("vital_documents", "Obtain vital documents"),
        ("positive_relationships", "Establish positive relationships"),
        ("drivers_license", "Get my driver's license"),
        ("health_counseling", "Access health care and/or counseling"),
        ("housing", "Find or maintain housing"),
        ("job", "Get or keep a job"),
        ("education", "Finish or further my education"),
        ("everyday_skills", "Learn everyday skills"),
        ("parenting", "Improve parenting skills"),
        ("something_else", "Something else"),
    ]
    q17_df = csv_df[(csv_df["q17_program_helped"].astype(str).str.strip() != "") & (csv_df["_age"].isin(q17_age_groups))].copy()
    q17_n_by_age = {age: len(q17_df[q17_df["_age"] == age]) for age in q17_age_groups}
    q17_total_known = sum(q17_n_by_age.values())
    q17_rows = []
    for code, label in q17_chart_order:
        counts_by_age = []
        total_count = 0
        for age in q17_age_groups:
            age_df = q17_df[q17_df["_age"] == age]
            count = int(sum(code in split_pipe(val) for val in age_df["q17_program_helped"]))
            counts_by_age.append(count)
            total_count += count
        if total_count == 0:
            continue
        q17_rows.append({
            "label": label,
            "counts": counts_by_age,
            "pcts": [100 * counts_by_age[i] / q17_n_by_age[age] if q17_n_by_age[age] else 0 for i, age in enumerate(q17_age_groups)],
            "total_count": total_count,
            "total_pct": 100 * total_count / q17_total_known if q17_total_known else 0,
        })
    q17_rows = sorted(q17_rows, key=lambda row: (-row["total_count"], row["label"]))

    if q17_rows and q17_total_known > 0:
        fig_h = max(5.4, 1.7 + 0.31 * len(q17_rows))
        fig, ax = plt.subplots(figsize=(11.0, fig_h))
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis("off")

        left = 0.02
        reason_w = 0.22
        data_w = 0.17
        total_w = 0.17
        cell_h = 0.052
        header_h = 0.064
        top = 0.90
        header_fill = "#DCE6F1"
        bar_fill = "#5B88C2"
        border = "#6B7280"

        ax.text(0.5, 0.985, "My Coach or the Youth Zone has Helped Me To... (by Age)",
                ha="center", va="top", fontsize=13.5, fontweight="bold", color=TEXT_COLOR)

        y = top - header_h
        x_positions = [left, left + reason_w, left + reason_w + data_w, left + reason_w + 2 * data_w, left + reason_w + 3 * data_w]
        widths = [reason_w, data_w, data_w, data_w, total_w]
        headers = ["Age", *q17_age_headers]

        for x0, width, label in zip(x_positions, widths, headers):
            fill = "white" if label == "Age" else header_fill
            ax.add_patch(plt.Rectangle((x0, y), width, header_h, facecolor=fill, edgecolor=border, linewidth=0.7))
            ax.text(x0 + width / 2, y + header_h / 2, label, ha="center", va="center",
                    fontsize=9.4, color=TEXT_COLOR, fontweight="bold" if label != "Age" else None,
                    style="italic" if label == "Age" else None)

        y -= cell_h
        counts = ["Number of Youth", str(q17_n_by_age[q17_age_groups[0]]), str(q17_n_by_age[q17_age_groups[1]]), str(q17_n_by_age[q17_age_groups[2]]), str(q17_total_known)]
        for x0, width, label in zip(x_positions, widths, counts):
            ax.add_patch(plt.Rectangle((x0, y), width, cell_h, facecolor="white", edgecolor=border, linewidth=0.7))
            ax.text(x0 + width / 2, y + cell_h / 2, label, ha="center", va="center",
                    fontsize=10, color=TEXT_COLOR, fontweight="bold" if label != "Number of Youth" else None,
                    style="italic" if label == "Number of Youth" else None)

        bar_cols = [
            (left + reason_w, data_w, 0),
            (left + reason_w + data_w, data_w, 1),
            (left + reason_w + 2 * data_w, data_w, 2),
            (left + reason_w + 3 * data_w, total_w, "total"),
        ]

        for row in q17_rows:
            y -= cell_h
            ax.add_patch(plt.Rectangle((left, y), reason_w, cell_h, facecolor="white", edgecolor=border, linewidth=0.7))
            ax.text(left + 0.006, y + cell_h / 2, row["label"], ha="left", va="center", fontsize=9.5, color=TEXT_COLOR)

            for x0, width, idx in bar_cols:
                ax.add_patch(plt.Rectangle((x0, y), width, cell_h, facecolor="white", edgecolor=border, linewidth=0.7))
                pct_val = row["total_pct"] if idx == "total" else row["pcts"][idx]
                if pct_val > 0:
                    bar_margin = 0.003
                    usable_w = width - 2 * bar_margin
                    bar_w = usable_w * min(pct_val / 100, 1)
                    ax.add_patch(plt.Rectangle((x0 + bar_margin, y + 0.006), bar_w, cell_h - 0.012,
                                               facecolor=bar_fill, edgecolor="none"))
                ax.text(x0 + width - 0.008, y + cell_h / 2, f"{int(round(pct_val))}%", ha="right", va="center",
                        fontsize=9.5, color=TEXT_COLOR)

        _save(fig, "chart_12_program_helped_combo.png")

    # ------------------------------------------------------------------
    # Chart 13: Respect and acceptance by staff vs peers — 100% stacked
    # ------------------------------------------------------------------
    respect_order = ["all_the_time", "often", "sometimes", "rarely", "never"]
    respect_labels = {
        "all_the_time": "All the time",
        "often": "Often",
        "sometimes": "Sometimes",
        "rarely": "Rarely",
        "never": "Never",
    }
    respect_colors = {
        "all_the_time": "#2E568F",
        "often": "#2E86AB",
        "sometimes": "#D9DDE3",
        "rarely": "#F4A261",
        "never": "#C1440E",
    }
    respect_items = [
        ("q18_staff_respect", "By Staff"),
        ("q19_peer_respect", "By Peers"),
    ]
    respect_counts = {}
    respect_totals = {}
    for field, label in respect_items:
        valid = csv_df[csv_df[field].astype(str).str.strip() != ""][field]
        respect_totals[label] = len(valid)
        respect_counts[label] = {code: int((valid == code).sum()) for code in respect_order}

    if all(respect_totals.values()):
        fig, ax = plt.subplots(figsize=(8.8, 4.9))
        y_pos = np.array([0.0, 0.48])
        lefts = np.zeros(len(respect_items))
        top2_pcts = []

        for _, label in respect_items:
            total = respect_totals[label]
            top2_pcts.append(
                100 * (respect_counts[label]["all_the_time"] + respect_counts[label]["often"]) / total if total else 0
            )

        for code in respect_order:
            pcts = []
            for _, label in respect_items:
                total = respect_totals[label]
                pct_val = 100 * respect_counts[label][code] / total if total else 0
                pcts.append(pct_val)
            bars = ax.barh(y_pos, pcts, left=lefts, height=0.28, color=respect_colors[code], label=respect_labels[code], zorder=3)
            for i, (bar, pval) in enumerate(zip(bars, pcts)):
                if pval >= 3:
                    x_text = lefts[i] + pval / 2
                    text_color = "white" if code in ("all_the_time", "often", "never") else TEXT_COLOR
                    ax.text(x_text, bar.get_y() + bar.get_height() / 2,
                            f"{int(round(pval))}%", va="center", ha="center", fontsize=10, color=text_color)
            lefts += np.array(pcts)

        for i, top2_pct in enumerate(top2_pcts):
            y_line = y_pos[i] + 0.225
            ax.plot([0, top2_pct], [y_line, y_line], color=TEXT_COLOR, linewidth=1)
            ax.plot([0, 0], [y_line - 0.04, y_line + 0.04], color=TEXT_COLOR, linewidth=1)
            ax.plot([top2_pct, top2_pct], [y_line - 0.04, y_line + 0.04], color=TEXT_COLOR, linewidth=1)
            ax.text(top2_pct + 1.4, y_line, f"Top-2 {int(round(top2_pct))}%", va="center", ha="left", fontsize=10, color=TEXT_COLOR)

        ax.set_yticks(y_pos)
        ax.set_yticklabels([label for _, label in respect_items], fontsize=10)
        ax.set_xlim(0, 100)
        ax.set_title("How Often Do You Feel You Are Treated with\nRespect and Accepted for Who You Are?",
                     fontsize=10, fontweight="bold", color=TEXT_COLOR, pad=55)
        _apply_hbar_style(ax)
        ax.spines["bottom"].set_visible(False)
        ax.xaxis.grid(False)
        ax.xaxis.set_visible(False)
        ax.invert_yaxis()
        ax.legend(loc="lower center", bbox_to_anchor=(0.5, 1.04), ncol=5, fontsize=9, frameon=False,
                  handlelength=0.8, handletextpad=0.3, columnspacing=1.2)
        _save(fig, "chart_13_respect_acceptance.png")

    # ------------------------------------------------------------------
    # Chart 14: Program environment (Q20) — 100% stacked, top-2 shown
    # ------------------------------------------------------------------
    environment_order = ["5", "4", "3", "2", "1"]
    environment_labels = {
        "5": "All the time",
        "4": "Often",
        "3": "Sometimes",
        "2": "Rarely",
        "1": "Never",
    }
    environment_colors = {
        "5": respect_colors["all_the_time"],
        "4": respect_colors["often"],
        "3": respect_colors["sometimes"],
        "2": respect_colors["rarely"],
        "1": respect_colors["never"],
    }
    environment_items = list(Q20_FIELDS)
    environment_counts = {}
    environment_totals = {}
    for field, label in environment_items:
        valid = csv_df[csv_df[field].astype(str).str.strip() != ""][field].astype(str).str.replace(r"\.0$", "", regex=True)
        environment_totals[label] = len(valid)
        environment_counts[label] = {code: int((valid == code).sum()) for code in environment_order}

    if all(environment_totals.values()):
        fig, ax = plt.subplots(figsize=(9.8, 6.0))
        y_pos = np.arange(len(environment_items)) * 0.62
        lefts = np.zeros(len(environment_items))
        top2_pcts = []

        for _, label in environment_items:
            total = environment_totals[label]
            top2_pcts.append(
                100 * (environment_counts[label]["5"] + environment_counts[label]["4"]) / total if total else 0
            )

        for code in environment_order:
            pcts = []
            for _, label in environment_items:
                total = environment_totals[label]
                pct_val = 100 * environment_counts[label][code] / total if total else 0
                pcts.append(pct_val)
            bars = ax.barh(y_pos, pcts, left=lefts, height=0.38, color=environment_colors[code], label=environment_labels[code], zorder=3)
            for i, (bar, pval) in enumerate(zip(bars, pcts)):
                if pval >= 3:
                    x_text = lefts[i] + pval / 2
                    text_color = "white" if code in ("5", "4", "1") else TEXT_COLOR
                    ax.text(x_text, bar.get_y() + bar.get_height() / 2,
                            f"{int(round(pval))}%", va="center", ha="center", fontsize=10, color=text_color)
            lefts += np.array(pcts)

        for i, top2_pct in enumerate(top2_pcts):
            y_line = y_pos[i] + 0.26
            ax.plot([0, top2_pct], [y_line, y_line], color=TEXT_COLOR, linewidth=1)
            ax.plot([0, 0], [y_line - 0.04, y_line + 0.04], color=TEXT_COLOR, linewidth=1)
            ax.plot([top2_pct, top2_pct], [y_line - 0.04, y_line + 0.04], color=TEXT_COLOR, linewidth=1)
            ax.text(top2_pct + 1.3, y_line, f"Top-2 {int(round(top2_pct))}%", va="center", ha="left", fontsize=10, color=TEXT_COLOR)

        ax.set_yticks(y_pos)
        ax.set_yticklabels([label for _, label in environment_items], fontsize=10)
        ax.set_xlim(0, 100)
        ax.set_title("How Do Youth Rate the Program Environment?\n(1-5 scale; top-2 shown below each item)",
                     fontsize=10, fontweight="bold", color=TEXT_COLOR, pad=55)
        _apply_hbar_style(ax)
        ax.spines["bottom"].set_visible(False)
        ax.xaxis.grid(False)
        ax.xaxis.set_visible(False)
        ax.invert_yaxis()
        ax.legend(loc="lower center", bbox_to_anchor=(0.5, 1.04), ncol=5, fontsize=9, frameon=False,
                  handlelength=0.8, handletextpad=0.3, columnspacing=1.2)
        _save(fig, "chart_14_environment_ratings.png")

    # ------------------------------------------------------------------
    # Chart 15: Primary Way Youth Get to Work — horizontal bar
    # ------------------------------------------------------------------
    csv_df_t = pd.read_csv(str(CSV_PATH), encoding="utf-8-sig", dtype=str).fillna("")
    counter_t = Counter(v.strip() for v in csv_df_t["q9_primary_transport"] if v.strip())
    total_t = sum(counter_t.values())
    _Q9_LABEL_MAP = {
        "public_transit":    "Bus or public transportation",
        "driving_self":      "Driving myself",
        "rides_from_others": "Getting rides from someone else",
        "rideshare":         "RideShare app (Lyft, Uber)",
        "active_transport":  "Walking, biking, scooter, etc.",
        "other":             "Other",
    }
    _Q9_ORDER = ["public_transit", "driving_self", "rides_from_others", "rideshare", "active_transport", "other"]
    transport_rows_c = []
    for code in _Q9_ORDER + [k for k in counter_t if k not in _Q9_ORDER]:
        n = counter_t.get(code, 0)
        if n == 0:
            continue
        transport_rows_c.append((_Q9_LABEL_MAP.get(code, code), n, 100 * n / total_t if total_t else 0))
    transport_rows_c.sort(key=lambda r: -r[2])

    if transport_rows_c and total_t > 0:
        t_labels = [r[0] for r in transport_rows_c]
        t_vals   = [r[2] for r in transport_rows_c]
        t_counts = [r[1] for r in transport_rows_c]

        fig, ax = plt.subplots(figsize=(7.0, max(3.0, 0.52 * len(t_labels) + 1.4)))
        fig.patch.set_facecolor("white")
        y_pos = range(len(t_labels))
        bars = ax.barh(list(y_pos), t_vals, color="#185FA5", zorder=3, height=0.62)
        ax.set_yticks(list(y_pos))
        ax.set_yticklabels(t_labels, fontsize=11)
        ax.invert_yaxis()
        ax.set_xlim(0, 100)
        ax.xaxis.set_major_locator(mticker.MultipleLocator(20))
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x)}%"))
        ax.set_xlabel("Percent of employed youth", fontsize=9, color=TEXT_COLOR, labelpad=8)
        _apply_hbar_style(ax)
        fig.suptitle(
            "Primary Way Youth Get to Work",
            x=0.5, y=1.02, ha="center", va="bottom",
            fontsize=13, fontweight="bold", color="#111",
        )
        ax.set_title(
            f"Among youth who are employed, n={total_t}",
            loc="left", fontsize=9, color=TEXT_COLOR, pad=10,
        )
        for bar, pval, cnt in zip(bars, t_vals, t_counts):
            inside_bar = pval >= 18
            x_pos = pval - 1.5 if inside_bar else pval + 1.5
            ax.text(
                x_pos,
                bar.get_y() + bar.get_height() / 2,
                f"{int(round(pval))}%  (n={cnt})",
                va="center",
                ha="right" if inside_bar else "left",
                fontsize=9.5,
                fontweight="bold",
                color="white" if inside_bar else TEXT_COLOR,
            )
        _save(fig, "chart_15_primary_transport.png")

    # ------------------------------------------------------------------
    # Chart 5: NPS — single 100% stacked horizontal bar
    # ------------------------------------------------------------------
    rows21 = _load_sheet("21_nps")
    nps_counts = {}
    nps_score_val = None
    for r in rows21:
        cat = str(list(r.values())[0]).strip()
        cnt_raw = list(r.values())[1] if len(r) > 1 else ""
        try:
            cnt = int(float(str(cnt_raw)))
        except (ValueError, TypeError):
            cnt = 0
        if "Promoter" in cat:
            nps_counts["promoter"] = cnt
        elif "Passive" in cat:
            nps_counts["passive"] = cnt
        elif "Detractor" in cat:
            nps_counts["detractor"] = cnt
        elif "NPS Score" in cat:
            nps_score_val = cnt_raw

    total_nps = sum(nps_counts.get(k, 0) for k in ("promoter", "passive", "detractor"))
    if total_nps > 0:
        det_pct  = 100 * nps_counts.get("detractor", 0) / total_nps
        pas_pct  = 100 * nps_counts.get("passive",   0) / total_nps
        pro_pct  = 100 * nps_counts.get("promoter",  0) / total_nps

        fig, ax = plt.subplots(figsize=(8, 2))
        # Order: Detractor | Passive | Promoter (left to right)
        seg_labels = ["Detractors", "Passives",  "Promoters"]
        seg_pcts   = [det_pct,       pas_pct,     pro_pct]
        seg_colors = ["#FF0000",     "#FFC000",   "#4472C4"]
        left = 0
        for lbl, pval, col in zip(seg_labels, seg_pcts, seg_colors):
            ax.barh([0], [pval], left=left, color=col, height=0.6, zorder=3)
            if pval >= 5:
                ax.text(left + pval / 2, 0, f"{int(round(pval))}%",
                        va="center", ha="center", fontsize=10, color="white", fontweight="bold")
            left += pval
        ax.set_xlim(0, 100)
        ax.set_ylim(-0.8, 0.8)
        ax.axis("off")
        nps_label = f"NPS = {nps_score_val}" if nps_score_val is not None else "NPS"
        ax.text(50, -0.55, nps_label, va="center", ha="center", fontsize=11, fontweight="bold")
        _save(fig, "chart_05_nps.png")

    # ------------------------------------------------------------------
    # Chart 6: Race Distribution — horizontal bar by count
    # ------------------------------------------------------------------
    rows03 = _load_sheet("03_race_once")
    race_labels, race_counts = [], []
    for r in rows03:
        label = str(list(r.values())[0]).strip()
        if label in ("", "Race/Ethnicity", "Total"):
            continue
        total_val = r.get("Total", "")
        try:
            cnt = int(float(str(total_val)))
        except (ValueError, TypeError):
            continue
        race_labels.append(label)
        race_counts.append(cnt)

    if race_labels:
        paired = sorted(zip(race_counts, race_labels), reverse=True)
        race_counts_s, race_labels_s = zip(*paired)
        race_total = sum(race_counts_s)
        race_display_labels = [
            "Other single\nracial identity" if label == "Other single racial identity" else label
            for label in race_labels_s
        ]

        fig, ax = plt.subplots(figsize=(7.4, 3.2))
        y_pos = range(len(race_display_labels))
        bars = ax.barh(list(y_pos), list(race_counts_s), color=PALETTE[0], zorder=3)
        ax.set_yticks(list(y_pos))
        ax.set_yticklabels(race_display_labels, fontsize=10)
        ax.set_xlim(0, max(race_counts_s) * 1.16)
        ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True, nbins=6))
        ax.set_xlabel("Number of respondents", fontsize=9, color=TEXT_COLOR, labelpad=8)
        _apply_hbar_style(ax)
        for bar, cnt in zip(bars, race_counts_s):
            pct = round(100 * cnt / race_total) if race_total else 0
            ax.text(
                cnt + (max(race_counts_s) * 0.02),
                bar.get_y() + bar.get_height() / 2,
                f"{cnt} ({pct}%)",
                va="center",
                ha="left",
                fontsize=9,
                fontweight="bold",
                color=TEXT_COLOR,
            )
        _save(fig, "chart_06_race_distribution.png")

    # ------------------------------------------------------------------
    # Chart 7: Communication Satisfaction (Q3) — horizontal bar
    # ------------------------------------------------------------------
    comm_df = pd.read_csv(str(CSV_PATH), encoding="utf-8-sig", dtype=str).fillna("")
    Q3_CODES  = ["good_amount", "not_enough", "too_much"]
    Q3_LABELS = ["Good amount", "Not enough", "Too much"]
    q3_total  = comm_df[comm_df["q3_communication_level"].isin(Q3_CODES)].shape[0]
    if q3_total > 0:
        q3_vals = [100 * int((comm_df["q3_communication_level"] == c).sum()) / q3_total
                   for c in Q3_CODES]
        labels7 = Q3_LABELS[::-1]   # reverse for barh (Good amount at top)
        vals7   = q3_vals[::-1]
        fig, ax = plt.subplots(figsize=(6, 3.2))
        fig.patch.set_edgecolor("white")
        fig.patch.set_linewidth(0)
        y_pos = range(len(labels7))
        bars = ax.barh(list(y_pos), vals7, color=PALETTE[0], zorder=3, height=0.8)
        ax.set_yticks(list(y_pos))
        ax.set_yticklabels(labels7, fontsize=10)
        ax.set_xlim(0, 100)
        ax.xaxis.set_major_locator(mticker.MultipleLocator(20))
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x)}%"))
        ax.set_xlabel("Percent of respondents", fontsize=9, color=TEXT_COLOR, labelpad=8)
        _apply_hbar_style(ax)
        fig.suptitle(
            "Rating of Coach Communication Amount, Mar-26",
            x=0.5,
            y=1.0,
            ha="center",
            va="top",
            fontsize=12,
            fontweight="bold",
            color=TEXT_COLOR,
        )
        ax.set_title(
            "How youth rated the frequency of communication with their coach",
            loc="left",
            fontsize=9,
            color=TEXT_COLOR,
            pad=10,
        )
        for bar, val in zip(bars, vals7):
            inside_bar = val >= 14
            x_pos = val - 1.8 if inside_bar else val + 1.2
            ax.text(
                x_pos,
                bar.get_y() + bar.get_height() / 2,
                f"{int(round(val))}%",
                va="center",
                ha="right" if inside_bar else "left",
                fontsize=9,
                fontweight="bold",
                color="white" if inside_bar else TEXT_COLOR,
            )
        _save(fig, "chart_07_communication_satisfaction.png")

    # ------------------------------------------------------------------
    # Chart 8: Q2 Frequency for Not Enough group — horizontal bar
    # ------------------------------------------------------------------
    Q2_CODES      = ["almost_every_day", "about_once_a_week", "1_2_times_per_month", "less_than_once_a_month"]
    Q2_LABELS_8   = ["Almost every day", "About once a week", "1\u20132 times per month", "Less than once a month"]
    ne_df         = comm_df[comm_df["q3_communication_level"] == "not_enough"]
    ne_q2_counts  = [int((ne_df["q2_communication_frequency"] == c).sum()) for c in Q2_CODES]
    ne_total = len(ne_df)
    if any(v > 0 for v in ne_q2_counts):
        labels8 = Q2_LABELS_8[::-1]   # reverse for barh (Almost every day at top)
        vals8   = ne_q2_counts[::-1]
        fig, ax = plt.subplots(figsize=(6, 3.2))
        fig.patch.set_edgecolor("white")
        fig.patch.set_linewidth(0)
        y_pos = range(len(labels8))
        bars = ax.barh(list(y_pos), vals8, color=PALETTE[0], zorder=3, height=0.8)
        ax.set_yticks(list(y_pos))
        ax.set_yticklabels(labels8, fontsize=10)
        max_val = max(vals8) if vals8 else 1
        ax.set_xlim(0, max_val * 1.18)
        ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: str(int(x))))
        ax.set_xlabel("Number of respondents", fontsize=9, color=TEXT_COLOR, labelpad=8)
        _apply_hbar_style(ax)
        fig.suptitle(
            f"Communication Frequency Among Youth Who Wanted More, Mar-26",
            x=0.5,
            y=1.0,
            ha="center",
            va="top",
            fontsize=12,
            fontweight="bold",
            color=TEXT_COLOR,
        )
        ax.set_title(
            f"Among {ne_total} youth who said communication was \u2018Not Enough\u2019 (Q3)",
            loc="left",
            fontsize=9,
            color=TEXT_COLOR,
            pad=10,
        )
        for bar, val in zip(bars, vals8):
            if val > 0:
                inside_bar = val >= max_val * 0.25
                x_pos = val - 0.3 if inside_bar else val + 0.3
                ax.text(
                    x_pos,
                    bar.get_y() + bar.get_height() / 2,
                    str(val),
                    va="center",
                    ha="right" if inside_bar else "left",
                    fontsize=9,
                    fontweight="bold",
                    color="white" if inside_bar else TEXT_COLOR,
                )
        _save(fig, "chart_08_communication_freq_not_enough.png")

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    if normalized_selected is not None:
        print(f"\nSelected charts written to {CHARTS_DIR}:")
    else:
        print(f"\nCharts written to {CHARTS_DIR}:")
    for png in sorted(written_charts, key=lambda path: path.name):
        size_kb = png.stat().st_size / 1024
        print(f"  {png.name}  ({size_kb:.1f} KB)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--charts",
        nargs="+",
        help="Optional list of chart names to regenerate, e.g. chart_10_stay_focused_visit_frequency or chart_10_stay_focused_visit_frequency.png",
    )
    args = parser.parse_args()

    if not CSV_PATH.exists():
        print(f"CSV not found: {CSV_PATH}")
        sys.exit(1)

    df = pd.read_csv(str(CSV_PATH), encoding="utf-8-sig", dtype=str).fillna("")
    df["_age"]    = df["age_range"].apply(age_label)
    df["_gender"] = df["gender"].apply(gender_group)

    print(f"Loaded {len(df)} surveys.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, title, func in SECTIONS:
        print(f"  {sheet_name}...")
        result = func(df)
        ws = wb.create_sheet(sheet_name)
        write_section(ws, title, result)
        autofit_columns(ws)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))
    print(f"\nSaved: {OUT_PATH}")

    generate_charts(args.charts)


if __name__ == "__main__":
    main()

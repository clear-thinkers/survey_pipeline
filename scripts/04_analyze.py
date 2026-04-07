"""
04_analyze.py
Compute descriptive statistics for output/412YZ/survey_data_412YZ.csv and write
output/412YZ/analysis_412YZ.xlsx with one sheet per reporting component (22 sheets).

Usage:
    python scripts/04_analyze.py
"""

import sys
from pathlib import Path
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
CSV_PATH = BASE_DIR / "output" / "412YZ" / "survey_data_412YZ.csv"
OUT_PATH = BASE_DIR / "output" / "412YZ" / "analysis_412YZ.xlsx"

# ---------------------------------------------------------------------------
# Display constants
# ---------------------------------------------------------------------------

AGE_MAP = {
    "16_17": "16-17 years old",
    "18_20": "18-20 years old",
    "21_23": "21-23 years old",
}
AGE_ORDER = ["16-17 years old", "18-20 years old", "21-23 years old", "Unknown"]
GENDER_ORDER = ["Female", "Male", "Trans, Non-binary"]

Q1_FIELDS = [
    ("q1_trustworthy",      "Is trustworthy"),
    ("q1_reliable",         "Is reliable"),
    ("q1_values_opinions",  "Values my opinions about my life"),
    ("q1_available",        "Is available to me when I need them"),
    ("q1_heard_understood", "Makes me feel heard and understood"),
]

Q20_FIELDS = [
    ("q20_people_care",      "People around me care about my success"),
    ("q20_no_judgment",      "I feel accepted without judgment"),
    ("q20_diversity_valued", "Diversity of backgrounds is valued"),
    ("q20_treated_fairly",   "I am treated fairly"),
    ("q20_safe_sharing",     "I feel safe sharing my thoughts"),
]

RACE_GROUP_ORDER = [
    "Black", "White", "Multiracial", "Other single racial identity",
    "Hispanic or Latinx", "Asian", "Native American or Native Hawaiian",
    "Prefer not to answer", "Other",
]

HOUSING_LABEL_MAP = {
    "stable":          "Safe and stable (90+ days)",
    "safe_not_90days": "Safe, but cannot stay 90 days",
    "90days_not_safe": "Can stay 90 days, but not safe",
    "no_place":        "No place to stay",
}
HOUSING_ORDER = ["stable", "safe_not_90days", "90days_not_safe", "no_place"]

SLEEP_LABEL_MAP = {
    "friends_family":     "Home, w/ Family/Friends",
    "shelter":            "Homeless shelter",
    "couch_surfing":      "Couch surfing",
    "car":                "In a car",
    "outside":            "Outside",
    "abandoned_building": "Abandoned building",
    "other":              "Other",
}

Q14_LABEL_MAP = {
    "lost_informal_housing": "Family/friends no longer let me stay",
    "left_unsafe":           "Left - felt unsafe",
    "evicted_nonpayment":    "Evicted, didn't/couldn't pay rent",
    "evicted_other":         "Evicted, other reason",
    "other":                 "Other",
}

Q8_LABEL_MAP = {
    "yes_full_time":        "Full time",
    "yes_part_time":        "Part time",
    "job_training_program": "Job training program",
    "no":                   "Not working",
}
Q8_ORDER = ["yes_full_time", "yes_part_time", "job_training_program", "no"]

Q8A_LABEL_MAP = {
    "less_3mo": "Less than 3 months",
    "3_6mo":    "3 to 6 months",
    "more_6mo": "More than 6 months",
}
Q8A_ORDER = ["less_3mo", "3_6mo", "more_6mo"]

Q10_LABEL_MAP = {
    "transportation":        "Transportation issues",
    "not_getting_called":    "Applying and not getting called",
    "mental_physical_health":"Mental or physical health",
    "limited_experience":    "Limited work experience",
    "childcare":             "Childcare/parenting challenges",
    "no_references":         "Weak references",
    "interview_skills":      "Interview skills",
    "no_diploma":            "Lack of HS diploma/GED",
    "criminal_background":   "Criminal background",
    "drugs_alcohol":         "Use of drugs or alcohol",
    "something_else":        "Something else",
}

Q11_LABEL_MAP = {
    "quit":               "Quit",
    "found_better":       "Found a better job",
    "seasonal":           "Seasonal/temporary",
    "fired_attendance":   "Fired: Missed work",
    "pregnancy_parenting":"Pregnant or Parenting",
    "fired_performance":  "Fired: Poor performance",
    "other":              "Other",
}

Q11A_LABEL_MAP = {
    "low_pay_hours":          "Low pay or not enough hours",
    "schedule_conflict":      "Schedule conflict",
    "lack_of_support":        "Lack of support or respect",
    "poor_conditions":        "Poor working conditions",
    "mental_emotional_health":"Mental/emotional health",
    "transportation":         "Transportation challenges",
    "not_good_fit":           "Not a good fit",
    "personal_family":        "Personal or family reasons",
    "other":                  "Other",
}

Q6_LABEL_MAP = {
    "yes":            "Yes",
    "learners_permit":"Learner's Permit",
    "no":             "No",
}
Q6_ORDER = ["yes", "learners_permit", "no"]

Q6A_LABEL_MAP = {
    "own_reliable":    "Own reliable vehicle",
    "own_unreliable":  "Own vehicle, but unreliable",
    "share_reliable":  "Share reliable vehicle",
    "share_unreliable":"Share vehicle, but unreliable",
    "borrow":          "Can borrow a car",
    "no_access":       "No reliable vehicle access",
}
Q6A_ORDER = ["own_reliable", "share_reliable", "borrow", "own_unreliable", "share_unreliable", "no_access"]

Q9_LABEL_MAP = {
    "public_transit":   "Bus or public transportation",
    "driving_self":     "Driving myself",
    "rides_from_others":"Getting rides from someone else",
    "rideshare":        "RideShare app (Lyft, Uber)",
    "active_transport": "Walking, biking, scooter, etc.",
    "other":            "Other",
}
Q9_ORDER = ["public_transit", "driving_self", "rides_from_others", "rideshare", "active_transport", "other"]
NO_CAR_SET = {"rideshare", "rides_from_others", "active_transport"}

Q7A_LABEL_MAP = {
    "vote_wont_matter": "My vote won't make a difference",
    "dont_understand":  "I do not understand politics",
    "dont_know_how":    "I do not know how to register",
    "not_old_enough":   "I am not old enough to vote",
    "other":            "Other",
}

Q15_LABEL_MAP = {
    "every_week":               "Every week",
    "1_3_times_per_month":      "1-3 times per month",
    "less_than_once_per_month": "Less than once per month",
    "never":                    "Never",
}
Q15_ORDER = ["every_week", "1_3_times_per_month", "less_than_once_per_month", "never"]
FREQUENT_VISIT = {"every_week", "1_3_times_per_month"}
INFREQUENT_VISIT = {"less_than_once_per_month", "never"}

Q15A_LABEL_MAP = {
    "see_coach_staff":   "See my Youth Coach or other Zone Staff",
    "work_on_goals":     "Work toward my goals",
    "safe_place":        "Be in a safe place",
    "computers":         "Access to computers",
    "laundry_shower":    "Do laundry, shower, etc.",
    "food":              "Eat food",
    "escape_problems":   "Escape problems/issues",
    "health_counseling": "Health care and/or counseling",
    "learn_skills":      "Learn new things, gain skills",
    "service_providers": "Meet with service providers",
    "socialize":         "Socialize/see friends",
    "scheduled_activity":"Participate in a scheduled activity",
    "other":             "Other",
}

Q15B_LABEL_MAP = {
    "better_activities": "More activities that interest me",
    "coach_invitation":  "Invitation from my Youth Coach",
    "more_info":         "Knowing more about activities",
    "other":             "Other",
}

Q17_LABEL_MAP = {
    "education":              "Finish or further my education",
    "job":                    "Get or keep a job",
    "housing":                "Find or maintain housing",
    "health_counseling":      "Access health care/counseling",
    "positive_relationships": "Establish positive relationships",
    "handle_problems":        "Figure out how to handle problems",
    "drivers_license":        "Get my driver's license",
    "parenting":              "Improve parenting skills",
    "everyday_skills":        "Learn everyday skills",
    "decision_making":        "Develop decision-making skills",
    "vital_documents":        "Obtain vital documents",
    "future":                 "Think about my future",
    "something_else":         "Something else",
}

Q25_LABEL_MAP = {
    "checking":    "Checking account",
    "savings":     "Savings account",
    "had_in_past": "Had an account in the past",
    "never_had":   "Never had an account",
}
Q25_ORDER = ["checking", "savings", "had_in_past", "never_had"]

Q24_LABEL_MAP = {
    "digital_apps":  "Venmo, Zelle, CashApp, etc.",
    "bank_account":  "Bank account",
    "cash_at_home":  "Saving/storing cash at home",
    "paypal":        "PayPal",
    "check_cashing": "Check cashing service (ex. ACE)",
    "money_order":   "Money Order",
    "other":         "Other",
}

Q26B_LABEL_MAP = {
    "debit_card":       "Bank credit/debit card",
    "online_banking":   "Online banking/Bank app",
    "direct_deposit":   "Direct deposit",
    "saving":           "Saving money",
    "budgeting":        "Budgeting money",
    "atm":              "ATMs",
    "keep_safe":        "Keep my money safe",
    "transferring":     "Transferring money",
    "cashing_checks":   "Cashing checks",
    "paying_bills":     "Paying household bills",
    "writing_checks":   "Writing checks",
    "in_person_banking":"Banking in person",
    "none":             "None - I don't use my account",
    "other":            "Other",
}

ORIENT_ORDER = [
    "Heterosexual/Straight", "Bisexual", "Gay, Lesbian, or Same Gender Loving",
    "Asexual", "Pansexual", "Queer", "Demisexual",
    "I am not sure yet", "I don't understand the question", "No answer provided",
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def split_pipe(val):
    if not str(val).strip():
        return []
    return [t.strip() for t in str(val).split("|") if t.strip()]


def pct_str(n, d, decimals=0):
    if d == 0:
        return ""
    val = round(100 * n / d, decimals)
    return f"{int(val)}%" if decimals == 0 else f"{val:.{decimals}f}%"


def age_label(code):
    return AGE_MAP.get(str(code).strip(), "Unknown")


def gender_group(g):
    g = str(g).strip()
    if g == "Female":
        return "Female"
    if g == "Male":
        return "Male"
    if not g:
        return "No answer"
    return "Trans, Non-binary"


def token_to_race_group(t):
    t = str(t).strip()
    if "Black" in t:
        return "Black"
    if "White" in t:
        return "White"
    if "Multi" in t:
        return "Multiracial"
    if "Hispanic" in t or "Latinx" in t:
        return "Hispanic or Latinx"
    if "Asian" in t:
        return "Asian"
    if "Native" in t:
        return "Native American or Native Hawaiian"
    if "Prefer not" in t:
        return "Prefer not to answer"
    return "Other"


def crosstab_age(df, field, label_map, age_col="_age", sort_by_total=True):
    """
    Count pipe-sep tokens in `field` by age group.
    Returns DataFrame: Label | 16-17... | 18-20... | 21-23... | Unknown | Total
    """
    rows = []
    for _, row in df.iterrows():
        for token in split_pipe(row[field]):
            rows.append({"_label": token, "_age": row[age_col]})
    if not rows:
        return pd.DataFrame()
    tmp = pd.DataFrame(rows)
    ct = tmp.groupby("_label")["_age"].value_counts().unstack(fill_value=0)
    for a in AGE_ORDER:
        if a not in ct.columns:
            ct[a] = 0
    ct = ct[AGE_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    ct = ct.reset_index()
    ct["Label"] = ct["_label"].map(label_map).fillna(ct["_label"])
    ct = ct.drop(columns=["_label"])
    ct = ct[["Label"] + AGE_ORDER + ["Total"]]
    if sort_by_total:
        ct = ct.sort_values("Total", ascending=False).reset_index(drop=True)
    return ct


def race_crosstab(df, multi_count=False):
    """
    Race x gender cross-tab.
    multi_count=False: each youth counted once (2+ tokens -> Multi-Racial).
    multi_count=True:  each token counted independently.
    """
    df2 = df[df["_gender"] != "No answer"].copy()

    # Groups collapsed into 'Other single racial identity' when counting each youth once
    _SMALL_GROUPS = {"Hispanic or Latinx", "Asian", "Native American or Native Hawaiian", "Other"}

    if not multi_count:
        def race_once(raw):
            tokens = split_pipe(raw)
            if not tokens:
                return None
            if len(tokens) >= 2:
                return "Multiracial"
            group = token_to_race_group(tokens[0])
            return "Other single racial identity" if group in _SMALL_GROUPS else group
        df2 = df2.copy()
        df2["_race"] = df2["race_ethnicity"].apply(race_once)
        df2 = df2.dropna(subset=["_race"])
        ct = pd.crosstab(df2["_race"], df2["_gender"])
    else:
        rows = []
        for _, row in df2.iterrows():
            for t in split_pipe(row["race_ethnicity"]):
                rows.append({"_race": token_to_race_group(t), "_gender": row["_gender"]})
        if not rows:
            return pd.DataFrame()
        tmp = pd.DataFrame(rows)
        ct = pd.crosstab(tmp["_race"], tmp["_gender"])

    for g in GENDER_ORDER:
        if g not in ct.columns:
            ct[g] = 0
    ct = ct[GENDER_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    n_total = ct["Total"].sum()
    ct["Percent"] = ct["Total"].apply(lambda n: pct_str(n, n_total))
    ct = ct.reset_index().rename(columns={"_race": "Race/Ethnicity"})

    present = [r for r in RACE_GROUP_ORDER if r in ct["Race/Ethnicity"].values]
    extra = [r for r in ct["Race/Ethnicity"].values if r not in RACE_GROUP_ORDER]
    ct = ct.set_index("Race/Ethnicity").reindex(present + extra).reset_index()

    total_row = {"Race/Ethnicity": "Total", "Percent": "100%"}
    for col in GENDER_ORDER + ["Total"]:
        total_row[col] = ct[col].sum()
    return pd.concat([ct, pd.DataFrame([total_row])], ignore_index=True)


# ---------------------------------------------------------------------------
# Section computations
# ---------------------------------------------------------------------------

def sec01_age(df):
    counts = df["_age"].value_counts().reindex(AGE_ORDER, fill_value=0)
    t = pd.DataFrame({"Age": counts.index, "Count": counts.values})
    return pd.concat([t, pd.DataFrame([{"Age": "Total", "Count": t["Count"].sum()}])],
                     ignore_index=True)


def sec02_gender_orient(df):
    df2 = df[df["_gender"] != "No answer"].copy()
    df2["_orient"] = df2["sexual_orientation"].apply(
        lambda x: x.strip() if x.strip() else "No answer provided"
    )
    # Merge Gay or Lesbian and Same Gender Loving (matches prior report format)
    df2["_orient"] = df2["_orient"].replace(
        {"Gay or Lesbian": "Gay, Lesbian, or Same Gender Loving",
         "Same Gender Loving": "Gay, Lesbian, or Same Gender Loving"}
    )
    known = df2["_orient"].unique().tolist()
    order = ORIENT_ORDER + [o for o in known if o not in ORIENT_ORDER]

    ct = pd.crosstab(df2["_orient"], df2["_gender"])
    for g in GENDER_ORDER:
        if g not in ct.columns:
            ct[g] = 0
    ct = ct[GENDER_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    ct = ct.reset_index().rename(columns={"_orient": "Sexual Orientation"})
    present = [o for o in order if o in ct["Sexual Orientation"].values]
    ct = ct.set_index("Sexual Orientation").reindex(present).reset_index()

    header = pd.DataFrame([{
        "Sexual Orientation": "Number of Youth",
        "Female": (df2["_gender"] == "Female").sum(),
        "Male": (df2["_gender"] == "Male").sum(),
        "Trans, Non-binary": (df2["_gender"] == "Trans, Non-binary").sum(),
        "Total": len(df2),
    }])
    return pd.concat([header, ct], ignore_index=True)


def sec03_race_once(df):
    return race_crosstab(df, multi_count=False)


def sec04_race_multi(df):
    return race_crosstab(df, multi_count=True)


def sec05_q1(df):
    TOP2 = {"4", "5"}
    rows = []
    for field, label in Q1_FIELDS:
        valid = df[field].astype(str)
        valid = valid[valid.str.strip() != ""]
        n_t2 = valid.isin(TOP2).sum()
        rows.append({
            "My Youth Coach...": label,
            "n": len(valid),
            "% Often or All the Time (4-5)": pct_str(n_t2, len(valid)),
        })
    return pd.DataFrame(rows)


def sec06_communication(df):
    Q2_LABEL_MAP = {
        "almost_every_day":       "Almost every day",
        "about_once_a_week":      "About once a week",
        "1_2_times_per_month":    "1-2 times per month",
        "less_than_once_a_month": "Less than once a month",
    }
    Q3_LABEL_MAP = {
        "not_enough": "Not enough",
        "good_amount": "Good amount",
        "too_much":    "Too much",
    }

    q3 = df["q3_communication_level"].replace("", pd.NA).dropna()
    total_q3 = len(q3)
    part_a = pd.DataFrame([{
        "Communication Level (Q3)": Q3_LABEL_MAP.get(k, k),
        "Count": v,
        "Percent": pct_str(v, total_q3),
    } for k, v in q3.value_counts().items()])

    not_enough = df[df["q3_communication_level"] == "not_enough"]
    q2 = not_enough["q2_communication_frequency"].replace("", pd.NA).dropna().value_counts()
    part_b = pd.DataFrame([{
        "Q2 Frequency (among 'Not Enough' responses)": Q2_LABEL_MAP.get(k, k),
        "Count": v,
    } for k, v in q2.items()])

    return {"Q3 Communication Level": part_a, "Q2 Frequency (Not Enough only)": part_b}


def sec07_housing(df):
    sleep_codes = []
    for v in df[df["q12_housing_stability"] != "stable"]["q13_sleeping_location"]:
        for t in split_pipe(v):
            if t not in sleep_codes:
                sleep_codes.append(t)
    ordered_sleep = [c for c in SLEEP_LABEL_MAP if c in sleep_codes]
    ordered_sleep += [c for c in sleep_codes if c not in SLEEP_LABEL_MAP]

    valid = df[df["q12_housing_stability"] != ""]
    total = len(valid)

    rows = []
    for code in HOUSING_ORDER:
        sub = valid[valid["q12_housing_stability"] == code]
        if sub.empty:
            continue
        row_data = {
            "Housing Status": HOUSING_LABEL_MAP.get(code, code),
            "Count": len(sub),
            "Percent": pct_str(len(sub), total),
        }
        for sc in ordered_sleep:
            cnt = sub["q13_sleeping_location"].apply(lambda v: sc in split_pipe(v)).sum()
            row_data[SLEEP_LABEL_MAP.get(sc, sc)] = cnt if cnt > 0 else ""
        rows.append(row_data)

    sleep_labels = [SLEEP_LABEL_MAP.get(sc, sc) for sc in ordered_sleep]
    total_row = {"Housing Status": "Total", "Count": total, "Percent": "100%"}
    for sl in sleep_labels:
        total_cnt = sum(int(r.get(sl, 0) or 0) for r in rows)
        total_row[sl] = total_cnt if total_cnt > 0 else ""
    rows.append(total_row)

    return pd.DataFrame(rows)


def sec08_housing_reasons(df):
    UNSTABLE = {"safe_not_90days", "90days_not_safe", "no_place"}

    stable_row = {"Reason": "No unstable housing in past 6 months"}
    for age in AGE_ORDER:
        stable_row[age] = int(((df["_age"] == age) & ~df["q12_housing_stability"].isin(UNSTABLE)).sum())
    stable_row["Total"] = int((~df["q12_housing_stability"].isin(UNSTABLE)).sum())

    unstable_row = {"Reason": "Total youth reporting unstable housing"}
    for age in AGE_ORDER:
        unstable_row[age] = int(((df["_age"] == age) & df["q12_housing_stability"].isin(UNSTABLE)).sum())
    unstable_row["Total"] = int(df["q12_housing_stability"].isin(UNSTABLE).sum())

    header_df = pd.DataFrame([stable_row, unstable_row])
    header_df = header_df.reindex(columns=["Reason"] + AGE_ORDER + ["Total"])

    ct = crosstab_age(df, "q14_housing_instability_reasons", Q14_LABEL_MAP)
    if not ct.empty:
        ct = ct.rename(columns={"Label": "Reason"})
        return pd.concat([header_df, ct], ignore_index=True)
    return header_df


def sec09_education(df):
    total_resp = df[df["q5_school_status"] != ""].shape[0]
    in_school = df[(df["q5_school_status"] != "") & (df["q5_school_status"] != "not_in_school")]
    not_school = df[df["q5_school_status"] == "not_in_school"]

    rows = []
    for code, label in [("graduate", "Graduate School"), ("college_career", "College/Vocational"),
                        ("ged", "GED Program"), ("high_school", "High School")]:
        n = (in_school["q5_school_status"] == code).sum()
        rows.append({
            "Group": "Currently Enrolled in School",
            "Level": label,
            "Count": n,
            "Percent of All": pct_str(n, total_resp),
        })

    rows.append({
        "Group": "Not in School",
        "Level": "Total Not in School",
        "Count": len(not_school),
        "Percent of All": pct_str(len(not_school), total_resp),
    })
    for code, label in [("college_degree", "College Degree or Certificate"),
                        ("some_college", "Some College"),
                        ("hs_diploma_ged", "HS Diploma or GED"),
                        ("some_hs", "Some High School")]:
        n = (not_school["q5a_highest_education"] == code).sum()
        if n > 0:
            rows.append({
                "Group": "Not in School",
                "Level": label,
                "Count": n,
                "Percent of All": pct_str(n, total_resp),
            })

    rows.append({"Group": "Total", "Level": "", "Count": total_resp, "Percent of All": "100%"})
    return pd.DataFrame(rows)


def sec10_employment(df):
    rows = []
    for code in Q8_ORDER:
        row_data = {"Employment Status": Q8_LABEL_MAP[code]}
        for age in AGE_ORDER:
            row_data[age] = int(((df["_age"] == age) & (df["q8_employment_status"] == code)).sum())
        row_data["Total"] = int((df["q8_employment_status"] == code).sum())
        rows.append(row_data)

        # Insert Seeking Employment sub-row after job training and after not working
        if code in ("job_training_program", "no"):
            seeking = (df["q8_employment_status"] == code) & (df["q8b_job_seeking"] == "yes")
            seek_row = {"Employment Status": "  Seeking Employment"}
            for age in AGE_ORDER:
                seek_row[age] = int(((df["_age"] == age) & seeking).sum())
            seek_row["Total"] = int(seeking.sum())
            rows.append(seek_row)

    total_row = {"Employment Status": "Total"}
    for age in AGE_ORDER:
        total_row[age] = int(((df["_age"] == age) & (df["q8_employment_status"] != "")).sum())
    total_row["Total"] = int((df["q8_employment_status"] != "").sum())
    rows.append(total_row)

    return pd.DataFrame(rows)


def sec11_job_tenure(df):
    employed = df[df["q8_employment_status"].isin(["yes_full_time", "yes_part_time"])]
    rows = []
    for code in Q8A_ORDER:
        ft = employed[(employed["q8a_job_tenure"] == code) & (employed["q8_employment_status"] == "yes_full_time")].shape[0]
        pt = employed[(employed["q8a_job_tenure"] == code) & (employed["q8_employment_status"] == "yes_part_time")].shape[0]
        rows.append({"Job Tenure": Q8A_LABEL_MAP[code], "Full time": ft, "Part time": pt, "Total": ft + pt})
    rows.append({
        "Job Tenure": "Total",
        "Full time": employed[employed["q8_employment_status"] == "yes_full_time"].shape[0],
        "Part time": employed[employed["q8_employment_status"] == "yes_part_time"].shape[0],
        "Total": len(employed),
    })
    return pd.DataFrame(rows)


def sec12_job_barriers(df):
    non_ft = df[df["q8_employment_status"] != "yes_full_time"]
    # Denominator = youth who actually answered Q10 (had job barriers to report)
    denom = int((non_ft["q10_job_barriers"].str.strip() != "").sum())

    counter = Counter()
    for v in non_ft["q10_job_barriers"]:
        for t in split_pipe(v):
            counter[t] += 1

    rows = []
    order = list(Q10_LABEL_MAP.keys()) + [k for k in counter if k not in Q10_LABEL_MAP]
    for code in order:
        n = counter.get(code, 0)
        if n > 0:
            rows.append({
                "Barrier": Q10_LABEL_MAP.get(code, code),
                "Total Youth": n,
                "Percent of Youth with Barriers": pct_str(n, denom),
            })
    rows.sort(key=lambda r: -r["Total Youth"])
    return pd.DataFrame(rows)


def sec13_left_job(df):
    total_left = df[df["q11_left_job_reasons"] != ""].shape[0]

    counter = Counter()
    for v in df["q11_left_job_reasons"]:
        for t in split_pipe(v):
            counter[t] += 1

    quit_counter = Counter()
    for v in df["q11a_quit_reasons"]:
        for t in split_pipe(v):
            quit_counter[t] += 1

    rows = []

    # Quit first, followed immediately by its Q11a sub-reasons (matches prior report format)
    quit_n = counter.get("quit", 0)
    rows.append({
        "Reason": Q11_LABEL_MAP.get("quit", "Quit"),
        "Total Youth": quit_n,
        "Percent Who Left a Job": pct_str(quit_n, total_left),
        "Source": "Q11",
    })
    quit_rows = []
    order_a = list(Q11A_LABEL_MAP.keys()) + [k for k in quit_counter if k not in Q11A_LABEL_MAP]
    for code in order_a:
        n = quit_counter.get(code, 0)
        if n > 0:
            quit_rows.append({
                "Reason": "  " + Q11A_LABEL_MAP.get(code, code),
                "Total Youth": n,
                "Percent Who Left a Job": pct_str(n, total_left),
                "Source": "Q11a",
            })
    quit_rows.sort(key=lambda r: -r["Total Youth"])
    rows.extend(quit_rows)

    # Remaining Q11 reasons (everything except quit), sorted by count
    other_rows = []
    order = [k for k in Q11_LABEL_MAP if k != "quit"] + [
        k for k in counter if k not in Q11_LABEL_MAP and k != "quit"
    ]
    for code in order:
        n = counter.get(code, 0)
        if n > 0:
            other_rows.append({
                "Reason": Q11_LABEL_MAP.get(code, code),
                "Total Youth": n,
                "Percent Who Left a Job": pct_str(n, total_left),
                "Source": "Q11",
            })
    other_rows.sort(key=lambda r: -r["Total Youth"])
    rows.extend(other_rows)

    return pd.DataFrame(rows)


def sec14_transport(df):
    # Table A: Driver's license x age
    valid_q6 = df[df["q6_drivers_license"] != ""]
    n_per_age = {age: (valid_q6["_age"] == age).sum() for age in AGE_ORDER}
    n_total = len(valid_q6)

    rows_a = [{"Driver's License": "Number of Youth",
               **{age: n_per_age[age] for age in AGE_ORDER}, "Total": n_total}]
    for code in Q6_ORDER:
        sub = valid_q6[valid_q6["q6_drivers_license"] == code]
        row_data = {"Driver's License": Q6_LABEL_MAP[code]}
        for age in AGE_ORDER:
            row_data[age] = pct_str((sub["_age"] == age).sum(), n_per_age[age])
        row_data["Total"] = pct_str(len(sub), n_total)
        rows_a.append(row_data)
    rows_a.append({"Driver's License": "Total",
                   **{age: "100%" if n_per_age[age] > 0 else "" for age in AGE_ORDER},
                   "Total": "100%"})
    table_a = pd.DataFrame(rows_a)

    # Table B: Vehicle access x age (license holders)
    licensed = df[df["q6_drivers_license"] == "yes"]
    rows_b = []
    for code in Q6A_ORDER:
        sub = licensed[licensed["q6a_vehicle_access"] == code]
        if sub.empty:
            continue
        row_data = {"Vehicle Access (license holders)": Q6A_LABEL_MAP.get(code, code)}
        for age in AGE_ORDER:
            row_data[age] = (sub["_age"] == age).sum()
        row_data["Total"] = len(sub)
        rows_b.append(row_data)
    rows_b.append({"Vehicle Access (license holders)": "Total",
                   **{age: (licensed["_age"] == age).sum() for age in AGE_ORDER},
                   "Total": len(licensed)})
    table_b = pd.DataFrame(rows_b)

    # Table C: Primary transport
    counter = Counter(v.strip() for v in df["q9_primary_transport"] if v.strip())
    total_t = sum(counter.values())

    rows_c = []
    all_codes = Q9_ORDER + [k for k in counter if k not in Q9_ORDER]
    for code in all_codes:
        n = counter.get(code, 0)
        if n == 0:
            continue
        rows_c.append({
            "Primary Transport": Q9_LABEL_MAP.get(code, code),
            "Total": n,
            "Percent": pct_str(n, total_t),
        })
    rows_c.sort(key=lambda r: -r["Total"])
    no_car_n = sum(v for k, v in counter.items() if k in NO_CAR_SET)
    if no_car_n > 0:
        rows_c.append({"Primary Transport": "No-Car Combination (non-bus)",
                       "Total": no_car_n, "Percent": pct_str(no_car_n, total_t)})
    rows_c.append({"Primary Transport": "Total", "Total": total_t, "Percent": "100%"})
    table_c = pd.DataFrame(rows_c)

    return {
        "Driver's License by Age": table_a,
        "Vehicle Access (licensed)": table_b,
        "Primary Transport": table_c,
    }


def sec15_voter_reg(df):
    voter_ages = ["18-20 years old", "21-23 years old"]
    vdf = df[df["_age"].isin(voter_ages)].copy()
    n_per_age = {age: (vdf[(vdf["_age"] == age) & (vdf["q7_registered_to_vote"] != "")]).shape[0]
                 for age in voter_ages}
    n_total = sum(n_per_age.values())

    rows_a = [{"Status": "n =", **{age: n_per_age[age] for age in voter_ages}, "Total": n_total}]
    for code, label in [("yes", "Registered to Vote"), ("no", "Not Registered")]:
        row_data = {"Status": label}
        for age in voter_ages:
            n = (vdf[(vdf["_age"] == age) & (vdf["q7_registered_to_vote"] == code)]).shape[0]
            row_data[age] = pct_str(n, n_per_age[age])
        total_n = (vdf["q7_registered_to_vote"] == code).sum()
        row_data["Total"] = pct_str(total_n, n_total)
        rows_a.append(row_data)
    table_a = pd.DataFrame(rows_a)

    not_reg = vdf[vdf["q7_registered_to_vote"] == "no"]
    counter_by_age = {age: Counter() for age in voter_ages}
    counter_all = Counter()
    for _, row in not_reg.iterrows():
        for t in split_pipe(row["q7a_not_registered_reasons"]):
            counter_by_age.get(row["_age"], Counter())[t] += 1
            counter_all[t] += 1

    rows_b = []
    order = list(Q7A_LABEL_MAP.keys()) + [k for k in counter_all if k not in Q7A_LABEL_MAP]
    for code in order:
        n = counter_all.get(code, 0)
        if n == 0:
            continue
        row_data = {"Reason Not Registered": Q7A_LABEL_MAP.get(code, code)}
        for age in voter_ages:
            row_data[age] = counter_by_age[age].get(code, 0)
        row_data["Total"] = n
        rows_b.append(row_data)
    rows_b.sort(key=lambda r: -r["Total"])
    table_b = pd.DataFrame(rows_b)

    return {"Voter Registration by Age": table_a, "Not Registered Reasons by Age": table_b}


def sec16_visit(df):
    rows_a = []
    for code in Q15_ORDER:
        row_data = {"Visit Frequency": Q15_LABEL_MAP[code]}
        for age in AGE_ORDER:
            row_data[age] = int(((df["_age"] == age) & (df["q15_visit_frequency"] == code)).sum())
        row_data["Total"] = int((df["q15_visit_frequency"] == code).sum())
        rows_a.append(row_data)
    total_row = {"Visit Frequency": "Total"}
    for age in AGE_ORDER:
        total_row[age] = int(((df["_age"] == age) & (df["q15_visit_frequency"] != "")).sum())
    total_row["Total"] = int((df["q15_visit_frequency"] != "").sum())
    rows_a.append(total_row)
    table_a = pd.DataFrame(rows_a)

    frequent = df[df["q15_visit_frequency"].isin(FREQUENT_VISIT)]
    ct_b = crosstab_age(frequent, "q15a_visit_reasons", Q15A_LABEL_MAP)
    if not ct_b.empty:
        ct_b = ct_b.rename(columns={"Label": "Visit Reason (frequent visitors)"})
    else:
        ct_b = None

    infreq = df[df["q15_visit_frequency"].isin(INFREQUENT_VISIT)]
    ct_c = crosstab_age(infreq, "q15b_visit_barriers", Q15B_LABEL_MAP)
    if not ct_c.empty:
        ct_c = ct_c.rename(columns={"Label": "Visit Barrier (infrequent/never visitors)"})
        # Add header row showing total respondents per age (matches prior report format)
        total_row = {"Visit Barrier (infrequent/never visitors)": "Total respondents"}
        for age in AGE_ORDER:
            total_row[age] = int((infreq["_age"] == age).sum())
        total_row["Total"] = len(infreq)
        ct_c = pd.concat([pd.DataFrame([total_row]), ct_c], ignore_index=True)
    else:
        ct_c = None

    return {
        "Visit Frequency by Age": table_a,
        "Visit Reasons (frequent)": ct_b,
        "Visit Barriers (infrequent)": ct_c,
    }


def sec17_program_impact(df):
    ct_a = crosstab_age(df, "q17_program_helped", Q17_LABEL_MAP)
    if not ct_a.empty:
        ct_a = ct_a.rename(columns={"Label": "Program Helped With"})
    else:
        ct_a = None

    Q16_LABEL_MAP = {"agree": "Agree", "somewhat_agree": "Somewhat agree",
                     "disagree": "Disagree", "unsure": "Unsure"}
    q16 = df["q16_stay_focused"].replace("", pd.NA).dropna().value_counts()
    total_q16 = q16.sum()
    table_b = pd.DataFrame([{
        "Q16 - Coach/Zone helps stay focused": Q16_LABEL_MAP.get(k, k),
        "Count": v, "Percent": pct_str(v, total_q16),
    } for k, v in q16.items()])

    Q21_LABEL_MAP = {"agree": "Agree", "somewhat": "Somewhat agree",
                     "disagree": "Disagree", "unsure": "Unsure"}
    q21 = df["q21_gained_independence"].replace("", pd.NA).dropna().value_counts()
    total_q21 = q21.sum()
    table_c = pd.DataFrame([{
        "Q21 - Gained independence": Q21_LABEL_MAP.get(k, k),
        "Count": v, "Percent": pct_str(v, total_q21),
    } for k, v in q21.items()])

    return {
        "Program Helped With (Q17) by Age": ct_a,
        "Stay Focused (Q16)": table_b,
        "Gained Independence (Q21)": table_c,
    }


def sec18_respect(df):
    RESP_ORDER = ["never", "rarely", "sometimes", "often", "all_the_time"]
    RESP_LABELS = {"never": "Never", "rarely": "Rarely", "sometimes": "Sometimes",
                   "often": "Often", "all_the_time": "All the time"}
    TOP2 = {"often", "all_the_time"}

    rows = []
    for field, label in [
        ("q18_staff_respect", "Staff treat me with respect and acceptance"),
        ("q19_peer_respect",  "Peers treat me with respect and acceptance"),
    ]:
        col = df[field].replace("", pd.NA).dropna()
        n_t2 = col.isin(TOP2).sum()
        row_data = {
            "Statement": label,
            "n": len(col),
            "% Often or All the Time": pct_str(n_t2, len(col)),
        }
        for code in RESP_ORDER:
            row_data[RESP_LABELS[code]] = (col == code).sum()
        rows.append(row_data)
    return pd.DataFrame(rows)


def sec19_environment(df):
    TOP2 = {"4", "5"}
    rows = []
    for field, label in Q20_FIELDS:
        valid = df[field].astype(str)
        valid = valid[valid.str.strip() != ""]
        n_t2 = valid.isin(TOP2).sum()
        rows.append({
            "The Zone is a place where...": label,
            "n": len(valid),
            "% Top-2 Box (4-5)": pct_str(n_t2, len(valid)),
        })
    return pd.DataFrame(rows)


def sec20_banking(df):
    has_account = df["q25_bank_account"].apply(
        lambda v: any(t in split_pipe(v) for t in ["checking", "savings"])
    )

    # Table A: bank account status x age
    rows_a = [{"Account Status": "Number of Youth",
               **{age: int((df["_age"] == age).sum()) for age in AGE_ORDER},
               "Total": len(df), "Percent of Total": ""}]
    rows_a.append({
        "Account Status": "Currently have a bank account",
        **{age: int(has_account[df["_age"] == age].sum()) for age in AGE_ORDER},
        "Total": int(has_account.sum()),
        "Percent of Total": pct_str(has_account.sum(), len(df)),
    })
    for code in Q25_ORDER:
        sub = df["q25_bank_account"].apply(lambda v: code in split_pipe(v))
        n = int(sub.sum())
        if n == 0:
            continue
        rows_a.append({
            "Account Status": Q25_LABEL_MAP[code],
            **{age: int(sub[df["_age"] == age].sum()) for age in AGE_ORDER},
            "Total": n,
            "Percent of Total": pct_str(n, len(df)),
        })
    table_a = pd.DataFrame(rows_a)

    # Table B: money methods x age
    ct_b = crosstab_age(df, "q24_money_methods", Q24_LABEL_MAP)
    if not ct_b.empty:
        ct_b = ct_b.rename(columns={"Label": "Money Method (Q24)"})
        ct_b["Percent of All"] = ct_b["Total"].apply(lambda n: pct_str(n, len(df)))
    else:
        ct_b = None

    # Table C: account usage x age (account holders only)
    has_acc_df = df[has_account]
    ct_c = crosstab_age(has_acc_df, "q26b_account_usage", Q26B_LABEL_MAP)
    if not ct_c.empty:
        ct_c = ct_c.rename(columns={"Label": "Account Usage (Q26b)"})
        ct_c["% of Account Holders"] = ct_c["Total"].apply(
            lambda n: pct_str(n, len(has_acc_df))
        )
    else:
        ct_c = None

    return {
        "Bank Account Status by Age": table_a,
        "Money Methods by Age (Q24)": ct_b,
        "Account Usage by Age (Q26b)": ct_c,
    }


def sec21_nps(df):
    nps_raw = df["q22_nps"].replace("", pd.NA).dropna()
    try:
        nps_int = nps_raw.astype(float).astype(int)
    except Exception:
        nps_int = pd.Series([], dtype=int)

    total = len(nps_int)
    promoters = int((nps_int >= 9).sum())
    passives   = int(((nps_int >= 7) & (nps_int <= 8)).sum())
    detractors = int((nps_int <= 6).sum())
    nps_score = round(100 * promoters / total - 100 * detractors / total) if total > 0 else ""

    return pd.DataFrame([
        {"Category": "Promoters (9-10)", "Count": promoters, "Percent": pct_str(promoters, total)},
        {"Category": "Passives (7-8)",   "Count": passives,  "Percent": pct_str(passives, total)},
        {"Category": "Detractors (0-6)", "Count": detractors,"Percent": pct_str(detractors, total)},
        {"Category": "Total Responded",  "Count": total,     "Percent": ""},
        {"Category": "NPS Score",        "Count": nps_score, "Percent": ""},
    ])


def sec22_comments(df):
    comments = df[df["q23_other_comments"].str.strip() != ""][
        ["survey_id", "q23_other_comments"]
    ].copy()
    comments.columns = ["Survey ID", "Comment"]
    return comments.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

TITLE_FONT  = Font(bold=True, color="FFFFFF", size=12)
TITLE_FILL  = PatternFill("solid", fgColor="2F5496")
SUB_FONT    = Font(bold=True, italic=True)
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")


def write_section(ws, title, tables):
    """Write title then one or more DataFrames to worksheet."""
    if isinstance(tables, pd.DataFrame):
        tables = [("", tables)]
    elif isinstance(tables, dict):
        tables = list(tables.items())

    row = 1
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    row += 1

    for subtitle, df in tables:
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            continue
        if subtitle:
            ws.cell(row=row, column=1, value=subtitle).font = SUB_FONT
            row += 1
        for col_idx, col_name in enumerate(df.columns, 1):
            c = ws.cell(row=row, column=col_idx, value=str(col_name))
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        row += 1
        for _, data_row in df.iterrows():
            for col_idx, val in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=val)
            row += 1
        row += 1  # blank row between tables


def autofit_columns(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 55)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

SECTIONS = [
    ("01_age",            "1. Age Distribution",                         sec01_age),
    ("02_gender_orient",  "2. Gender x Sexual Orientation",              sec02_gender_orient),
    ("03_race_once",      "3. Race/Ethnicity (Counted Once)",            sec03_race_once),
    ("04_race_multi",     "4. Race/Ethnicity (Counted Multiple Times)",  sec04_race_multi),
    ("05_q1",             "5. Coach Satisfaction (Q1)",                  sec05_q1),
    ("06_communication",  "6. Communication Frequency & Level (Q2/Q3)", sec06_communication),
    ("07_housing",        "7. Housing Stability (Q12 x Q13)",            sec07_housing),
    ("08_housing_reasons","8. Housing Instability Reasons (Q14 x Age)",  sec08_housing_reasons),
    ("09_education",      "9. Education (Q5 + Q5a)",                    sec09_education),
    ("10_employment",     "10. Employment Status (Q8 x Age)",            sec10_employment),
    ("11_job_tenure",     "11. Job Tenure (Q8a x Full/Part-time)",       sec11_job_tenure),
    ("12_job_barriers",   "12. Job Barriers (Q10)",                      sec12_job_barriers),
    ("13_left_job",       "13. Reasons Left Job (Q11 + Q11a)",           sec13_left_job),
    ("14_transport",      "14. Transportation (Q6 + Q6a + Q9)",          sec14_transport),
    ("15_voter_reg",      "15. Voter Registration (Q7 + Q7a)",           sec15_voter_reg),
    ("16_visit",          "16. Visit Frequency + Reasons + Barriers",    sec16_visit),
    ("17_impact",         "17. Program Impact (Q17 + Q16 + Q21)",        sec17_program_impact),
    ("18_respect",        "18. Staff & Peer Respect (Q18 + Q19)",        sec18_respect),
    ("19_environment",    "19. Program Environment (Q20)",               sec19_environment),
    ("20_banking",        "20. Banking (Q25 + Q24 + Q26b)",              sec20_banking),
    ("21_nps",            "21. NPS (Q22)",                               sec21_nps),
    ("22_comments",       "22. Additional Comments (Q23)",               sec22_comments),
]


def main():
    if not CSV_PATH.exists():
        print(f"CSV not found: {CSV_PATH}")
        sys.exit(1)

    df = pd.read_csv(str(CSV_PATH), encoding="utf-8-sig", dtype=str).fillna("")
    df["_age"]    = df["age_range"].apply(age_label)
    df["_gender"] = df["gender"].apply(gender_group)

    print(f"Loaded {len(df)} surveys.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, title, func in SECTIONS:
        print(f"  {sheet_name}...")
        result = func(df)
        ws = wb.create_sheet(sheet_name)
        write_section(ws, title, result)
        autofit_columns(ws)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))
    print(f"\nSaved: {OUT_PATH}")


if __name__ == "__main__":
    main()

"""
03b_apply_corrections_IL.py
Read reviewer decisions from output/IL/qa_questions_IL.xlsx
and apply them to output/IL/survey_data_IL.csv.

Actions:
  recode  - replace current_value with corrected_value
  clear   - blank the field (scalar) or remove the flagged token (array)
  accept  - no change (skipped)

apply_to rules:
  this_survey  - change only the specified survey_id row
  all_surveys  - change every row where field contains current_value
  (blank)      - treated as this_survey

For array (pipe-separated) fields, current_value is the specific token to
act on; other tokens in the same cell are left untouched.

Overwrites survey_data_IL.csv in place and prints a change log.

Usage:
    python scripts/03b_apply_corrections_IL.py
"""

import sys
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).parent.parent

# Fields stored as pipe-separated arrays in the CSV
ARRAY_FIELDS = {
    "q6b_job_types",
    "q7_barriers",
    "q8_left_job_reasons",
    "q8a_quit_reasons",
    "q9_bank_account",
    "q9a_no_account_reasons",
    "q11_program_helped",
    "race_ethnicity",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def split_pipe(val) -> list[str]:
    if val is None or str(val).strip() == "":
        return []
    return [token.strip() for token in str(val).split("|") if token.strip()]


def join_pipe(tokens: list[str]) -> str:
    return " | ".join(tokens)


def normalize_cell(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def dedupe_tokens(tokens: list[str]) -> list[str]:
    seen = set()
    unique = []
    for token in tokens:
        if token not in seen:
            seen.add(token)
            unique.append(token)
    return unique


# ---------------------------------------------------------------------------
# Load corrections from workbook
# ---------------------------------------------------------------------------

def load_corrections(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    corrections = []
    for sheet_name in ("QA Questions", "Accepted - No Action"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            rec = dict(zip(headers, row))
            action = str(rec.get("action") or "").strip().lower()
            if action == "exclude":
                action = "clear"
                rec["action"] = "clear"
            if not action or action == "accept":
                continue
            corrections.append(rec)

    return corrections


# ---------------------------------------------------------------------------
# Apply corrections
# ---------------------------------------------------------------------------

def apply_all(df: pd.DataFrame, corrections: list[dict]) -> tuple[pd.DataFrame, list[str]]:
    log = []

    for rec in corrections:
        survey_id = str(rec.get("survey_id") or "").strip()
        field = str(rec.get("field") or "").strip()
        current_val = str(rec.get("current_value") or "").strip()
        action = str(rec.get("action") or "").strip().lower()
        corrected = str(rec.get("corrected_value") or "").strip()
        scope_raw = str(rec.get("apply_to") or "").strip().lower()
        apply_to = "all_surveys" if scope_raw == "all_surveys" else "this_survey"

        if field not in df.columns:
            log.append(f"  [SKIP]   {survey_id} / {field}: column not in CSV")
            continue

        is_array = field in ARRAY_FIELDS
        survey_mask = df["survey_id"] == survey_id
        mask = pd.Series(True, index=df.index) if apply_to == "all_surveys" else survey_mask
        changed = 0

        if is_array:
            def fix_token(cell_val, cur=current_val, act=action, corr=corrected):
                cell_text = normalize_cell(cell_val)
                if act == "clear" and cur and cell_text == cur:
                    return ""
                if act == "recode" and cur and cell_text == cur:
                    return corr

                tokens = split_pipe(cell_val)
                if cur not in tokens:
                    return cell_val
                if act == "clear":
                    tokens = [token for token in tokens if token != cur]
                elif act == "recode":
                    tokens = [corr if token == cur else token for token in tokens]
                return join_pipe(dedupe_tokens(tokens))

            before = df.loc[mask, field].copy()
            df.loc[mask, field] = df.loc[mask, field].apply(fix_token)
            changed = (df.loc[mask, field] != before).sum()

        else:
            if action == "clear":
                if apply_to == "all_surveys":
                    rows = mask & (df[field] == current_val)
                else:
                    rows = mask
                before = df.loc[rows, field].copy()
                df.loc[rows, field] = ""
                changed = (df.loc[rows, field] != before).sum()

            elif action == "recode":
                if apply_to == "all_surveys":
                    rows = mask & (df[field] == current_val)
                else:
                    rows = mask
                before = df.loc[rows, field].copy()
                df.loc[rows, field] = corrected
                changed = (df.loc[rows, field] != before).sum()

        scope_label = "all surveys" if apply_to == "all_surveys" else survey_id
        new_val_label = f"'{corrected}'" if action == "recode" else "blank"
        log.append(
            f"  [{action.upper():<7}] {field}: '{current_val}' -> {new_val_label}"
            f"  ({scope_label}, {changed} row(s))"
        )

    return df, log


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    csv_path = BASE_DIR / "output" / "IL" / "survey_data_IL.csv"
    xlsx_path = BASE_DIR / "output" / "IL" / "qa_questions_IL.xlsx"

    if not csv_path.exists():
        print(f"CSV not found: {csv_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}")
        sys.exit(1)

    df = pd.read_csv(str(csv_path), encoding="utf-8-sig", dtype=str).fillna("")
    corrections = load_corrections(xlsx_path)

    print(f"Loaded {len(df)} surveys. Applying {len(corrections)} correction(s)...\n")

    df, log = apply_all(df, corrections)

    df.to_csv(str(csv_path), index=False, encoding="utf-8-sig")

    print("Change log:")
    for entry in log:
        print(entry)
    print(f"\nSaved: {csv_path}")


if __name__ == "__main__":
    main()
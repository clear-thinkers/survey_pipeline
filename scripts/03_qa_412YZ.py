"""
03_qa_412YZ.py
Run QA validation on output/412YZ/survey_data_412YZ.csv.

Produces two outputs:
  output/412YZ/flagged_412YZ.csv         — machine-readable issue log
  output/412YZ/qa_questions_412YZ.xlsx   — reviewer workbook: one row per issue,
                                           colour-coded by category, with blank
                                           Action / Corrected Value / Apply To columns
                                           for the reviewer to fill in before running
                                           03b_apply_corrections_412YZ.py.

Usage:
    python scripts/03_qa_412YZ.py
"""

import sys
from collections import Counter
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent.parent))

BASE_DIR = Path(__file__).parent.parent

# ---------------------------------------------------------------------------
# Allowed-value definitions (from extraction_prompt_412YZ.txt)
# ---------------------------------------------------------------------------

# Single-select fields
SINGLE_SELECT = {
    "q2_communication_frequency": {
        "almost_every_day", "about_once_a_week", "1_2_times_per_month", "less_than_once_a_month"
    },
    "q3_communication_level": {"not_enough", "good_amount", "too_much"},
    "q4_program_duration":    {"less_6mo", "6_12mo", "1_3yr", "3yr_plus"},
    "q5_school_status":       {"high_school", "ged", "college_career", "graduate", "not_in_school"},
    "q5a_highest_education":  {"some_hs", "hs_diploma_ged", "some_college", "college_degree", "graduate"},
    "q6_drivers_license":     {"yes", "learners_permit", "no"},
    "q6a_vehicle_access":     {"own_reliable", "own_unreliable", "share_reliable", "share_unreliable", "borrow", "no_access"},
    "q7_registered_to_vote":  {"yes", "no"},
    "q8_employment_status":   {"yes_part_time", "yes_full_time", "no", "job_training_program"},
    "q8a_job_tenure":         {"less_3mo", "3_6mo", "more_6mo"},
    "q8b_job_seeking":        {"yes", "no", "not_yet_in_training"},
    "q9_primary_transport":   {"public_transit", "driving_self", "rides_from_others", "rideshare", "active_transport", "other"},
    "q12_housing_stability":  {"stable", "safe_not_90days", "90days_not_safe", "no_place"},
    "q15_visit_frequency":    {"every_week", "1_3_times_per_month", "less_than_once_per_month", "never"},
    "q16_stay_focused":       {"agree", "somewhat_agree", "disagree", "unsure"},
    "q18_staff_respect":      {"never", "rarely", "sometimes", "often", "all_the_time"},
    "q19_peer_respect":       {"never", "rarely", "sometimes", "often", "all_the_time"},
    "q21_gained_independence": {"agree", "somewhat", "disagree", "unsure"},
    "age_range":              {"16_17", "18_20", "21_23"},
}

# Multi-select (pipe-separated) fields: each token validated against allowed codes
MULTI_SELECT = {
    "q13_sleeping_location": {
        "friends_family", "shelter", "couch_surfing", "car", "outside", "abandoned_building", "other"
    },
    "q7a_not_registered_reasons": {
        "not_old_enough", "dont_know_how", "dont_understand", "vote_wont_matter", "other"
    },
    "q10_job_barriers": {
        "childcare", "criminal_background", "no_references", "interview_skills", "no_diploma",
        "limited_experience", "mental_physical_health", "transportation", "drugs_alcohol",
        "not_getting_called", "something_else"
    },
    "q11_left_job_reasons": {
        "found_better", "quit", "fired_attendance", "fired_performance",
        "seasonal", "pregnancy_parenting", "other"
    },
    "q11a_quit_reasons": {
        "low_pay_hours", "schedule_conflict", "lack_of_support", "poor_conditions",
        "mental_emotional_health", "transportation", "not_good_fit", "personal_family", "other"
    },
    "q14_housing_instability_reasons": {
        "evicted_nonpayment", "evicted_other", "lost_informal_housing", "left_unsafe", "other"
    },
    "q15a_visit_reasons": {
        "computers", "safe_place", "laundry_shower", "food", "escape_problems",
        "health_counseling", "learn_skills", "service_providers", "see_coach_staff",
        "socialize", "work_on_goals", "scheduled_activity", "other"
    },
    "q15b_visit_barriers": {"coach_invitation", "more_info", "better_activities", "other"},
    "q17_program_helped": {
        "health_counseling", "positive_relationships", "handle_problems", "housing",
        "education", "job", "drivers_license", "parenting", "everyday_skills",
        "decision_making", "vital_documents", "future", "something_else"
    },
    "q24_money_methods": {
        "bank_account", "check_cashing", "digital_apps", "paypal",
        "money_order", "cash_at_home", "other"
    },
    "q25_bank_account": {"checking", "savings", "had_in_past", "never_had"},
    "q26a_account_setup": {
        "self_online", "self_inperson", "self_with_help", "added_by_other", "other"
    },
    "q26b_account_usage": {
        "budgeting", "saving", "cashing_checks", "writing_checks", "keep_safe",
        "transferring", "direct_deposit", "debit_card", "online_banking",
        "atm", "in_person_banking", "paying_bills", "none", "other"
    },
}

LIKERT_FIELDS = [
    "q1_trustworthy", "q1_reliable", "q1_values_opinions", "q1_available", "q1_heard_understood",
    "q20_people_care", "q20_no_judgment", "q20_diversity_valued", "q20_treated_fairly", "q20_safe_sharing",
]

REQUIRED_FIELDS = [
    "q1_trustworthy", "q2_communication_frequency", "q3_communication_level",
    "q4_program_duration", "q5_school_status", "q8_employment_status",
    "q12_housing_stability", "q15_visit_frequency", "q16_stay_focused",
    "q18_staff_respect", "q19_peer_respect", "q21_gained_independence",
    "q22_nps", "age_range", "gender",
]

# Conditional fields: child → (parent_field, set_of_parent_values_that_allow_child)
# q11a, q26a, q26b have array-valued parents and are handled separately below.
CONDITIONAL = {
    "q5a_highest_education":      ("q5_school_status",     {"not_in_school"}),
    "q6a_vehicle_access":         ("q6_drivers_license",   {"yes"}),
    "q7a_not_registered_reasons": ("q7_registered_to_vote", {"no"}),
    "q8a_job_tenure":             ("q8_employment_status", {"yes_part_time", "yes_full_time"}),
    "q8b_job_seeking":            ("q8_employment_status", {"no", "job_training_program"}),
    "q13_sleeping_location":      ("q12_housing_stability", {"safe_not_90days", "90days_not_safe", "no_place"}),
    "q15a_visit_reasons":         ("q15_visit_frequency",  {"every_week", "1_3_times_per_month"}),
    "q15b_visit_barriers":        ("q15_visit_frequency",  {"less_than_once_per_month", "never"}),
    "q16a_what_would_help":       ("q16_stay_focused",     {"somewhat_agree", "disagree"}),
}

# Known printed labels for free-text demographic fields
KNOWN_RACE_LABELS = {
    "Black or of African or Caribbean Descent",
    "White or of European Descent",
    "Multi-Racial",
    "Hispanic or Latinx",
    "East Asian", "Asian",
    "Native American", "Native Hawaiian", "Native American or Native Hawaiian",
    "Native American or Indigenous peoples of America",
    "Native Hawaiian or Pacific Islander",
    "South Asian or Indian (Subcontinent)",
    "Southeast Asian",
    "Western Asian or Middle Eastern",
    "Other Asian",
    "I prefer not to answer", "Prefer not to answer",
}

KNOWN_GENDER_LABELS = {
    "Female", "Male", "Non-binary",
    "Transgender Male", "Transgender Female",
    "Genderqueer", "Two-Spirit", "Gender Nonconforming",
    "Trans-Non-binary",
    "I prefer not to say", "Prefer not to answer",
}

STANDARD_GENDER_VALUES = [
    "Female",
    "Male",
    "Non-binary",
    "Transgender Male",
    "Transgender Female",
    "Genderqueer",
    "Two-Spirit",
    "Gender Nonconforming",
    "Prefer not to answer",
]

KNOWN_ORIENTATION_LABELS = {
    "Heterosexual", "Heterosexual/Straight", "Gay or Lesbian", "Bisexual", "Asexual",
    "Pansexual", "Queer", "Demisexual", "Mostly heterosexual", "I am not sure yet",
    "I don't understand the question", "I prefer not to answer",
    "Same Gender Loving",
}

STANDARD_RACE_VALUES = [
    "Black or of African or Caribbean Descent",
    "White or of European Descent",
    "Multi-Racial",
    "Hispanic or Latinx",
    "East Asian",
    "Native American or Indigenous peoples of America",
    "Native Hawaiian or Pacific Islander",
    "South Asian or Indian (Subcontinent)",
    "Southeast Asian",
    "Western Asian or Middle Eastern",
    "Other Asian",
    "Prefer not to answer",
]

STANDARD_ORIENTATION_VALUES = [
    "Heterosexual/Straight",
    "Gay or Lesbian",
    "Bisexual",
    "Asexual",
    "Pansexual",
    "Queer",
    "Demisexual",
    "Mostly heterosexual",
    "I am not sure yet",
    "I don't understand the question",
    "Same Gender Loving",
    "I prefer not to answer",
]

# Demographic normalization defaults.
# Each entry becomes a prefilled QA row so the reviewer only needs to confirm it.
GENDER_NORMALIZATION = {
    "Prefer not to say": {
        "normalized": "Prefer not to answer",
        "scope": "all_surveys",
        "detail": "alternate wording for the printed nonresponse option",
    },
    "Female | Straight": {
        "normalized": "Female",
        "scope": "this_survey",
        "detail": "gender value includes an orientation label; keeping the gender selection only",
    },
}

RACE_NORMALIZATION = {
    "Black": {
        "normalized": "Black or of African or Caribbean Descent",
        "scope": "all_surveys",
        "detail": "shorthand variant of the printed race label",
    },
    "White": {
        "normalized": "White or of European Descent",
        "scope": "all_surveys",
        "detail": "shorthand variant of the printed race label",
    },
    "Biracial": {
        "normalized": "Multi-Racial",
        "scope": "all_surveys",
        "detail": "common alternate wording for the printed multiracial label",
    },
    "Black African": {
        "normalized": "Black or of African or Caribbean Descent",
        "scope": "this_survey",
        "detail": "self-described race maps directly to the closest printed Black/African label",
    },
    "I don't know what to put in white": {
        "normalized": "White or of European Descent",
        "scope": "this_survey",
        "detail": "free-text response explicitly indicates White as the closest printed race label",
    },
}

ORIENTATION_NORMALIZATION = {
    "Straight": {
        "normalized": "Heterosexual/Straight",
        "scope": "all_surveys",
        "detail": "common shorthand for the printed label",
    },
    "Heterosexual": {
        "normalized": "Heterosexual/Straight",
        "scope": "all_surveys",
        "detail": "truncated variant of the printed label",
    },
    "Demi-Sexual": {
        "normalized": "Demisexual",
        "scope": "all_surveys",
        "detail": "hyphenated spelling variant",
    },
    "Demi Sexye": {
        "normalized": "Demisexual",
        "scope": "all_surveys",
        "detail": "likely OCR misspelling of Demisexual",
    },
    "Decline to answer": {
        "normalized": "I prefer not to answer",
        "scope": "all_surveys",
        "detail": "alternate wording for the printed nonresponse option",
    },
    "I don't know what I am I just say I'm bisexual because I'm still exploring": {
        "normalized": "Bisexual",
        "scope": "this_survey",
        "detail": "respondent explicitly says they identify as bisexual despite exploratory wording",
    },
    "Achillean and minsexual": {
        "normalized": "Queer",
        "scope": "this_survey",
        "detail": "non-standard queer-spectrum identity best grouped under the printed umbrella label",
    },
}

DEMOGRAPHIC_PLACEHOLDERS = {
    "self-describe",
    "self-describe:",
    "self describe",
    "self describe:",
}

LOW_CONF_THRESHOLD = 0.75

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def split_pipe(val) -> list[str]:
    if val is None or str(val).strip() == "":
        return []
    return [t.strip() for t in str(val).split("|") if t.strip()]


def is_blank(val) -> bool:
    return val is None or str(val).strip() == ""


def has_value(row, field: str) -> bool:
    v = row.get(field, "")
    if is_blank(v):
        return False
    if field in MULTI_SELECT:
        return bool(split_pipe(v))
    return True


def make_issue(survey_id, field, value, issue_type, detail, question,
               suggestion="", action="", corrected_value="", apply_to=""):
    return {
        "survey_id":             survey_id,
        "field":                 field,
        "current_value":         "" if is_blank(value) else str(value),
        "issue_type":            issue_type,
        "detail":                detail,
        "question_for_reviewer": question,
        "suggested_correction":  suggestion,
        "action":                action,
        "corrected_value":       corrected_value,
        "apply_to":              apply_to,
        "notes":                 "",
    }


def make_recode_issue(survey_id, field, value, issue_type, detail, question,
                      corrected_value, apply_to="this_survey"):
    return make_issue(
        survey_id,
        field,
        value,
        issue_type,
        detail,
        question,
        suggestion=corrected_value,
        action="recode",
        corrected_value=corrected_value,
        apply_to=apply_to,
    )


def make_accept_issue(survey_id, field, value, issue_type, detail, question,
                      suggestion=""):
    return make_issue(
        survey_id,
        field,
        value,
        issue_type,
        detail,
        question,
        suggestion=suggestion,
        action="accept",
        apply_to="this_survey",
    )


def allowed_demographic_values(field: str) -> str:
    if field == "gender":
        return ", ".join(STANDARD_GENDER_VALUES)
    if field == "race_ethnicity":
        return ", ".join(STANDARD_RACE_VALUES)
    if field == "sexual_orientation":
        return ", ".join(STANDARD_ORIENTATION_VALUES)
    return ""


def make_demographic_recode_issue(survey_id, field, value, mapping):
    corrected = mapping["normalized"]
    return make_recode_issue(
        survey_id,
        field,
        value,
        "F_likely_typo",
        f"'{value}' is {mapping['detail']}",
        f"[{survey_id}] {field}='{value}'. Suggested standardization is '{corrected}'. "
        f"Review the default and change it only if needed.",
        corrected,
        apply_to=mapping.get("scope", "this_survey"),
    )


def make_demographic_accept_issue(survey_id, field, value, detail):
    return make_accept_issue(
        survey_id,
        field,
        value,
        "F_self_describe" if field != "race_ethnicity" else "F_unrecognized_race_label",
        detail,
        f"[{survey_id}] {field}='{value}'. Default action is accept as-is. "
        f"Review only if you want to override the captured wording for analysis.",
        suggestion="accept_as_is",
    )


def make_demographic_review_issue(survey_id, field, value, detail):
    allowed = allowed_demographic_values(field)
    issue_type = "F_self_describe" if field != "race_ethnicity" else "F_unrecognized_race_label"
    return make_issue(
        survey_id,
        field,
        value,
        issue_type,
        detail,
        f"[{survey_id}] {field}='{value}'. Reviewer approval required. Choose the best standard value in corrected_value, "
        f"or explicitly set action=accept only if the wording should be preserved.",
        suggestion=f"Allowed values: {allowed}",
        action="",
        corrected_value="",
        apply_to="this_survey",
    )


def make_demographic_clear_issue(survey_id, field, value, detail):
    return make_issue(
        survey_id,
        field,
        value,
        "F_likely_typo",
        detail,
        f"[{survey_id}] {field}='{value}'. Suggested action is clear because this looks like "
        f"prompt text or OCR noise rather than a respondent answer.",
        suggestion="clear",
        action="clear",
        apply_to="this_survey",
    )


def comma_parts(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def normalize_label_text(value: str) -> str:
    return value.strip().lower().rstrip(":")


def is_demographic_placeholder(value: str) -> bool:
    return normalize_label_text(value) in DEMOGRAPHIC_PLACEHOLDERS


def is_numeric_ocr_noise(value: str) -> bool:
    return value.strip().isdigit()


def all_parts_in_set(value: str, allowed: set[str]) -> bool:
    parts = split_pipe(value)
    return bool(parts) and all(part in allowed for part in parts)


# ---------------------------------------------------------------------------
# Rule A — Type / range validation
# ---------------------------------------------------------------------------

def check_A(row, issues):
    sid = row["survey_id"]

    for f in LIKERT_FIELDS:
        v = row.get(f, "")
        if is_blank(v):
            continue
        try:
            iv = int(float(v))
            assert 1 <= iv <= 5
        except (ValueError, TypeError, AssertionError):
            issues.append(make_issue(
                sid, f, v, "A_invalid_range",
                f"Expected integer 1–5, got '{v}'",
                f"[{sid}] {f} = '{v}'. Valid range is 1–5. What is the correct value?",
            ))

    nps = row.get("q22_nps", "")
    if not is_blank(nps):
        try:
            iv = int(float(nps))
            assert 0 <= iv <= 10
        except (ValueError, TypeError, AssertionError):
            issues.append(make_issue(
                sid, "q22_nps", nps, "A_invalid_range",
                f"Expected integer 0–10, got '{nps}'",
                f"[{sid}] q22_nps = '{nps}'. Valid range is 0–10. What is the correct value?",
            ))


# ---------------------------------------------------------------------------
# Rule B — Categorical allowed-values
# ---------------------------------------------------------------------------

def check_B(row, issues):
    sid = row["survey_id"]
    for f, allowed in SINGLE_SELECT.items():
        v = str(row.get(f, "")).strip()
        if not v:
            continue
        if v not in allowed:
            issues.append(make_issue(
                sid, f, v, "B_invalid_category",
                f"'{v}' not in allowed set",
                f"[{sid}] {f} = '{v}'. Not a recognized code. What should this be?",
                f"Allowed: {', '.join(sorted(allowed))}",
            ))


# ---------------------------------------------------------------------------
# Rule C — Array token allowed-values
# ---------------------------------------------------------------------------

def check_C(row, issues):
    sid = row["survey_id"]
    for f, allowed in MULTI_SELECT.items():
        for tok in split_pipe(row.get(f, "")):
            if tok not in allowed:
                issues.append(make_issue(
                    sid, f, tok, "C_invalid_token",
                    f"Token '{tok}' not in allowed codes for {f}",
                    f"[{sid}] {f} contains '{tok}'. Not a recognized code. What should this be?",
                    f"Allowed: {', '.join(sorted(allowed))}",
                ))


# ---------------------------------------------------------------------------
# Rule D — Conditional logic
# ---------------------------------------------------------------------------

def check_D(row, issues):
    sid = row["survey_id"]

    # Standard scalar-parent conditionals
    for child, (parent, allowed_vals) in CONDITIONAL.items():
        parent_val = str(row.get(parent, "")).strip()
        condition_met = parent_val in allowed_vals
        child_filled = has_value(row, child)

        if not condition_met and child_filled:
            # Special case: q12 blank + q13 filled → infer q12=no_place rather than clear q13
            if child == "q13_sleeping_location" and not parent_val:
                issues.append(make_issue(
                    sid, "q12_housing_stability", "", "D_infer_parent",
                    "q12_housing_stability is blank but q13_sleeping_location has a value",
                    f"[{sid}] q13='{row.get('q13_sleeping_location','')}' "
                    f"but q12 is blank — auto-setting q12=no_place.",
                    "no_place",
                    action="recode", corrected_value="no_place", apply_to="this_survey",
                ))
            else:
                issues.append(make_issue(
                    sid, child, row.get(child, ""), "D_auto_clear",
                    f"{child} is non-blank but {parent}='{parent_val}' (condition not met)",
                    f"[{sid}] {child}='{row.get(child,'')}' but {parent}='{parent_val}' "
                    f"— condition not met. Auto-clearing {child}.",
                    "clear",
                    action="clear", apply_to="this_survey",
                ))

        # Flag missing conditional value only for non-free-text fields
        if condition_met and not child_filled and child != "q16a_what_would_help":
            issues.append(make_issue(
                sid, child, "", "D_conditional_missing",
                f"{child} is blank but {parent}='{parent_val}' (condition is met)",
                f"[{sid}] {parent}='{parent_val}' so {child} is expected, "
                f"but it's blank. Was this intentionally left blank on the survey?",
            ))

    # q11a: condition = "quit" in q11_left_job_reasons (array parent)
    q11_tokens = set(split_pipe(row.get("q11_left_job_reasons", "")))
    q11a_filled = has_value(row, "q11a_quit_reasons")
    if "quit" not in q11_tokens and q11a_filled:
        issues.append(make_issue(
            sid, "q11a_quit_reasons", row.get("q11a_quit_reasons", ""),
            "D_auto_clear",
            "'quit' not in q11_left_job_reasons but q11a_quit_reasons is non-blank",
            f"[{sid}] q11a='{row.get('q11a_quit_reasons','')}' "
            f"but 'quit' not in q11_left_job_reasons — auto-clearing q11a.",
            "clear",
            action="clear", apply_to="this_survey",
        ))
    if "quit" in q11_tokens and not q11a_filled:
        issues.append(make_issue(
            sid, "q11a_quit_reasons", "", "D_conditional_missing",
            "'quit' is in q11_left_job_reasons but q11a_quit_reasons is blank",
            f"[{sid}] 'quit' is selected in q11_left_job_reasons "
            f"but q11a_quit_reasons is blank. Was this intentionally left blank?",
        ))

    # q26a / q26b: condition = q25 includes "checking" or "savings"
    q25_tokens = set(split_pipe(row.get("q25_bank_account", "")))
    has_account = bool(q25_tokens & {"checking", "savings"})
    for child in ("q26a_account_setup", "q26b_account_usage"):
        child_filled = has_value(row, child)
        if not has_account and child_filled:
            issues.append(make_issue(
                sid, child, row.get(child, ""), "D_auto_clear",
                f"{child} is non-blank but q25 has no checking/savings",
                f"[{sid}] {child}='{row.get(child,'')}' but q25 has no active account "
                f"— auto-clearing {child}.",
                "clear",
                action="clear", apply_to="this_survey",
            ))
        if has_account and not child_filled:
            issues.append(make_issue(
                sid, child, "", "D_conditional_missing",
                f"{child} is blank but q25 includes checking/savings",
                f"[{sid}] q25 includes an active account but {child} is blank. "
                f"Was this intentionally left blank on the survey?",
            ))


# ---------------------------------------------------------------------------
# Rule E — Missing required fields
# ---------------------------------------------------------------------------

def check_E(row, issues):
    sid = row["survey_id"]
    for f in REQUIRED_FIELDS:
        if is_blank(row.get(f, "")):
            issues.append(make_issue(
                sid, f, "", "E_missing_required",
                f"Required field '{f}' is blank",
                f"[{sid}] Required field '{f}' is blank. Was this left blank on the survey, "
                f"or is there a value that was missed during extraction?",
            ))


# ---------------------------------------------------------------------------
# Rule F — Free-text / demographic normalization
# ---------------------------------------------------------------------------

def check_F(row, issues):
    sid = row["survey_id"]

    # DOB
    dob = str(row.get("dob", "")).strip()
    if dob:
        try:
            pd.to_datetime(dob)
        except Exception:
            issues.append(make_issue(
                sid, "dob", dob, "F_invalid_date",
                f"'{dob}' cannot be parsed as a date",
                f"[{sid}] dob='{dob}' is not a valid date. What is the correct date of birth?",
            ))

    # Gender
    gender = str(row.get("gender", "")).strip()
    if gender:
        if is_demographic_placeholder(gender):
            issues.append(make_demographic_clear_issue(
                sid,
                "gender",
                gender,
                "placeholder label captured instead of a gender response",
            ))
        elif is_numeric_ocr_noise(gender):
            issues.append(make_demographic_clear_issue(
                sid,
                "gender",
                gender,
                "numeric OCR noise captured in the gender field",
            ))
        elif all_parts_in_set(gender, KNOWN_GENDER_LABELS):
            pass
        elif gender in GENDER_NORMALIZATION:
            issues.append(make_demographic_recode_issue(sid, "gender", gender, GENDER_NORMALIZATION[gender]))
        elif gender not in KNOWN_GENDER_LABELS:
            issues.append(make_demographic_review_issue(
                sid,
                "gender",
                gender,
                f"'{gender}' is not a standard printed label",
            ))

    # Sexual orientation
    orientation = str(row.get("sexual_orientation", "")).strip()
    if orientation:
        if is_demographic_placeholder(orientation):
            issues.append(make_demographic_clear_issue(
                sid,
                "sexual_orientation",
                orientation,
                "placeholder label captured instead of a sexual-orientation response",
            ))
        elif is_numeric_ocr_noise(orientation):
            issues.append(make_demographic_clear_issue(
                sid,
                "sexual_orientation",
                orientation,
                "numeric OCR noise captured in the sexual-orientation field",
            ))
        elif orientation in ORIENTATION_NORMALIZATION:
            issues.append(make_demographic_recode_issue(
                sid,
                "sexual_orientation",
                orientation,
                ORIENTATION_NORMALIZATION[orientation],
            ))
        elif orientation not in KNOWN_ORIENTATION_LABELS:
            parts = comma_parts(orientation)
            if len(parts) > 1 and all(part in KNOWN_ORIENTATION_LABELS for part in parts):
                issues.append(make_demographic_review_issue(
                    sid,
                    "sexual_orientation",
                    orientation,
                    "contains multiple recognizable orientation labels and needs reviewer grouping",
                ))
            else:
                issues.append(make_demographic_review_issue(
                    sid,
                    "sexual_orientation",
                    orientation,
                    f"'{orientation}' is not a standard printed label",
                ))

    # Race / ethnicity tokens
    for tok in split_pipe(row.get("race_ethnicity", "")):
        if is_demographic_placeholder(tok):
            issues.append(make_demographic_clear_issue(
                sid,
                "race_ethnicity",
                tok,
                "placeholder label captured instead of a race response",
            ))
        elif is_numeric_ocr_noise(tok):
            issues.append(make_demographic_clear_issue(
                sid,
                "race_ethnicity",
                tok,
                "numeric OCR noise captured in the race field",
            ))
        elif tok in RACE_NORMALIZATION:
            issues.append(make_demographic_recode_issue(sid, "race_ethnicity", tok, RACE_NORMALIZATION[tok]))
        elif tok not in KNOWN_RACE_LABELS:
            issues.append(make_demographic_review_issue(
                sid,
                "race_ethnicity",
                tok,
                f"Race token '{tok}' is not a recognized printed label",
            ))


# ---------------------------------------------------------------------------
# Rule G — Low-confidence flags
# ---------------------------------------------------------------------------

def check_G(row, issues):
    sid = row["survey_id"]
    for col in row.index:
        if not col.endswith("_conf"):
            continue
        v = row[col]
        if is_blank(v):
            continue
        try:
            conf = float(v)
        except (ValueError, TypeError):
            continue
        if conf < LOW_CONF_THRESHOLD:
            field = col[:-5]
            val = row.get(field, "")
            issues.append(make_issue(
                sid, field, val, "G_low_confidence",
                f"Confidence {conf:.2f} < threshold {LOW_CONF_THRESHOLD}",
                f"[{sid}] {field}='{val}' — confidence {conf:.2f}. "
                f"Please verify this extraction is correct.",
            ))


# ---------------------------------------------------------------------------
# Reviewer workbook
# ---------------------------------------------------------------------------

ISSUE_CATEGORIES = {
    "A_invalid_range":           ("A — Type / Range",               "FF6B6B"),
    "B_invalid_category":        ("B — Invalid Category",           "FFA07A"),
    "C_invalid_token":           ("C — Invalid Array Token",        "FFD700"),
    "D_auto_clear":              ("D — Auto-Clear Child",           "ADD8E6"),
    "D_infer_parent":            ("D — Infer Parent Value",         "87CEFA"),
    "D_conditional_missing":     ("D — Conditional Missing",        "B0E0E6"),
    "E_missing_required":        ("E — Missing Required",           "FFB6C1"),
    "F_invalid_date":            ("F — Invalid Date",               "98FB98"),
    "F_likely_typo":             ("F — Likely Typo",                "90EE90"),
    "F_self_describe":           ("F — Self-Describe",              "ADFF2F"),
    "F_unrecognized_race_label": ("F — Unrecognized Race Label",    "7FFF00"),
    "G_low_confidence":          ("G — Low Confidence",             "D3D3D3"),
}

# These issue types are accepted as-is (no reviewer action needed).
# Correction decisions recorded in ROADMAP.md Phase 3 QA section.
ACCEPTED_TYPES = {
    "D_auto_clear",          # child field cleared automatically — parent condition not met
    "D_infer_parent",        # parent inferred from child (q12 from q13)
    "D_conditional_missing", # follow-up fields blank on survey — leave as-is
    "E_missing_required",    # fields blank on survey — leave as-is
    "G_low_confidence",      # reviewer already verified during Phase 1b review
}

HEADERS = [
    "Issue Category", "survey_id", "field", "current_value",
    "detail", "question_for_reviewer", "suggested_correction",
    "action", "corrected_value", "apply_to", "notes",
]

COL_WIDTHS = [28, 12, 30, 35, 50, 80, 35, 14, 30, 16, 25]

INSTRUCTIONS = (
    "HOW TO USE THIS WORKBOOK:\n"
    "Review the prefilled defaults in columns H–J before changing anything.\n"
    "  Demographic mapping rows are review-only by default: the action/value/scope is already filled in using the current crosswalk tables.\n"
    "  Difficult self-described demographic rows stay on QA Questions with no default action; use column G for the allowed standard values before approving a recode or accept decision.\n"
    "  Only edit the prefilled demographic defaults when a suggested recode or clear decision is wrong for that response.\n"
    "  action        → clear | recode | accept\n"
    "  corrected_value → new value if action=recode (leave blank otherwise)\n"
    "  apply_to      → this_survey | all_surveys\n"
    "When done, save and run:  python scripts/03b_apply_corrections_412YZ.py"
)


def _write_issue_rows(ws, issues, hdr_fill, hdr_font, include_reviewer_cols):
    """Write a header row + issue rows onto ws. Returns row count written."""
    reviewer_fill = PatternFill("solid", fgColor="FFF9C4")

    ws.append(HEADERS)
    for i, cell in enumerate(ws[1], 1):
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    if include_reviewer_cols:
        for col in (8, 9, 10, 11):
            ws.cell(1, col).fill = PatternFill("solid", fgColor="F9A825")
    ws.row_dimensions[1].height = 30

    for issue in issues:
        label, hex_col = ISSUE_CATEGORIES.get(
            issue["issue_type"], (issue["issue_type"], "FFFFFF")
        )
        ws.append([
            label,
            issue["survey_id"],
            issue["field"],
            issue["current_value"],
            issue["detail"],
            issue["question_for_reviewer"],
            issue["suggested_correction"],
            issue["action"],
            issue["corrected_value"],
            issue["apply_to"],
            issue["notes"],
        ])
        row_fill = PatternFill("solid", fgColor=hex_col)
        for i, cell in enumerate(ws[ws.max_row], 1):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = row_fill if i <= 7 else reviewer_fill

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def issue_needs_reviewer_input(issue: dict) -> bool:
    if issue["issue_type"] in ACCEPTED_TYPES:
        return False
    return str(issue.get("action", "")).strip().lower() != "accept"


def write_reviewer_workbook(issues: list, out_path: Path):
    active   = [i for i in issues if issue_needs_reviewer_input(i)]
    accepted = [i for i in issues if not issue_needs_reviewer_input(i)]

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    # ── Sheet 1: Instructions ───────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst["A1"] = INSTRUCTIONS
    ws_inst["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_inst.row_dimensions[1].height = 120
    ws_inst.column_dimensions["A"].width = 90

    # ── Sheet 2: QA Questions (active — needs reviewer input) ───────────────
    ws_active = wb.create_sheet("QA Questions")
    _write_issue_rows(ws_active, active, hdr_fill, hdr_font, include_reviewer_cols=True)

    n = len(active) + 1
    if n >= 2:
        dv_action = DataValidation(
            type="list", formula1='"clear,recode,accept"',
            allow_blank=True, showDropDown=False,
        )
        ws_active.add_data_validation(dv_action)
        dv_action.add(f"H2:H{n}")

        dv_scope = DataValidation(
            type="list", formula1='"this_survey,all_surveys"',
            allow_blank=True, showDropDown=False,
        )
        ws_active.add_data_validation(dv_scope)
        dv_scope.add(f"J2:J{n}")

    # ── Sheet 3: Accepted — No Action ───────────────────────────────────────
    ws_acc = wb.create_sheet("Accepted - No Action")
    _write_issue_rows(ws_acc, accepted, hdr_fill, hdr_font, include_reviewer_cols=False)

    # ── Sheet 4: Summary ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Issue Category", "Count", "Tab"])
    for cell in ws_sum[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    counts_active = Counter(i["issue_type"] for i in active)
    counts_accepted = Counter(i["issue_type"] for i in accepted)
    for itype, (label, hex_col) in ISSUE_CATEGORIES.items():
        n_cat = counts_active.get(itype, 0) + counts_accepted.get(itype, 0)
        if counts_active.get(itype, 0) and counts_accepted.get(itype, 0):
            tab = "Both"
        elif counts_active.get(itype, 0):
            tab = "QA Questions"
        else:
            tab = "Accepted - No Action"
        ws_sum.append([label, n_cat, tab])
        row_fill = PatternFill("solid", fgColor=hex_col)
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = row_fill

    ws_sum.append(["TOTAL — needs review", len(active),   "QA Questions"])
    ws_sum.append(["TOTAL — accepted",     len(accepted), "Accepted - No Action"])
    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 25

    wb.save(str(out_path))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    csv_path = BASE_DIR / "output" / "412YZ" / "survey_data_412YZ.csv"
    if not csv_path.exists():
        print(f"CSV not found: {csv_path}")
        sys.exit(1)

    df = pd.read_csv(str(csv_path), encoding="utf-8-sig", dtype=str).fillna("")
    print(f"Loaded {len(df)} surveys, {len(df.columns)} columns.")

    all_issues: list[dict] = []
    for _, row in df.iterrows():
        check_A(row, all_issues)
        check_B(row, all_issues)
        check_C(row, all_issues)
        check_D(row, all_issues)
        check_E(row, all_issues)
        check_F(row, all_issues)
        check_G(row, all_issues)

    out_dir = BASE_DIR / "output" / "412YZ"

    # Machine-readable flag log
    flag_df = pd.DataFrame(all_issues)
    flag_path = out_dir / "flagged_412YZ.csv"
    flag_df.to_csv(str(flag_path), index=False, encoding="utf-8-sig")

    # Reviewer workbook
    xlsx_path = out_dir / "qa_questions_412YZ.xlsx"
    write_reviewer_workbook(all_issues, xlsx_path)

    # Console summary (ASCII-safe for Windows terminal)
    active   = [i for i in all_issues if issue_needs_reviewer_input(i)]
    accepted = [i for i in all_issues if not issue_needs_reviewer_input(i)]
    counts_active = Counter(i["issue_type"] for i in active)
    counts_accepted = Counter(i["issue_type"] for i in accepted)
    print(f"\n{'='*62}")
    print(f"QA Summary -- {len(df)} surveys, {len(all_issues)} total issues")
    print(f"{'='*62}")
    print(f"  Needs review (QA Questions tab):")
    for itype, (label, _) in ISSUE_CATEGORIES.items():
        n = counts_active.get(itype, 0)
        if n:
            safe_label = label.replace("\u2014", "-").replace("\u2013", "-")
            print(f"    {safe_label:<38} {n:>4}")
    print(f"  Accepted -- no action (separate tab):")
    for itype, (label, _) in ISSUE_CATEGORIES.items():
        n = counts_accepted.get(itype, 0)
        if n:
            safe_label = label.replace("\u2014", "-").replace("\u2013", "-")
            print(f"    {safe_label:<38} {n:>4}")
    print(f"  {'-'*46}")
    print(f"  {'Needs review':<40} {len(active):>4}")
    print(f"  {'Accepted':<40} {len(accepted):>4}")
    print(f"\nOutputs written:")
    print(f"  {flag_path}")
    print(f"  {xlsx_path}")


if __name__ == "__main__":
    main()

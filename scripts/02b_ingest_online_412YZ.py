"""
02b_ingest_online_412YZ.py
Load the SurveyMonkey online export for the 412YZ survey, normalize all
responses to the paper survey schema, and merge into
output/412YZ/survey_data_412YZ.csv.

- Appends rows with survey_id o001–oNNN and source="online".
- Paper rows get source="paper" added if not already present.
- Idempotent: drops any existing online rows (survey_id starts with "o")
  before re-appending.
- Sets all _conf columns to 1.0 for online rows (no OCR uncertainty).
- Fields not collected online (dob, first_initial, last_name) are left blank.

Usage:
    python scripts/02b_ingest_online_412YZ.py
"""

import sys
from pathlib import Path

import pandas as pd
import openpyxl

BASE_DIR    = Path(__file__).parent.parent
ONLINE_PATH = BASE_DIR / "data" / "online" / "Youth Zone Survey - Feb 2026.xlsx"
CSV_PATH    = BASE_DIR / "output" / "412YZ" / "survey_data_412YZ.csv"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(val):
    """Strip, collapse whitespace, and normalise smart quotes for dict lookup."""
    if val is None:
        return ""
    s = " ".join(str(val).split())
    # Replace curly/smart apostrophes and quotes with straight equivalents
    s = s.replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"')
    return s


def s(row, col):
    """Return normalized string for row[col], or '' if None/missing."""
    try:
        return _norm(row[col])
    except IndexError:
        return ""


def lk(val, d, default=None):
    """Lookup normalized val in dict d; return default if not found."""
    key = _norm(val)
    if not key:
        return default
    return d.get(key, default)


def collect_multi(row, col_to_code, skip_cols=()):
    """
    Collect a multi-select field from checkbox columns.

    col_to_code : {col_idx: code}  — code=None is a no-op (never included)
    skip_cols   : if any of these columns is non-blank the entire field is
                  treated as "Does not apply" → returns ""

    Returns a pipe-separated string of codes, e.g. "token_a | token_b",
    or "" if nothing selected.
    """
    for c in skip_cols:
        if s(row, c):
            return ""
    tokens = []
    for col, code in col_to_code.items():
        if s(row, col) and code is not None:
            tokens.append(code)
    return " | ".join(tokens)


def first_multi(row, col_to_code, skip_val_col=None):
    """
    Like collect_multi but returns only the first matching code (string).
    Used for fields that are single-select in the paper schema.
    skip_val_col: if this col is non-blank, return None immediately.
    """
    if skip_val_col is not None and s(row, skip_val_col):
        return None
    for col, code in col_to_code.items():
        if s(row, col) and code is not None:
            return code
    return None


def insert_after_first_present(df, new_col, value, anchor_cols):
    """Insert new_col after the first anchor column that exists, else append."""
    for anchor in anchor_cols:
        if anchor in df.columns:
            df.insert(df.columns.get_loc(anchor) + 1, new_col, value)
            return
    df[new_col] = value


# ---------------------------------------------------------------------------
# Lookup tables  (keys are whitespace-normalised to guard against double-spaces)
# ---------------------------------------------------------------------------

LIKERT5 = {
    "Never": "1", "Rarely": "2", "Sometimes": "3", "Often": "4", "All the time": "5",
}

FREQ = {
    "Almost every day":        "almost_every_day",
    "About once a week":       "about_once_a_week",
    "1-2 times per month":     "1_2_times_per_month",
    "Less than once a month":  "less_than_once_a_month",
}

COMM_LEVEL = {
    "Not enough":   "not_enough",
    "A good amount": "good_amount",
    "Too much":     "too_much",
}

DURATION = {
    "Less than 6 months":                     "less_6mo",
    "6-12 months":                            "6_12mo",
    "More than 1 year, but less than 3 years": "1_3yr",
    "3 years or more":                        "3yr_plus",
}

SCHOOL_STATUS = {
    "High school":                                   "high_school",
    "GED program":                                   "ged",
    "College/Career or technical school":            "college_career",
    "Graduate school (ex. Master's degree program)": "graduate",
    "Not in school":                                 "not_in_school",
}

EDUCATION = {
    "Some high school":                          "some_hs",
    "High school diploma or GED":                "hs_diploma_ged",
    "Some college/career or technical school":   "some_college",
    "College degree or Certification":           "college_degree",
    "Graduate degree (ex. Master's, PhD)":       "graduate",
}

LICENSE = {
    "Yes":                      "yes",
    "Learner's Permit":         "learners_permit",
    "I have a Learner's Permit": "learners_permit",
    "No":                       "no",
}

VEHICLE = {
    "I have my own reliable vehicle.":                                          "own_reliable",
    "I have my own vehicle, but it often breaks down or needs repairs.":        "own_unreliable",
    "I share a reliable vehicle with someone else.":                            "share_reliable",
    "I share a vehicle with someone else, but it often breaks down or needs repairs.": "share_unreliable",
    "I can borrow a car from friends or family when I need to.":                "borrow",
    "I do not usually have access to a reliable vehicle.":                      "no_access",
}

VOTE = {"Yes": "yes", "No": "no"}

NOT_REG = {
    "I am not old enough to vote":                        "not_old_enough",
    "I do not know how to register":                      "dont_know_how",
    "I do not understand politics":                       "dont_understand",
    "I do not believe my one vote will make a difference": "vote_wont_matter",
    "Other (please describe)":                            "other",
}

EMPLOYMENT = {
    "Yes, part-time":                      "yes_part_time",
    "Yes, full-time":                      "yes_full_time",
    "No":                                  "no",
    "I am in a job training program":      "job_training_program",
}

JOB_TENURE = {
    "Less than 3 months": "less_3mo",
    "3-6 months":         "3_6mo",
    "More than 6 months": "more_6mo",
}

JOB_SEEKING = {
    "Yes":                                       "yes",
    "No":                                        "no",
    "Not yet - I am in a job training program":  "not_yet_in_training",
}

TRANSPORT = {
    "Bus or other public transportation":                       "public_transit",
    "Driving myself":                                           "driving_self",
    "Getting rides from someone else (coworker, family, friend, etc.)": "rides_from_others",
    "RideShare app (Lyft, Uber)":                               "rideshare",
    "Walking, biking, scooter, etc.":                           "active_transport",
    "Other (please specify)":                                   "other",
}

HOUSING = {
    "I have housing that I feel safe in, and I can stay there for at least the next 90 days.": "stable",
    "I have housing that feels safe, but I can not stay there for the next 90 days.":          "safe_not_90days",
    "I have somewhere to stay for the next 90 days, but it's not safe.":                       "90days_not_safe",
    "I do not have a place to stay right now.":                                                "no_place",
}

VISIT_FREQ = {
    "Every week":              "every_week",
    "1-3 times per month":     "1_3_times_per_month",
    "Less than once per month": "less_than_once_per_month",
    "Never":                   "never",
}

STAY_FOCUSED = {
    "Agree":                                    "agree",
    "Somewhat agree":                           "somewhat_agree",
    "Disagree":                                 "disagree",
    "Unsure. I don't have clear goals right now": "unsure",
    "Unsure":                                   "unsure",
}

RESPECT = {
    "Never": "never", "Rarely": "rarely", "Sometimes": "sometimes",
    "Often": "often", "All the time": "all_the_time",
}

INDEPENDENCE = {
    "Agree":         "agree",
    "Somewhat":      "somewhat",
    "Somewhat agree": "somewhat",   # guard for variation
    "Disagree":      "disagree",
    "Unsure":        "unsure",
}

AGE_RANGE = {
    "16-17 years old": "16_17",
    "18-20 years old": "18_20",
    "21-23 years old": "21_23",
}

# ---------------------------------------------------------------------------
# Multi-select column maps  {col_index: output_code}
# ---------------------------------------------------------------------------

# Q10 job barriers (cols 31-42; col 31 = "Does not apply" skip trigger)
Q10_COLS = {
    32: "childcare", 33: "criminal_background", 34: "no_references",
    35: "interview_skills", 36: "no_diploma", 37: "limited_experience",
    38: "mental_physical_health", 39: "transportation", 40: "drugs_alcohol",
    41: "not_getting_called", 42: "something_else",
}

# Q11 left job (cols 43-50; col 43 = "Does not apply" skip trigger)
# col 50 value = free text for "other"
Q11_COLS = {
    44: "found_better", 45: "quit", 46: "fired_attendance",
    47: "fired_performance", 48: "seasonal", 49: "pregnancy_parenting",
    50: "other",
}

# Q11a quit reasons (cols 51-59; col 59 value = free text for "other")
Q11A_COLS = {
    51: "low_pay_hours", 52: "schedule_conflict", 53: "lack_of_support",
    54: "poor_conditions", 55: "mental_emotional_health", 56: "transportation",
    57: "not_good_fit", 58: "personal_family", 59: "other",
}

# Q13 sleeping location (cols 61-67; single-select in paper — take first match)
Q13_COLS = {
    61: "friends_family", 62: "shelter", 63: "couch_surfing",
    64: "car", 65: "outside", 66: "abandoned_building", 67: "other",
}

# Q14 housing instability (cols 68-73; col 72 = "Does not apply" skip trigger)
Q14_COLS = {
    68: "evicted_nonpayment", 69: "evicted_other",
    70: "lost_informal_housing", 71: "left_unsafe", 73: "other",
}

# Q15a visit reasons (cols 75-87; col 87 value = free text for "other")
Q15A_COLS = {
    75: "computers", 76: "safe_place", 77: "laundry_shower", 78: "food",
    79: "escape_problems", 80: "health_counseling", 81: "learn_skills",
    82: "service_providers", 83: "see_coach_staff", 84: "socialize",
    85: "work_on_goals", 86: "scheduled_activity", 87: "other",
}

# Q15b visit barriers (cols 88-91; col 91 value = free text for "other")
Q15B_COLS = {
    88: "coach_invitation", 89: "more_info", 90: "better_activities", 91: "other",
}

# Q17 program helped (cols 94-107; col 94 = "None of the above" skip trigger)
Q17_COLS = {
    95: "health_counseling", 96: "positive_relationships", 97: "handle_problems",
    98: "housing", 99: "education", 100: "job", 101: "drivers_license",
    102: "parenting", 103: "everyday_skills", 104: "decision_making",
    105: "vital_documents", 106: "future", 107: "something_else",
}

# Q24 money methods (cols 118-124; col 124 value = free text for "other")
Q24_COLS = {
    118: "bank_account", 119: "check_cashing", 120: "digital_apps",
    121: "paypal", 122: "money_order", 123: "cash_at_home", 124: "other",
}

# Q25 bank account status (cols 125-128)
Q25_COLS = {
    125: "checking", 126: "savings", 127: "had_in_past", 128: "never_had",
}

# Q26a account setup (cols 129-133; col 133 value = free text for "other")
Q26A_COLS = {
    129: "self_online", 130: "self_inperson", 131: "self_with_help",
    132: "added_by_other", 133: "other",
}

# Q26b account usage (cols 134-147; col 146 = "none" token; col 147 free text)
Q26B_COLS = {
    134: "budgeting", 135: "saving", 136: "cashing_checks",
    137: "writing_checks", 138: "keep_safe", 139: "transferring",
    140: "direct_deposit", 141: "debit_card", 142: "online_banking",
    143: "atm", 144: "in_person_banking", 145: "paying_bills",
    146: "none", 147: "other",
}

# Gender checkboxes (cols 148-155; col 156 = self-describe free text)
GENDER_COLS = {
    148: "Female", 149: "Gender Nonconforming", 150: "Genderqueer",
    151: "Male", 152: "Non-binary", 153: "Transgender Female",
    154: "Transgender Male", 155: "Two-Spirit",
}

# Race/ethnicity checkboxes (cols 158-168; col 169 = self-describe free text)
RACE_COLS = {
    158: "Black or of African or Caribbean Descent",
    159: "East Asian",
    160: "Hispanic or Latinx",
    161: "Native American or Indigenous peoples of America",
    162: "Native Hawaiian or Pacific Islander",
    163: "South Asian or Indian (Subcontinent)",
    164: "Southeast Asian",
    165: "Western Asian or Middle Eastern",
    166: "Other Asian",
    167: "White or of European Descent",
    168: "Multi-Racial",
}

# ---------------------------------------------------------------------------
# Row mapper
# ---------------------------------------------------------------------------

def map_row(row, idx):
    """Map one online Excel data row (tuple) to the paper CSV schema dict."""

    survey_id = f"o{idx:03d}"

    # Coach name — col 10 is the write-in field (col 9 echoes the sub-label)
    coach_raw = s(row, 10)
    coach_name = coach_raw if coach_raw and coach_raw.lower() != "coach name:" else None

    # --- Q1 Likert (cols 11-15) ---
    q1_trustworthy      = lk(s(row, 11), LIKERT5)
    q1_reliable         = lk(s(row, 12), LIKERT5)
    q1_values_opinions  = lk(s(row, 13), LIKERT5)
    q1_available        = lk(s(row, 14), LIKERT5)
    q1_heard_understood = lk(s(row, 15), LIKERT5)

    # --- Q2, Q3, Q4 ---
    q2 = lk(s(row, 16), FREQ)
    q3 = lk(s(row, 17), COMM_LEVEL)
    q4 = lk(s(row, 18), DURATION)

    # --- Q5, Q5a ---
    q5  = lk(s(row, 19), SCHOOL_STATUS)
    q5a = lk(s(row, 20), EDUCATION) if q5 == "not_in_school" else None

    # --- Q6, Q6a ---
    q6  = lk(s(row, 21), LICENSE)
    q6a = lk(s(row, 22), VEHICLE) if q6 == "yes" else None

    # --- Q7, Q7a ---
    q7 = lk(s(row, 23), VOTE)
    if q7 == "no":
        q7a_code = lk(s(row, 24), NOT_REG)
        q7a_pipe = q7a_code if q7a_code else ""
        # "Other (please describe)" in col 24 → code = "other"; free text in col 25
        q7a_other = s(row, 25) or None
    else:
        q7a_pipe  = ""
        q7a_other = None

    # --- Q8, Q8a, Q8b ---
    q8  = lk(s(row, 26), EMPLOYMENT)
    q8a = lk(s(row, 27), JOB_TENURE)  if q8 in ("yes_part_time", "yes_full_time") else None
    q8b = lk(s(row, 28), JOB_SEEKING) if q8 in ("no", "job_training_program")     else None

    # --- Q9 ---
    q9       = lk(s(row, 29), TRANSPORT)
    q9_other = s(row, 30) or None

    # --- Q10 job barriers ---
    q10       = collect_multi(row, Q10_COLS, skip_cols=(31,))
    q10_other = s(row, 42) or None   # "something else" free text lives in col 42

    # --- Q11 left job ---
    q11       = collect_multi(row, Q11_COLS, skip_cols=(43,))
    q11_other = s(row, 50) or None   # "other" free text lives in col 50

    # --- Q11a quit reasons (conditional on "quit" in q11) ---
    if "quit" in q11.split(" | "):
        q11a       = collect_multi(row, Q11A_COLS)
        q11a_other = s(row, 59) or None
    else:
        q11a       = ""
        q11a_other = None

    # --- Q12, Q13 ---
    q12      = lk(s(row, 60), HOUSING)
    if q12 == "stable":
        q13       = None
        q13_other = None
    else:
        q13       = first_multi(row, Q13_COLS)
        q13_other = s(row, 67) or None

    # --- Q14 housing instability ---
    q14       = collect_multi(row, Q14_COLS, skip_cols=(72,))
    q14_other = s(row, 73) or None

    # --- Q15, Q15a, Q15b ---
    q15 = lk(s(row, 74), VISIT_FREQ)
    if q15 in ("every_week", "1_3_times_per_month"):
        q15a       = collect_multi(row, Q15A_COLS)
        q15a_other = s(row, 87) or None
    else:
        q15a       = ""
        q15a_other = None
    if q15 in ("less_than_once_per_month", "never"):
        q15b       = collect_multi(row, Q15B_COLS)
        q15b_other = s(row, 91) or None
    else:
        q15b       = ""
        q15b_other = None

    # --- Q16 ---
    q16      = lk(s(row, 92), STAY_FOCUSED)
    q16a_txt = s(row, 93) or None   # "what would help" open text

    # --- Q17 program helped ---
    q17       = collect_multi(row, Q17_COLS, skip_cols=(94,))
    q17_other = s(row, 107) or None   # "something else" free text

    # --- Q18, Q19 ---
    q18 = lk(s(row, 108), RESPECT)
    q19 = lk(s(row, 109), RESPECT)

    # --- Q20 environment (same Never=1…All the time=5 Likert as Q1) ---
    q20_pc = lk(s(row, 110), LIKERT5)
    q20_nj = lk(s(row, 111), LIKERT5)
    q20_dv = lk(s(row, 112), LIKERT5)
    q20_tf = lk(s(row, 113), LIKERT5)
    q20_ss = lk(s(row, 114), LIKERT5)

    # --- Q21, Q22, Q23 ---
    q21 = lk(s(row, 115), INDEPENDENCE)
    q22_raw = s(row, 116)
    try:
        q22 = str(int(float(q22_raw))) if q22_raw else None
    except (ValueError, TypeError):
        q22 = None
    q23 = s(row, 117) or None

    # --- Q24 money methods ---
    q24       = collect_multi(row, Q24_COLS)
    q24_other = s(row, 124) or None

    # --- Q25 bank account ---
    q25 = collect_multi(row, Q25_COLS)
    has_account = bool(q25) and any(t in q25.split(" | ") for t in ("checking", "savings"))

    # --- Q26a, Q26b (only if has checking or savings) ---
    if has_account:
        q26a       = collect_multi(row, Q26A_COLS)
        q26a_other = s(row, 133) or None
        q26b       = collect_multi(row, Q26B_COLS)
        q26b_other = s(row, 147) or None
    else:
        q26a = q26b = ""
        q26a_other = q26b_other = None

    # --- Demographics ---
    # Gender: collect checked boxes + self-describe
    gender_parts = [label for col, label in GENDER_COLS.items() if s(row, col)]
    gender_self  = s(row, 156)
    if gender_self:
        gender_parts.append(gender_self)
    gender = " | ".join(gender_parts) if gender_parts else None

    age_range = lk(s(row, 157), AGE_RANGE)

    # Race: collect checked boxes (store raw label — token_to_race_group uses
    # substring matching so online labels map correctly to groups)
    race_parts = [label for col, label in RACE_COLS.items() if s(row, col)]
    race_self  = s(row, 169)
    if race_self:
        race_parts.append(race_self)
    race = " | ".join(race_parts)

    # Sexual orientation: self-describe overrides the standard selection
    orient_self = s(row, 171)
    orient_std  = s(row, 170)
    orientation = orient_self if orient_self else (orient_std if orient_std else None)

    return {
        "survey_id":                    survey_id,
        "dob":                          None,
        "first_initial":                None,
        "last_name":                    None,
        "coach_name":                   coach_name,
        "coach_name_corrected":         None,
        "q1_trustworthy":               q1_trustworthy,
        "q1_reliable":                  q1_reliable,
        "q1_values_opinions":           q1_values_opinions,
        "q1_available":                 q1_available,
        "q1_heard_understood":          q1_heard_understood,
        "q2_communication_frequency":   q2,
        "q3_communication_level":       q3,
        "q4_program_duration":          q4,
        "q5_school_status":             q5,
        "q5a_highest_education":        q5a,
        "q6_drivers_license":           q6,
        "q6a_vehicle_access":           q6a,
        "q7_registered_to_vote":        q7,
        "q7a_not_registered_reasons":   q7a_pipe,
        "q7a_other_text":               q7a_other,
        "q8_employment_status":         q8,
        "q8a_job_tenure":               q8a,
        "q8b_job_seeking":              q8b,
        "q9_primary_transport":         q9,
        "q9_other_text":                q9_other,
        "q10_job_barriers":             q10,
        "q10_something_else_text":      q10_other,
        "q11_left_job_reasons":         q11,
        "q11_other_text":               q11_other,
        "q11a_quit_reasons":            q11a,
        "q11a_other_text":              q11a_other,
        "q12_housing_stability":        q12,
        "q13_sleeping_location":        q13,
        "q13_other_text":               q13_other,
        "q14_housing_instability_reasons": q14,
        "q14_other_text":               q14_other,
        "q15_visit_frequency":          q15,
        "q15a_visit_reasons":           q15a,
        "q15a_other_text":              q15a_other,
        "q15b_visit_barriers":          q15b,
        "q15b_other_text":              q15b_other,
        "q16_stay_focused":             q16,
        "q16a_what_would_help":         q16a_txt,
        "q17_program_helped":           q17,
        "q17_something_else_text":      q17_other,
        "q17_none_explain_text":        None,
        "q18_staff_respect":            q18,
        "q19_peer_respect":             q19,
        "q20_people_care":              q20_pc,
        "q20_no_judgment":              q20_nj,
        "q20_diversity_valued":         q20_dv,
        "q20_treated_fairly":           q20_tf,
        "q20_safe_sharing":             q20_ss,
        "q21_gained_independence":      q21,
        "q22_nps":                      q22,
        "q23_other_comments":           q23,
        "q24_money_methods":            q24,
        "q24_other_text":               q24_other,
        "q25_bank_account":             q25,
        "q26a_account_setup":           q26a,
        "q26a_other_text":              q26a_other,
        "q26b_account_usage":           q26b,
        "q26b_other_text":              q26b_other,
        "gender":                       gender,
        "age_range":                    age_range,
        "race_ethnicity":               race,
        "sexual_orientation":           orientation,
        "source":                       "online",
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not ONLINE_PATH.exists():
        print(f"Online file not found: {ONLINE_PATH}")
        sys.exit(1)
    if not CSV_PATH.exists():
        print(f"Paper CSV not found: {CSV_PATH}")
        print("Run 02_compile.py first.")
        sys.exit(1)

    # --- Load paper CSV ---
    print(f"Loading paper CSV: {CSV_PATH}")
    df_paper = pd.read_csv(CSV_PATH, encoding="utf-8-sig", dtype=str)

    # Add source column if missing
    if "source" not in df_paper.columns:
        insert_after_first_present(
            df_paper,
            "source",
            "paper",
            ["coach_name_corrected", "coach_name"],
        )
    else:
        df_paper["source"] = df_paper["source"].fillna("paper")

    # Drop any previously ingested online rows (idempotency)
    paper_only = df_paper[~df_paper["survey_id"].str.startswith("o", na=False)].copy()
    dropped = len(df_paper) - len(paper_only)
    if dropped:
        print(f"  Dropped {dropped} existing online rows (will re-ingest)")

    # Collect _conf column names from paper CSV to propagate to online rows
    conf_cols = [c for c in df_paper.columns if c.endswith("_conf")]

    # --- Load online xlsx ---
    print(f"Loading online xlsx: {ONLINE_PATH}")
    wb = openpyxl.load_workbook(ONLINE_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    # Rows 0 and 1 are header rows; data starts at row 2
    data_rows = [r for r in all_rows[2:] if any(v is not None for v in r)]
    print(f"  Found {len(data_rows)} data rows")

    # --- Map online rows ---
    online_records = []
    for idx, row in enumerate(data_rows, start=1):
        online_records.append(map_row(row, idx))

    df_online = pd.DataFrame(online_records)

    # Add _conf = 1.0 for all conf columns (no OCR uncertainty)
    for col in conf_cols:
        df_online[col] = "1.0"

    # Ensure all paper columns exist in online df (fill missing with blank)
    for col in paper_only.columns:
        if col not in df_online.columns:
            df_online[col] = ""

    # Align column order to paper CSV
    df_online = df_online.reindex(columns=paper_only.columns, fill_value="")

    # --- Merge and save ---
    df_merged = pd.concat([paper_only, df_online], ignore_index=True)

    # Replace Python None with blank string for consistent CSV output
    df_merged = df_merged.fillna("").infer_objects(copy=False)

    df_merged.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")

    paper_n  = len(paper_only)
    online_n = len(df_online)
    print(f"\nMerged: {paper_n} paper + {online_n} online = {len(df_merged)} total rows")
    print(f"Saved: {CSV_PATH}")

    # --- Quick sanity check ---
    print("\nSanity check — source counts:")
    print(df_merged["source"].value_counts().to_string())
    print("\nSanity check — online age_range distribution:")
    print(df_merged[df_merged["source"] == "online"]["age_range"].value_counts().to_string())
    print("\nSanity check — online q8_employment_status distribution:")
    print(df_merged[df_merged["source"] == "online"]["q8_employment_status"].value_counts().to_string())


if __name__ == "__main__":
    main()

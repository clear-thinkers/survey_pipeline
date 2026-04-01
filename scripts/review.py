"""
review.py
Read a single extracted JSON and produce a human-review Excel workbook.

Usage:
    python scripts/review.py s001
"""

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
import config

BASE_DIR = Path(__file__).parent.parent

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

FILL_HEADER  = PatternFill("solid", fgColor="D9D9D9")  # light gray
FILL_SUMMARY = PatternFill("solid", fgColor="DEEAF1")  # light blue
FILL_RED     = PatternFill("solid", fgColor="FFCCCC")  # red    — must review
FILL_YELLOW  = PatternFill("solid", fgColor="FFF2CC")  # yellow — check
FILL_GREEN   = PatternFill("solid", fgColor="E2EFDA")  # green  — valid values present
FONT_BOLD    = Font(bold=True)
FONT_SMALL   = Font(size=10)

COL_WIDTHS = {1: 30, 2: 45, 3: 12, 4: 14, 5: 35, 6: 35, 7: 60}

MUST_REVIEW_THRESHOLD = 0.75  # conf < 0.75 → red

VALID_VALUES = {
    # IL fields
    "q6b_job_types":           "retail_customer_service, food_service, office_admin, healthcare_childcare_helping, warehouse_construction_handson, technology_creative, other",
    "q7_barriers":             "childcare, criminal_background, no_references, interview_skills, no_diploma, limited_experience, mental_physical_health, transportation, drugs_alcohol, not_getting_called, something_else, does_not_apply",
    "q8_left_job_reasons":     "found_better, quit, fired_attendance, fired_performance, seasonal, other, does_not_apply",
    "q8a_quit_reasons":        "low_pay_hours, schedule_conflict, lack_of_support, poor_conditions, mental_emotional_health, transportation, not_good_fit, personal_family, other",
    "q9_bank_account":         "checking, savings, had_in_past, never_had",
    "q9a_no_account_reasons":  "dont_know_how, fees, bad_credit, not_enough_money, min_balance_requirements, no_trusted_adult, tried_and_failed, other",
    "q11_program_helped":      "health_counseling, positive_relationships, handle_problems, housing, education, job, drivers_license, parenting, everyday_skills, decision_making, vital_documents, future, something_else, none",
    # 412YZ fields
    "q7a_not_registered_reasons": "not_old_enough, dont_know_how, dont_understand, vote_wont_matter, other",
    "q10_job_barriers":           "childcare, criminal_background, no_references, interview_skills, no_diploma, limited_experience, mental_physical_health, transportation, drugs_alcohol, not_getting_called, something_else",
    "q11_left_job_reasons":       "found_better, quit, fired_attendance, fired_performance, seasonal, pregnancy_parenting, other",
    "q11a_quit_reasons":          "low_pay_hours, schedule_conflict, lack_of_support, poor_conditions, mental_emotional_health, transportation, not_good_fit, personal_family, other",
    "q14_housing_instability_reasons": "evicted_nonpayment, evicted_other, lost_informal_housing, left_unsafe, other",
    "q15a_visit_reasons":         "computers, safe_place, laundry_shower, food, escape_problems, health_counseling, learn_skills, service_providers, see_coach_staff, socialize, work_on_goals, scheduled_activity, other",
    "q15b_visit_barriers":        "coach_invitation, more_info, better_activities, other",
    "q17_program_helped":         "health_counseling, positive_relationships, handle_problems, housing, education, job, drivers_license, parenting, everyday_skills, decision_making, vital_documents, future, something_else",
    "q24_money_methods":          "bank_account, check_cashing, digital_apps, paypal, money_order, cash_at_home, other",
    "q25_bank_account":           "checking, savings, had_in_past, never_had",
    "q26a_account_setup":         "self_online, self_inperson, self_with_help, added_by_other, other",
    "q26b_account_usage":         "budgeting, saving, cashing_checks, writing_checks, keep_safe, transferring, direct_deposit, debit_card, online_banking, atm, in_person_banking, paying_bills, none, other",
    # shared
    "race_ethnicity":          "Black or of African or Caribbean Descent, East Asian, Hispanic or Latinx, Native American Indigenous peoples of America, Native Hawaiian or Pacific Islander, South Asian or Indian Subcontinent, Southeast Asian, Western Asian or Middle Eastern, Other Asian, White or of European Descent, Multi-Racial",
}

# Canonical field order — covers both IL and 412YZ survey types.
# Fields not present in a given survey's JSON are simply absent from the workbook;
# the sort key falls back to len(FIELD_ORDER) so any unexpected fields go to the end.
FIELD_ORDER = [
    # Cover page (shared)
    "dob", "first_initial", "last_name", "coach_name",
    # Q1 — Coach traits Likert (shared)
    "q1_trustworthy", "q1_reliable", "q1_values_opinions", "q1_available", "q1_heard_understood",
    # Q2–Q4 (shared)
    "q2_communication_frequency",
    "q3_communication_level",
    "q4_program_duration",
    # Q5 — School (shared)
    "q5_school_status", "q5a_highest_education",
    # Q6 — IL: employment | 412YZ: driver's license
    "q6_employment_status", "q6a_job_tenure", "q6b_job_seeking",
    "q6b_job_types", "q6b_job_types_other",
    "q6_drivers_license", "q6a_vehicle_access",
    # Q7 — IL: barriers | 412YZ: voter registration
    "q7_barriers", "q7_something_else_text",
    "q7_registered_to_vote", "q7a_not_registered_reasons", "q7a_other_text",
    # Q8 — IL: left job | 412YZ: employment
    "q8_left_job_reasons", "q8_other_text",
    "q8a_quit_reasons", "q8a_other_text",
    "q8_employment_status", "q8a_job_tenure", "q8b_job_seeking",
    # Q9 — IL: bank account | 412YZ: transport
    "q9_bank_account", "q9a_no_account_reasons", "q9a_tried_failed_text", "q9a_other_text",
    "q9_primary_transport", "q9_other_text",
    # Q10 — IL: stay focused | 412YZ: job barriers
    "q10_stay_focused", "q10a_what_would_help",
    "q10_job_barriers", "q10_something_else_text",
    # Q11 — IL: program helped | 412YZ: left job
    "q11_program_helped", "q11_something_else_text", "q11_none_explain_text",
    "q11_left_job_reasons", "q11_other_text",
    "q11a_quit_reasons", "q11a_other_text",
    # Q12 — IL: other supports | 412YZ: housing stability
    "q12_other_supports",
    "q12_housing_stability",
    # Q13 — IL: staff respect | 412YZ: sleeping location
    "q13_staff_respect",
    "q13_sleeping_location", "q13_other_text",
    # Q14 — IL: peer respect | 412YZ: housing instability reasons
    "q14_peer_respect",
    "q14_housing_instability_reasons", "q14_other_text",
    # Q15 — IL: environment Likert | 412YZ: visit frequency + reasons/barriers
    "q15_people_care", "q15_no_judgment", "q15_diversity_valued",
    "q15_treated_fairly", "q15_safe_sharing",
    "q15_visit_frequency",
    "q15a_visit_reasons", "q15a_other_text",
    "q15b_visit_barriers", "q15b_other_text",
    # Q16 — IL: independence | 412YZ: stay focused
    "q16_gained_independence",
    "q16_stay_focused", "q16a_what_would_help",
    # Q17 — IL: NPS | 412YZ: program helped
    "q17_nps",
    "q17_program_helped", "q17_something_else_text", "q17_none_explain_text",
    # Q18 — IL: comments | 412YZ: staff respect
    "q18_other_comments",
    "q18_staff_respect",
    # Q19–Q26 (412YZ only)
    "q19_peer_respect",
    "q20_people_care", "q20_no_judgment", "q20_diversity_valued",
    "q20_treated_fairly", "q20_safe_sharing",
    "q21_gained_independence",
    "q22_nps",
    "q23_other_comments",
    "q24_money_methods", "q24_other_text",
    "q25_bank_account",
    "q26a_account_setup", "q26a_other_text",
    "q26b_account_usage", "q26b_other_text",
    # Demographics (shared)
    "gender", "age_range", "race_ethnicity", "sexual_orientation",
]


def format_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value) if value else ""
    return str(value)


# ---------------------------------------------------------------------------
# Build worksheet
# ---------------------------------------------------------------------------

def build_workbook(survey_id: str, fields: dict, confidence: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    threshold = config.CONFIDENCE_THRESHOLD

    # --- Assemble field rows ------------------------------------------------
    rows = []
    for field, value in fields.items():
        conf = confidence.get(field, 0.0)
        if not isinstance(conf, (int, float)):
            conf = 0.0

        if conf < MUST_REVIEW_THRESHOLD:
            tier = 2          # red — must review
            flag_label = "\u26a0 Must review"
            fill = FILL_RED
        elif conf < threshold:
            tier = 1          # yellow — check
            flag_label = "\u26a0 Check"
            fill = FILL_YELLOW
        else:
            tier = 0          # clean
            flag_label = ""
            fill = None

        rows.append({
            "field": field,
            "value": format_value(value),
            "confidence": conf,
            "flag_label": flag_label,
            "tier": tier,
            "fill": fill,
        })

    # Sort by canonical question order; unknown fields go to the end
    order_index = {f: i for i, f in enumerate(FIELD_ORDER)}
    rows.sort(key=lambda r: order_index.get(r["field"], len(FIELD_ORDER)))

    total_fields = len(rows)
    must_review_count = sum(1 for r in rows if r["tier"] == 2)
    check_count       = sum(1 for r in rows if r["tier"] == 1)
    clean_count       = sum(1 for r in rows if r["tier"] == 0)

    # --- Row 1: Header ------------------------------------------------------
    headers = ["Field", "Extracted Value", "Confidence", "Flagged",
               "Reviewer Correction", "Notes", "Valid values"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER

    # --- Row 2: Summary (merged A2:G2) --------------------------------------
    summary_text = (
        f"Survey: {survey_id}    |    "
        f"{total_fields} fields    |    "
        f"Must review (<{MUST_REVIEW_THRESHOLD:.0%}): {must_review_count}    |    "
        f"Check ({MUST_REVIEW_THRESHOLD:.0%}-{threshold - 0.01:.0%}): {check_count}    |    "
        f"Clean (>={threshold:.0%}): {clean_count}"
    )
    ws.merge_cells("A2:G2")
    summary_cell = ws["A2"]
    summary_cell.value = summary_text
    summary_cell.font = FONT_BOLD
    summary_cell.fill = FILL_SUMMARY
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")

    # --- Row 3: blank separator ---------------------------------------------
    # (leave empty)

    # --- Rows 4+: field data ------------------------------------------------
    for i, row in enumerate(rows, start=4):
        valid_vals = VALID_VALUES.get(row["field"], "")
        values = [
            row["field"],
            row["value"],
            row["confidence"],
            row["flag_label"],
            "",           # Reviewer correction
            "",           # Notes
            valid_vals,   # Valid values
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if row["fill"]:
                cell.fill = row["fill"]
            if col == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 7:
                cell.font = FONT_SMALL
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if valid_vals:
                    cell.fill = FILL_GREEN

    # --- Column widths ------------------------------------------------------
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- Freeze pane below header row ---------------------------------------
    ws.freeze_panes = "A2"

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def resolve_json_files() -> list[Path]:
    """Return list of extracted JSON paths based on CLI arguments."""
    all_jsons = sorted(config.EXTRACTED_DIR.glob("*.json"))
    names = [p.stem for p in all_jsons]

    if len(sys.argv) == 3:
        # Range mode: python review.py s008 s010
        start_id = sys.argv[1].removesuffix(".json")
        end_id   = sys.argv[2].removesuffix(".json")
        for sid in (start_id, end_id):
            if sid not in names:
                print(f"ERROR: {sid}.json not found in {config.EXTRACTED_DIR}")
                sys.exit(1)
        start_i = names.index(start_id)
        end_i   = names.index(end_id)
        if start_i > end_i:
            print(f"ERROR: {start_id} comes after {end_id} in sort order.")
            sys.exit(1)
        return all_jsons[start_i : end_i + 1]

    elif len(sys.argv) == 2:
        # Single file mode: python review.py s001
        survey_id = sys.argv[1].removesuffix(".json")
        json_path = config.EXTRACTED_DIR / f"{survey_id}.json"
        if not json_path.exists():
            print(f"ERROR: {json_path} not found.")
            sys.exit(1)
        return [json_path]

    else:
        # All extracted files
        if not all_jsons:
            print(f"No extracted JSON files found in {config.EXTRACTED_DIR}")
            sys.exit(0)
        return all_jsons


def main():
    # Strip --force flag before positional argument parsing
    force = "--force" in sys.argv
    if force:
        sys.argv.remove("--force")

    json_files = resolve_json_files()

    print(f"Generating review workbook(s) for {len(json_files)} survey(s)...\n")

    for json_path in json_files:
        survey_id = json_path.stem
        data       = json.loads(json_path.read_text(encoding="utf-8"))

        survey_type = data.get("survey_type", "IL")
        if survey_type not in config.SURVEY_TYPES:
            print(f"[WARN] {survey_id}: unknown survey_type '{survey_type}', defaulting to IL.")
            survey_type = "IL"

        output_dir = BASE_DIR / config.SURVEY_TYPES[survey_type]["output_dir"]
        output_dir.mkdir(parents=True, exist_ok=True)
        out_path = output_dir / f"review_{survey_id}.xlsx"

        if out_path.exists() and not force:
            print(f"[SKIP] {survey_id} — review workbook already exists. Use --force to overwrite.")
            continue

        fields     = data.get("fields", {})
        confidence = data.get("confidence", {})

        wb = build_workbook(survey_id, fields, confidence)
        wb.save(str(out_path))
        print(f"[OK]   Saved: {out_path}")


if __name__ == "__main__":
    main()

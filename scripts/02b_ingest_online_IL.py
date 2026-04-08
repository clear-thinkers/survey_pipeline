"""
02b_ingest_online_IL.py
Load the SurveyMonkey online export for the IL survey, normalize all
responses to the paper survey schema, and merge into
output/IL/survey_data_IL.csv.

- Appends rows with survey_id o001-oNNN and source="online".
- Paper rows get source="paper" added if not already present.
- Idempotent: drops any existing online rows (survey_id starts with "o")
  before re-appending.
- Sets all _conf columns to 1.0 for online rows (no OCR uncertainty).
- Fields not collected online (dob, first_initial, last_name) are left blank.

Usage:
    python scripts/02b_ingest_online_IL.py
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

BASE_DIR = Path(__file__).parent.parent
ONLINE_PATH = BASE_DIR / "data" / "online" / "Crawford IL Participant Survey 2026.xlsx"
CSV_PATH = BASE_DIR / "output" / "IL" / "survey_data_IL.csv"


def _norm(val):
    """Strip and normalize whitespace/quotes for reliable lookup."""
    if val is None:
        return ""
    text = " ".join(str(val).split())
    return (
        text.replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .strip()
    )


def s(row, col):
    try:
        return _norm(row[col])
    except IndexError:
        return ""


def lk(val, mapping, default=None):
    key = _norm(val)
    if not key:
        return default
    return mapping.get(key, default)


def parse_nps(val):
    text = _norm(val)
    if not text:
        return None
    match = re.match(r"^(\d{1,2})", text)
    if not match:
        return None
    score = int(match.group(1))
    return str(score) if 0 <= score <= 10 else None


def other_text(row, col, label):
    text = s(row, col)
    if not text or text == _norm(label):
        return None
    return text


def collect_multi(row, col_to_code, skip_cols=()):
    for col in skip_cols:
        if s(row, col):
            return ""
    tokens = []
    for col, code in col_to_code.items():
        if s(row, col) and code is not None:
            tokens.append(code)
    return " | ".join(tokens)


LIKERT5 = {
    "Never": "1",
    "Rarely": "2",
    "Sometimes": "3",
    "Often": "4",
    "All the time": "5",
    "All of the time": "5",
}

FREQ = {
    "Almost every day": "almost_every_day",
    "About once a week": "about_once_a_week",
    "1-2 times per month": "1_2_times_per_month",
    "Less than once a month": "less_than_once_a_month",
}

COMM_LEVEL = {
    "Not enough": "not_enough",
    "A good amount": "good_amount",
    "Too much": "too_much",
}

DURATION = {
    "Less than 6 months": "less_6mo",
    "6-12 months": "6_12mo",
    "More than 1 year, but less than 3 years": "1_3yr",
    "3 years or more": "3yr_plus",
}

SCHOOL_STATUS = {
    "High school": "high_school",
    "GED program": "ged",
    "College/career or technical school": "college_career",
    "College/Career or technical school": "college_career",
    "Not in school": "not_in_school",
}

EDUCATION = {
    "Some high school": "some_hs",
    "High school diploma or GED": "hs_diploma_ged",
    "Some college/career or technical school": "some_college",
    "College degree or Certification": "college_degree",
    "Graduate degree (ex. Master's, PhD)": "graduate",
}

EMPLOYMENT = {
    "Yes, full-time": "yes_full_time",
    "Yes, part-time": "yes_part_time",
    "No": "no",
}

JOB_TENURE = {
    "Less than 3 months": "less_3mo",
    "3-6 months": "3_6mo",
    "More than 6 months": "more_6mo",
}

JOB_SEEKING = {
    "Yes": "yes",
    "No": "no",
    "Not yet - I am in a job training program": "not_yet_in_training",
}

STAY_FOCUSED = {
    "Agree": "agree",
    "Somewhat agree": "somewhat_agree",
    "Disagree": "disagree",
    "Unsure": "unsure",
    "Not sure": "unsure",
}

RESPECT = {
    "Never": "never",
    "Rarely": "rarely",
    "Sometimes": "sometimes",
    "Often": "often",
    "All the time": "all_the_time",
    "All of the time": "all_the_time",
}

INDEPENDENCE = {
    "Agree": "agree",
    "Somewhat": "somewhat",
    "Somewhat agree": "somewhat",
    "Disagree": "disagree",
    "Unsure": "unsure",
}

AGE_RANGE = {
    "14-17 years old": "14_17",
    "18-20 years old": "18_20",
    "21-23 years old": "21_23",
}

Q6B_JOB_TYPES_COLS = {
    24: "retail_customer_service",
    25: "food_service",
    26: "office_admin",
    27: "healthcare_childcare_helping",
    28: "warehouse_construction_handson",
    29: "technology_creative",
}

Q7_COLS = {
    32: "childcare",
    33: "criminal_background",
    34: "no_references",
    35: "interview_skills",
    36: "no_diploma",
    37: "limited_experience",
    38: "mental_physical_health",
    39: "transportation",
    40: "drugs_alcohol",
    41: "not_getting_called",
}

Q8_COLS = {
    44: "found_better",
    45: "quit",
    46: "fired_attendance",
    47: "fired_performance",
    48: "seasonal",
}

Q9_BANK_COLS = {
    51: "checking",
    52: "savings",
    53: "had_in_past",
    54: "never_had",
}

Q9A_COLS = {
    55: "dont_know_how",
    56: "fees",
    57: "bad_credit",
    58: "not_enough_money",
    59: "min_balance_requirements",
    60: "no_trusted_adult",
    61: "tried_and_failed",
}

Q11_COLS = {
    67: "health_counseling",
    68: "positive_relationships",
    69: "handle_problems",
    70: "housing",
    71: "education",
    72: "job",
    73: "drivers_license",
    74: "everyday_skills",
    75: "decision_making",
    76: "vital_documents",
    77: "future",
    78: "parenting",
}

GENDER_COLS = {
    92: "Female",
    93: "Gender Nonconforming",
    94: "Genderqueer",
    95: "Male",
    96: "Non-binary",
    97: "Transgender Female",
    98: "Transgender Male",
    99: "Two-Spirit",
    100: "Prefer not to answer",
}

RACE_COLS = {
    103: "Black or of African or Caribbean Descent",
    104: "East Asian",
    105: "Hispanic or Latinx",
    106: "Native American or Indigenous peoples of America",
    107: "Native Hawaiian or Pacific Islander",
    108: "South Asian or Indian (Subcontinent)",
    109: "Southeast Asian",
    110: "Western Asian or Middle Eastern",
    111: "Other Asian",
    112: "White or of European Descent",
    113: "Multi-Racial",
    114: "Prefer not to answer",
}


def map_row(row, idx):
    survey_id = f"o{idx:03d}"

    coach_raw = s(row, 10)
    coach_name = coach_raw if coach_raw and coach_raw.lower() != "coach name:" else None

    q1_trustworthy = lk(s(row, 11), LIKERT5)
    q1_reliable = lk(s(row, 12), LIKERT5)
    q1_values_opinions = lk(s(row, 13), LIKERT5)
    q1_available = lk(s(row, 14), LIKERT5)
    q1_heard_understood = lk(s(row, 15), LIKERT5)
    q2 = lk(s(row, 16), FREQ)
    q3 = lk(s(row, 17), COMM_LEVEL)
    q4 = lk(s(row, 18), DURATION)
    q5 = lk(s(row, 19), SCHOOL_STATUS)
    q5a = lk(s(row, 20), EDUCATION) if q5 == "not_in_school" else None
    q6 = lk(s(row, 21), EMPLOYMENT)
    q6a = lk(s(row, 22), JOB_TENURE) if q6 in ("yes_part_time", "yes_full_time") else None
    q6b = lk(s(row, 23), JOB_SEEKING) if q6 == "no" else None

    q6b_tokens = []
    for col, code in Q6B_JOB_TYPES_COLS.items():
        if s(row, col):
            q6b_tokens.append(code)
    q6b_other = other_text(row, 30, "Other (please specify)")
    if q6b_other:
        q6b_tokens.append("other")
    q6b_job_types = " | ".join(q6b_tokens) if q6b == "yes" else ""
    q6b_job_types_other = q6b_other if q6b == "yes" else None

    q7_tokens = []
    q7_skip = bool(s(row, 31))
    if not q7_skip:
        for col, code in Q7_COLS.items():
            if s(row, col):
                q7_tokens.append(code)
        q7_other = other_text(row, 42, "Something else (please specify)")
        if q7_other:
            q7_tokens.append("something_else")
    else:
        q7_other = None
    q7 = " | ".join(q7_tokens)

    q8_tokens = []
    q8_skip = bool(s(row, 43))
    if not q8_skip:
        for col, code in Q8_COLS.items():
            if s(row, col):
                q8_tokens.append(code)
        q8_other = other_text(row, 49, "Other (please describe)")
        if q8_other:
            q8_tokens.append("other")
    else:
        q8_other = None
    q8 = " | ".join(q8_tokens)

    q8a_other = s(row, 50) or None
    q8a = ""

    q9 = collect_multi(row, Q9_BANK_COLS)
    q9a_tokens = []
    if any(token in q9.split(" | ") for token in ("had_in_past", "never_had")):
        for col, code in Q9A_COLS.items():
            if s(row, col):
                q9a_tokens.append(code)
        q9a_other = other_text(row, 62, "Other reason. Please explain:")
        if q9a_other:
            q9a_tokens.append("other")
        q9a_tried_failed = s(row, 63) or None
    else:
        q9a_other = None
        q9a_tried_failed = None
    q9a = " | ".join(q9a_tokens)

    q10 = lk(s(row, 64), STAY_FOCUSED)
    q10a = (s(row, 65) or None) if q10 in ("somewhat_agree", "disagree") else None

    q11_tokens = []
    q11_skip = bool(s(row, 66))
    if not q11_skip:
        for col, code in Q11_COLS.items():
            if s(row, col):
                q11_tokens.append(code)
        q11_other = other_text(row, 79, "Something else:")
        if q11_other:
            q11_tokens.append("something_else")
    else:
        q11_other = None
    q11 = " | ".join(q11_tokens)
    q11_none = s(row, 80) or None

    q12 = s(row, 81) or None
    q13 = lk(s(row, 82), RESPECT)
    q14 = lk(s(row, 83), RESPECT)
    q15_people_care = lk(s(row, 84), LIKERT5)
    q15_no_judgment = lk(s(row, 85), LIKERT5)
    q15_diversity_valued = lk(s(row, 86), LIKERT5)
    q15_treated_fairly = lk(s(row, 87), LIKERT5)
    q15_safe_sharing = lk(s(row, 88), LIKERT5)
    q16 = lk(s(row, 89), INDEPENDENCE)
    q17 = parse_nps(row[90])
    q18 = s(row, 91) or None

    gender_parts = [label for col, label in GENDER_COLS.items() if s(row, col)]
    gender_self = s(row, 101)
    if gender_self:
        gender_parts.append(gender_self)
    gender = " | ".join(gender_parts) if gender_parts else None

    age_range = lk(s(row, 102), AGE_RANGE)

    race_parts = [label for col, label in RACE_COLS.items() if s(row, col)]
    race_self = s(row, 115)
    if race_self:
        race_parts.append(race_self)
    race_ethnicity = " | ".join(race_parts)

    orient_raw = s(row, 116)
    orient_self = s(row, 117)
    if orient_raw == "Self describe:" and orient_self:
        sexual_orientation = orient_self
    else:
        sexual_orientation = orient_raw or None

    return {
        "survey_id": survey_id,
        "dob": None,
        "first_initial": None,
        "last_name": None,
        "coach_name": coach_name,
        "source": "online",
        "q1_trustworthy": q1_trustworthy,
        "q1_reliable": q1_reliable,
        "q1_values_opinions": q1_values_opinions,
        "q1_available": q1_available,
        "q1_heard_understood": q1_heard_understood,
        "q2_communication_frequency": q2,
        "q3_communication_level": q3,
        "q4_program_duration": q4,
        "q5_school_status": q5,
        "q5a_highest_education": q5a,
        "q6_employment_status": q6,
        "q6a_job_tenure": q6a,
        "q6b_job_seeking": q6b,
        "q6b_job_types": q6b_job_types,
        "q6b_job_types_other": q6b_job_types_other,
        "q7_barriers": q7,
        "q7_something_else_text": q7_other,
        "q8_left_job_reasons": q8,
        "q8_other_text": q8_other,
        "q8a_quit_reasons": q8a,
        "q8a_other_text": q8a_other,
        "q9_bank_account": q9,
        "q9a_no_account_reasons": q9a,
        "q9a_tried_failed_text": q9a_tried_failed,
        "q9a_other_text": q9a_other,
        "q10_stay_focused": q10,
        "q10a_what_would_help": q10a,
        "q11_program_helped": q11,
        "q11_something_else_text": q11_other,
        "q11_none_explain_text": q11_none,
        "q12_other_supports": q12,
        "q13_staff_respect": q13,
        "q14_peer_respect": q14,
        "q15_people_care": q15_people_care,
        "q15_no_judgment": q15_no_judgment,
        "q15_diversity_valued": q15_diversity_valued,
        "q15_treated_fairly": q15_treated_fairly,
        "q15_safe_sharing": q15_safe_sharing,
        "q16_gained_independence": q16,
        "q17_nps": q17,
        "q18_other_comments": q18,
        "gender": gender,
        "age_range": age_range,
        "race_ethnicity": race_ethnicity,
        "sexual_orientation": sexual_orientation,
    }


def main():
    if not ONLINE_PATH.exists():
        print(f"Online file not found: {ONLINE_PATH}")
        sys.exit(1)
    if not CSV_PATH.exists():
        print(f"Paper CSV not found: {CSV_PATH}")
        print("Run 02_compile.py first.")
        sys.exit(1)

    print(f"Loading paper CSV: {CSV_PATH}")
    df_paper = pd.read_csv(CSV_PATH, encoding="utf-8-sig", dtype=str)

    if "source" not in df_paper.columns:
        source_idx = list(df_paper.columns).index("coach_name") + 1
        df_paper.insert(source_idx, "source", "paper")
    else:
        df_paper["source"] = df_paper["source"].fillna("paper")

    paper_only = df_paper[~df_paper["survey_id"].str.startswith("o", na=False)].copy()
    dropped = len(df_paper) - len(paper_only)
    if dropped:
        print(f"  Dropped {dropped} existing online rows (will re-ingest)")

    conf_cols = [col for col in df_paper.columns if col.endswith("_conf")]

    print(f"Loading online xlsx: {ONLINE_PATH}")
    wb = openpyxl.load_workbook(ONLINE_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    data_rows = [row for row in all_rows[2:] if any(v is not None for v in row)]
    print(f"  Found {len(data_rows)} data rows")

    online_records = [map_row(row, idx) for idx, row in enumerate(data_rows, start=1)]
    df_online = pd.DataFrame(online_records)

    for col in conf_cols:
        df_online[col] = "1.0"

    for col in paper_only.columns:
        if col not in df_online.columns:
            df_online[col] = ""

    df_online = df_online.reindex(columns=paper_only.columns, fill_value="")

    df_merged = pd.concat([paper_only, df_online], ignore_index=True)
    df_merged = df_merged.fillna("")
    df_merged.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")

    paper_n = len(paper_only)
    online_n = len(df_online)
    print(f"\nMerged: {paper_n} paper + {online_n} online = {len(df_merged)} total rows")
    print(f"Saved: {CSV_PATH}")

    print("\nSanity check — source counts:")
    print(df_merged["source"].value_counts().to_string())
    print("\nSanity check — online age_range distribution:")
    print(df_merged[df_merged["source"] == "online"]["age_range"].value_counts().to_string())
    print("\nSanity check — online q6_employment_status distribution:")
    print(df_merged[df_merged["source"] == "online"]["q6_employment_status"].value_counts().to_string())


if __name__ == "__main__":
    main()
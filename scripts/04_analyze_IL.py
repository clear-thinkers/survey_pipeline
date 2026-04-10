"""
04_analyze_IL.py
Build output/IL/analysis_IL.xlsx and chart assets for the Crawford County IL report.

Usage:
    python scripts/04_analyze_IL.py
"""

# pyright: reportMissingImports=false

from __future__ import annotations

import sys
from collections import Counter
import importlib
from pathlib import Path
from textwrap import fill

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

matplotlib = importlib.import_module("matplotlib")
matplotlib.use("Agg")
plt = importlib.import_module("matplotlib.pyplot")
np = importlib.import_module("numpy")


BASE_DIR = Path(__file__).parent.parent
CSV_PATH = BASE_DIR / "output" / "IL" / "survey_data_IL.csv"
OUT_PATH = BASE_DIR / "output" / "IL" / "analysis_IL.xlsx"
CHARTS_DIR = BASE_DIR / "output" / "IL" / "charts"

AGE_MAP = {
    "14_17": "14-17 years old",
    "18_20": "18-20 years old",
    "21_23": "21-23 years old",
}
AGE_ORDER = ["14-17 years old", "18-20 years old", "21-23 years old", "Did not answer"]

COMM_FREQ_MAP = {
    "almost_every_day": "Almost every day",
    "about_once_a_week": "About once a week",
    "1_2_times_per_month": "1-2 times per month",
    "less_than_once_a_month": "Less than once a month",
}

COMM_LEVEL_MAP = {
    "not_enough": "Not enough",
    "good_amount": "Good amount",
    "too_much": "Too much",
}

EMPLOYMENT_MAP = {
    "yes_full_time": "Full time",
    "yes_part_time": "Part time",
    "no": "Unemployed",
}

JOB_TENURE_MAP = {
    "less_3mo": "Less than 3 months",
    "3_6mo": "3 to 6 months",
    "more_6mo": "More than 6 months",
}

JOB_TYPE_MAP = {
    "retail_customer_service": "Retail/customer service",
    "food_service": "Food service",
    "office_admin": "Office/administrative",
    "healthcare_childcare_helping": "Healthcare, childcare, or helping roles",
    "warehouse_construction_handson": "Warehouse, construction, or hands-on work",
    "technology_creative": "Technology or creative work",
    "other": "Other",
}

JOB_BARRIER_MAP = {
    "transportation": "Transportation issues",
    "limited_experience": "Limited work experience",
    "mental_physical_health": "Mental or physical health",
    "no_references": "Do not have good references",
    "interview_skills": "Interview skills",
    "no_diploma": "Lack of high school diploma/GED",
    "childcare": "Childcare/parenting challenges",
    "not_getting_called": "Applying and not getting called",
    "criminal_background": "Criminal background",
    "drugs_alcohol": "Use of drugs or alcohol",
    "something_else": "Something else",
}

LEFT_JOB_MAP = {
    "other": "Other",
    "found_better": "Found a better job",
    "quit": "Quit",
    "fired_attendance": "Fired: attendance/tardiness",
    "fired_performance": "Fired: performance or behavior",
    "seasonal": "Seasonal/temporary",
}

BANK_MAP = {
    "checking": "Checking account",
    "savings": "Savings account",
    "had_in_past": "Had an account in the past",
    "never_had": "Never had an account",
}

NO_ACCOUNT_REASON_MAP = {
    "dont_know_how": "Do not know how to open one",
    "fees": "Fees",
    "bad_credit": "Bad credit",
    "not_enough_money": "Do not have enough money",
    "min_balance_requirements": "Minimum balance requirements",
    "no_trusted_adult": "No trusted adult to help",
    "tried_and_failed": "Tried and failed",
    "other": "Other",
}

HELPED_MAP = {
    "health_counseling": "Access health care and/or counseling",
    "positive_relationships": "Establish positive relationships",
    "handle_problems": "Handle problems or challenging situations",
    "housing": "Find or maintain housing",
    "education": "Finish or further education",
    "job": "Get or keep a job",
    "drivers_license": "Get a driver's license",
    "parenting": "Improve parenting skills",
    "everyday_skills": "Learn everyday skills",
    "decision_making": "Develop decision-making skills",
    "vital_documents": "Obtain vital documents",
    "future": "Think about the future",
    "something_else": "Something else",
}

ENVIRONMENT_FIELDS = [
    ("q15_people_care", "People in the program care about my success"),
    ("q15_no_judgment", "I feel accepted without judgment"),
    ("q15_diversity_valued", "Diversity of backgrounds is valued"),
    ("q15_treated_fairly", "I am treated fairly"),
    ("q15_safe_sharing", "I feel safe sharing my thoughts"),
]

TRIVIAL_COMMENT_SET = {
    "no",
    "nope",
    "naw",
    "none",
    "none.",
    "n/a",
    "na",
}

TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FILL = PatternFill("solid", fgColor="2F5496")
SUB_FONT = Font(bold=True, italic=True)
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")

PALETTE = ["#355070", "#6D597A", "#B56576", "#E56B6F", "#EAAC8B"]


def split_pipe(value: str) -> list[str]:
    if value is None or str(value).strip() == "":
        return []
    return [part.strip() for part in str(value).split("|") if part.strip()]


def pct_str(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return ""
    return f"{round(100 * numerator / denominator)}%"


def age_label(code: str) -> str:
    return AGE_MAP.get(str(code).strip(), "Did not answer")


def clean_gender(value: str) -> str:
    text = str(value).strip()
    if text in {"Female", "Male"}:
        return text
    if not text:
        return "Did not answer"
    return "Gender nonconforming, Non-binary"


def clean_orientation(value: str) -> str:
    text = str(value).strip()
    if text in {"Gay or Lesbian", "Same Gender Loving"}:
        return "Gay, Lesbian, Same Gender Loving"
    if not text:
        return "Did not answer"
    return text


def race_once(value: str) -> str:
    tokens = split_pipe(value)
    if not tokens:
        return "Did not answer"
    if len(tokens) > 1:
        return "Multi-Racial"
    token = tokens[0]
    known = {
        "White or of European Descent",
        "Black or of African or Caribbean Descent",
        "Multi-Racial",
        "Native American or Indigenous peoples of America",
        "East Asian",
        "Hispanic or Latinx",
        "Native Hawaiian or Pacific Islander",
        "South Asian or Indian (Subcontinent)",
        "Southeast Asian",
        "Western Asian or Middle Eastern",
        "Other Asian",
        "Prefer not to answer",
    }
    return token if token in known else "Self described"


def count_with_total(series: pd.Series, order: list[str], label_col: str) -> pd.DataFrame:
    counts = series.value_counts()
    rows = [{label_col: label, "Count": int(counts.get(label, 0))} for label in order]
    table = pd.DataFrame(rows)
    table = table[table["Count"] > 0].reset_index(drop=True)
    total = int(table["Count"].sum())
    table.loc[len(table)] = {label_col: "Total", "Count": total}
    return table


def crosstab_multi_by_age(df: pd.DataFrame, field: str, label_map: dict[str, str]) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        for token in split_pipe(row[field]):
            rows.append({"Label": label_map.get(token, token), "Age": row["_age"]})
    if not rows:
        return pd.DataFrame(columns=["Item"] + AGE_ORDER + ["Total"])
    temp = pd.DataFrame(rows)
    ct = pd.crosstab(temp["Label"], temp["Age"])
    for age in AGE_ORDER:
        if age not in ct.columns:
            ct[age] = 0
    ct = ct[AGE_ORDER].copy()
    ct["Total"] = ct.sum(axis=1)
    ct = ct.reset_index().rename(columns={"Label": "Item"})
    ct = ct.sort_values(["Total", "Item"], ascending=[False, True]).reset_index(drop=True)
    return ct


def coach_table(df: pd.DataFrame) -> pd.DataFrame:
    fields = [
        ("q1_trustworthy", "Is trustworthy"),
        ("q1_reliable", "Is reliable"),
        ("q1_values_opinions", "Values my opinions about my life"),
        ("q1_available", "Is available when I need them"),
        ("q1_heard_understood", "Makes me feel heard and understood"),
    ]
    rows = []
    for field, label in fields:
        valid = df[field].astype(str).str.strip()
        valid = valid[valid != ""]
        top2 = valid.isin(["4", "5"]).sum()
        rows.append({
            "My Coach...": label,
            "n": int(len(valid)),
            "% Often or All the Time": pct_str(int(top2), int(len(valid))),
        })
    return pd.DataFrame(rows)


def communication_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    q3_counts = []
    valid_q3 = df["q3_communication_level"].astype(str).str.strip()
    valid_q3 = valid_q3[valid_q3 != ""]
    total_q3 = int(len(valid_q3))
    for code in ["good_amount", "not_enough", "too_much"]:
        count = int((valid_q3 == code).sum())
        q3_counts.append({
            "Communication Level": COMM_LEVEL_MAP[code],
            "Count": count,
            "Percent": pct_str(count, total_q3),
        })
    q3_table = pd.DataFrame(q3_counts)

    q2_counts = []
    valid_q2 = df["q2_communication_frequency"].astype(str).str.strip()
    valid_q2 = valid_q2[valid_q2 != ""]
    total_q2 = int(len(valid_q2))
    for code in ["almost_every_day", "about_once_a_week", "1_2_times_per_month", "less_than_once_a_month"]:
        count = int((valid_q2 == code).sum())
        q2_counts.append({
            "Reported Frequency of Communication": COMM_FREQ_MAP[code],
            "Number of Youth": count,
            "Percent": pct_str(count, total_q2),
        })
    q2_table = pd.DataFrame(q2_counts)
    q2_table.loc[len(q2_table)] = {
        "Reported Frequency of Communication": "Total",
        "Number of Youth": total_q2,
        "Percent": "100%",
    }
    return {
        "Communication Satisfaction": q3_table,
        "Reported Frequency": q2_table,
    }


def employment_status_table(df: pd.DataFrame) -> pd.DataFrame:
    valid = df[df["q6_employment_status"].astype(str).str.strip() != ""]
    total = int(len(valid))
    ft = int((valid["q6_employment_status"] == "yes_full_time").sum())
    pt = int((valid["q6_employment_status"] == "yes_part_time").sum())
    unemployed = int((valid["q6_employment_status"] == "no").sum())
    seeking = int(((valid["q6_employment_status"] == "no") & (valid["q6b_job_seeking"] == "yes")).sum())
    return pd.DataFrame([
        {"Employment Status": "Full time", "Total": ft},
        {"Employment Status": "Part time", "Total": pt},
        {"Employment Status": "Unemployed", "Total": unemployed},
        {"Employment Status": "Seeking Employment", "Total": seeking},
        {"Employment Status": "Total", "Total": total},
    ])


def job_tenure_table(df: pd.DataFrame) -> pd.DataFrame:
    employed = df[df["q6_employment_status"].isin(["yes_full_time", "yes_part_time"])]
    rows = []
    for code in ["less_3mo", "3_6mo", "more_6mo"]:
        ft = int(((employed["q6_employment_status"] == "yes_full_time") & (employed["q6a_job_tenure"] == code)).sum())
        pt = int(((employed["q6_employment_status"] == "yes_part_time") & (employed["q6a_job_tenure"] == code)).sum())
        rows.append({
            "Length of Employment": JOB_TENURE_MAP[code],
            "Full time": ft,
            "Part time": pt,
            "Total": ft + pt,
        })
    rows.append({
        "Length of Employment": "Total",
        "Full time": int((employed["q6_employment_status"] == "yes_full_time").sum()),
        "Part time": int((employed["q6_employment_status"] == "yes_part_time").sum()),
        "Total": int(len(employed)),
    })
    return pd.DataFrame(rows)


def employment_by_school_table(df: pd.DataFrame) -> pd.DataFrame:
    def school_group(code: str) -> str:
        if code == "high_school":
            return "High school"
        if code in {"ged", "college_career"}:
            return "GED or college/career"
        if code == "not_in_school":
            return "Not in school"
        return "Did not answer"

    rows = []
    order = ["High school", "GED or college/career", "Not in school", "Did not answer"]
    for label in order:
        sub = df[df["q5_school_status"].apply(school_group) == label]
        if sub.empty:
            continue
        rows.append({
            "School Enrollment": label,
            "Full time": int((sub["q6_employment_status"] == "yes_full_time").sum()),
            "Part time": int((sub["q6_employment_status"] == "yes_part_time").sum()),
            "Unemployed, seeking": int(((sub["q6_employment_status"] == "no") & (sub["q6b_job_seeking"] == "yes")).sum()),
            "Unemployed, not seeking or not yet": int(((sub["q6_employment_status"] == "no") & (sub["q6b_job_seeking"] != "yes")).sum()),
            "Total": int(len(sub)),
        })
    return pd.DataFrame(rows)


def job_barriers_table(df: pd.DataFrame) -> pd.DataFrame:
    respondents = df[df["q7_barriers"].astype(str).str.strip() != ""]
    denom = int(len(respondents))
    counts = Counter()
    for value in respondents["q7_barriers"]:
        counts.update(split_pipe(value))
    rows = []
    for code, count in sorted(counts.items(), key=lambda item: (-item[1], JOB_BARRIER_MAP.get(item[0], item[0]))):
        rows.append({
            "Reason youth have trouble finding jobs": JOB_BARRIER_MAP.get(code, code),
            "Total Youth": int(count),
            "Percent of Respondents": pct_str(int(count), denom),
        })
    return pd.DataFrame(rows)


def left_job_table(df: pd.DataFrame) -> pd.DataFrame:
    respondents = df[df["q8_left_job_reasons"].astype(str).str.strip() != ""]
    denom = int(len(respondents))
    counts = Counter()
    for value in respondents["q8_left_job_reasons"]:
        counts.update(split_pipe(value))
    rows = []
    for code, count in sorted(counts.items(), key=lambda item: (-item[1], LEFT_JOB_MAP.get(item[0], item[0]))):
        rows.append({
            "Reason youth left a job": LEFT_JOB_MAP.get(code, code),
            "Total Youth": int(count),
            "Percent of Respondents": pct_str(int(count), denom),
        })
    return pd.DataFrame(rows)


def program_impact_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    helped = crosstab_multi_by_age(df, "q11_program_helped", HELPED_MAP)
    if not helped.empty:
        helped = helped.rename(columns={"Item": "My coach or the IL program has helped me to..."})

    stay = []
    valid_stay = df["q10_stay_focused"].astype(str).str.strip()
    valid_stay = valid_stay[valid_stay != ""]
    total_stay = int(len(valid_stay))
    stay_map = {
        "agree": "Agree",
        "somewhat_agree": "Somewhat agree",
        "disagree": "Disagree",
        "unsure": "Unsure",
    }
    for code in ["agree", "somewhat_agree", "disagree", "unsure"]:
        count = int((valid_stay == code).sum())
        if count > 0:
            stay.append({
                "Support from IL helps me stay focused on my goals": stay_map[code],
                "Count": count,
                "Percent": pct_str(count, total_stay),
            })

    independence = []
    valid_ind = df["q16_gained_independence"].astype(str).str.strip()
    valid_ind = valid_ind[valid_ind != ""]
    total_ind = int(len(valid_ind))
    ind_map = {
        "agree": "Agree",
        "somewhat": "Somewhat agree",
        "disagree": "Disagree",
        "unsure": "Unsure",
    }
    for code in ["agree", "somewhat", "disagree", "unsure"]:
        count = int((valid_ind == code).sum())
        if count > 0:
            independence.append({
                "The IL program has helped me gain independence": ind_map[code],
                "Count": count,
                "Percent": pct_str(count, total_ind),
            })

    return {
        "Program Helped By Age": helped,
        "Stay Focused": pd.DataFrame(stay),
        "Gained Independence": pd.DataFrame(independence),
    }


def respect_environment_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    respect_rows = []
    top2 = {"often", "all_the_time"}
    respect_items = [
        ("q13_staff_respect", "Staff treat me with respect and acceptance"),
        ("q14_peer_respect", "Peers treat me with respect and acceptance"),
    ]
    for field, label in respect_items:
        valid = df[field].astype(str).str.strip()
        valid = valid[valid != ""]
        respect_rows.append({
            "Statement": label,
            "n": int(len(valid)),
            "% Often or All the Time": pct_str(int(valid.isin(top2).sum()), int(len(valid))),
            "All the time": int((valid == "all_the_time").sum()),
            "Often": int((valid == "often").sum()),
            "Sometimes": int((valid == "sometimes").sum()),
            "Rarely": int((valid == "rarely").sum()),
            "Never": int((valid == "never").sum()),
        })

    env_rows = []
    for field, label in ENVIRONMENT_FIELDS:
        valid = df[field].astype(str).str.strip()
        valid = valid[valid != ""]
        env_rows.append({
            "Program environment statement": label,
            "n": int(len(valid)),
            "% Top-2 Box (4-5)": pct_str(int(valid.isin(["4", "5"]).sum()), int(len(valid))),
        })

    return {
        "Respect and Acceptance": pd.DataFrame(respect_rows),
        "Program Environment": pd.DataFrame(env_rows),
    }


def banking_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    age_cols = [age for age in AGE_ORDER if int((df["_age"] == age).sum()) > 0]
    has_account = df["q9_bank_account"].apply(lambda value: any(token in split_pipe(value) for token in ["checking", "savings"]))

    header_row = {
        "Banking Status": "Number of Youth",
        **{age: int((df["_age"] == age).sum()) for age in age_cols},
        "Total": int(len(df)),
        "Percent of Total": "100%",
    }
    rows = [header_row]

    rows.append({
        "Banking Status": "Currently have a bank account",
        **{age: int(has_account[df["_age"] == age].sum()) for age in age_cols},
        "Total": int(has_account.sum()),
        "Percent of Total": pct_str(int(has_account.sum()), int(len(df))),
    })

    for code in ["checking", "savings", "had_in_past", "never_had"]:
        mask = df["q9_bank_account"].apply(lambda value: code in split_pipe(value))
        rows.append({
            "Banking Status": BANK_MAP[code],
            **{age: int(mask[df["_age"] == age].sum()) for age in age_cols},
            "Total": int(mask.sum()),
            "Percent of Total": pct_str(int(mask.sum()), int(len(df))),
        })

    reasons_df = df[df["q9a_no_account_reasons"].astype(str).str.strip() != ""]
    reason_counts = Counter()
    for value in reasons_df["q9a_no_account_reasons"]:
        reason_counts.update(split_pipe(value))
    reason_rows = []
    for code, count in sorted(reason_counts.items(), key=lambda item: (-item[1], NO_ACCOUNT_REASON_MAP.get(item[0], item[0]))):
        reason_rows.append({
            "Reason for not having an account": NO_ACCOUNT_REASON_MAP.get(code, code),
            "Count": int(count),
            "Percent of Respondents": pct_str(int(count), int(len(reasons_df))),
        })

    other_text_rows = []
    for _, row in df[df["q9a_other_text"].astype(str).str.strip() != ""].iterrows():
        other_text_rows.append({
            "Survey ID": row["survey_id"],
            "Other text": row["q9a_other_text"],
        })

    return {
        "Banking Status by Age": pd.DataFrame(rows),
        "Reasons for No Account": pd.DataFrame(reason_rows),
        "Open-Ended No Account Reasons": pd.DataFrame(other_text_rows),
    }


def nps_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    valid = pd.to_numeric(df["q17_nps"], errors="coerce").dropna().astype(int)
    total = int(len(valid))
    promoters = int((valid >= 9).sum())
    passives = int(((valid >= 7) & (valid <= 8)).sum())
    detractors = int((valid <= 6).sum())
    nps_score = round(100 * promoters / total - 100 * detractors / total) if total else ""

    summary = pd.DataFrame([
        {"Category": "Promoters (9-10)", "Count": promoters, "Percent": pct_str(promoters, total)},
        {"Category": "Passives (7-8)", "Count": passives, "Percent": pct_str(passives, total)},
        {"Category": "Detractors (0-6)", "Count": detractors, "Percent": pct_str(detractors, total)},
        {"Category": "Total Responded", "Count": total, "Percent": ""},
        {"Category": "NPS Score", "Count": nps_score, "Percent": ""},
    ])

    dist = pd.DataFrame({
        "Score": list(range(0, 11)),
        "Count": [int((valid == score).sum()) for score in range(0, 11)],
    })
    return {"NPS Summary": summary, "NPS Distribution": dist}


def comments_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    support_rows = []
    for _, row in df[df["q12_other_supports"].astype(str).str.strip() != ""].iterrows():
        support_rows.append({"Survey ID": row["survey_id"], "Requested Support": row["q12_other_supports"]})

    comment_rows = []
    for _, row in df[df["q18_other_comments"].astype(str).str.strip() != ""].iterrows():
        comment_rows.append({"Survey ID": row["survey_id"], "Comment": row["q18_other_comments"]})

    substantive = [
        row for row in comment_rows
        if row["Comment"].strip().lower() not in TRIVIAL_COMMENT_SET
    ]
    summary = pd.DataFrame([
        {"Metric": "Non-blank support requests", "Count": int(len(support_rows))},
        {"Metric": "Non-blank comments", "Count": int(len(comment_rows))},
        {"Metric": "Substantive comments", "Count": int(len(substantive))},
    ])

    return {
        "Summary": summary,
        "Other Supports Requested": pd.DataFrame(support_rows),
        "Other Comments": pd.DataFrame(comment_rows),
    }


def demographics_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    age_table = count_with_total(df["_age"], AGE_ORDER, "Age")
    gender_order = ["Female", "Male", "Gender nonconforming, Non-binary", "Did not answer"]
    gender_table = count_with_total(df["_gender"], gender_order, "Gender")

    race_order = [
        "White or of European Descent",
        "Black or of African or Caribbean Descent",
        "Multi-Racial",
        "Native American or Indigenous peoples of America",
        "East Asian",
        "Hispanic or Latinx",
        "Other Asian",
        "Prefer not to answer",
        "Self described",
        "Did not answer",
    ]
    race_table = count_with_total(df["_race_once"], race_order, "Race")

    orient_order = [
        "Heterosexual",
        "Bisexual",
        "Gay, Lesbian, Same Gender Loving",
        "Asexual",
        "Pansexual",
        "Mostly heterosexual",
        "I am not sure yet",
        "I don't understand the question",
        "Prefer not to answer",
        "Did not answer",
    ]
    orient_table = count_with_total(df["_orientation"], orient_order, "Sexual Orientation")

    return {
        "Age": age_table,
        "Gender": gender_table,
        "Race": race_table,
        "Sexual Orientation": orient_table,
    }


def write_section(ws, title: str, tables):
    if isinstance(tables, pd.DataFrame):
        tables = [("", tables)]
    elif isinstance(tables, dict):
        tables = list(tables.items())

    row = 1
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    row += 1

    for subtitle, df in tables:
        if df is None or df.empty:
            continue
        if subtitle:
            ws.cell(row=row, column=1, value=subtitle).font = SUB_FONT
            row += 1
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=row, column=col_idx, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        row += 1
        for _, data_row in df.iterrows():
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col_idx, value=value)
            row += 1
        row += 1


def autofit_columns(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def save_chart(fig, filename: str):
    CHARTS_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(CHARTS_DIR / filename, dpi=150, bbox_inches="tight")
    plt.close(fig)


def make_employment_chart(df: pd.DataFrame):
    if df.empty:
        return
    plot_df = df.copy()
    labels = plot_df["School Enrollment"].tolist()
    segments = [
        ("Full time", PALETTE[0]),
        ("Part time", PALETTE[1]),
        ("Unemployed, seeking", PALETTE[2]),
        ("Unemployed, not seeking or not yet", PALETTE[3]),
    ]
    fig, ax = plt.subplots(figsize=(8, 3.8))
    left = [0] * len(plot_df)
    for segment, color in segments:
        values = plot_df[segment].tolist()
        ax.barh(labels, values, left=left, color=color, label=segment)
        left = [base + value for base, value in zip(left, values)]
    ax.set_title("Employment status by school enrollment")
    ax.set_xlabel("Number of youth")
    ax.grid(axis="x", color="#DDDDDD", linewidth=0.8)
    ax.set_axisbelow(True)
    ax.legend(frameon=False, fontsize=8, loc="lower right")
    save_chart(fig, "chart_01_employment_by_school.png")


def make_program_helped_chart(df: pd.DataFrame):
    """Table-with-embedded-bars chart matching 412YZ chart 12 style.

    Columns: 14-17 years old | 18-23 years old (combined) | Total
    Base: youth who answered both q11_program_helped and the age question.
    """
    TEXT_COLOR = "#1F2933"
    BAR_FILL = "#5B88C2"
    HEADER_CLR = "#DCE6F1"
    BORDER = "#6B7280"

    CHART_AGE_COLS = ["14-17 years old", "18-23 years old"]

    # Restrict to those who answered q11 AND gave a known age
    answered = df[
        (df["q11_program_helped"].astype(str).str.strip() != "") &
        (df["_age"].isin(["14-17 years old", "18-20 years old", "21-23 years old"]))
    ].copy()
    if answered.empty:
        return

    # Combine 18-20 and 21-23 into a single bucket
    answered["_age_chart"] = answered["_age"].apply(
        lambda a: "18-23 years old" if a in ("18-20 years old", "21-23 years old") else a
    )

    n_by_col = {col: int((answered["_age_chart"] == col).sum()) for col in CHART_AGE_COLS}
    n_total = sum(n_by_col.values())
    if n_total == 0:
        return

    chart_rows = []
    for code, label in HELPED_MAP.items():
        counts_by_col = []
        total_count = 0
        for col in CHART_AGE_COLS:
            col_df = answered[answered["_age_chart"] == col]
            count = int(sum(code in split_pipe(v) for v in col_df["q11_program_helped"]))
            counts_by_col.append(count)
            total_count += count
        if total_count == 0:
            continue
        chart_rows.append({
            "label": label,
            "pcts": [
                100.0 * counts_by_col[i] / n_by_col[col] if n_by_col[col] else 0.0
                for i, col in enumerate(CHART_AGE_COLS)
            ],
            "total_count": total_count,
            "total_pct": 100.0 * total_count / n_total,
        })
    chart_rows.sort(key=lambda r: (-r["total_count"], r["label"]))

    active_ages = CHART_AGE_COLS

    # Pre-wrap labels and compute per-row heights
    WRAP_WIDTH = 38
    BASE_CELL_H = 0.052
    LINE_EXTRA_H = 0.046  # extra height per additional wrapped line
    wrapped_rows = []
    for row in chart_rows:
        display = fill(row["label"], WRAP_WIDTH)
        n_lines = display.count("\n") + 1
        row_h = BASE_CELL_H + (n_lines - 1) * LINE_EXTRA_H
        wrapped_rows.append({**row, "display_label": display, "row_h": row_h})
    extra_lines = sum(r["row_h"] - BASE_CELL_H for r in wrapped_rows)

    n_data_cols = len(active_ages)
    fig_h = max(5.4, 1.7 + 0.31 * len(chart_rows) + extra_lines * 6.0)
    fig, ax = plt.subplots(figsize=(11.0, fig_h))
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")
    fig.patch.set_facecolor("white")

    col_left = 0.02
    reason_w = 0.30
    data_w = 0.19
    total_w = 0.17
    cell_h = BASE_CELL_H
    header_h = 0.064
    top = 0.90

    ax.text(0.5, 0.985,
            "My Coach or the IL Program has Helped Me To... (by Age)",
            ha="center", va="top", fontsize=13.5, fontweight="bold", color=TEXT_COLOR)

    # Column x-start positions and widths
    x_starts = [col_left] + [
        col_left + reason_w + i * data_w for i in range(n_data_cols)
    ] + [col_left + reason_w + n_data_cols * data_w]
    col_widths = [reason_w] + [data_w] * n_data_cols + [total_w]
    col_labels = ["Age"] + active_ages + ["Total"]

    # Header row
    y = top - header_h
    for x0, w, label in zip(x_starts, col_widths, col_labels):
        fill_clr = "white" if label == "Age" else HEADER_CLR
        ax.add_patch(plt.Rectangle((x0, y), w, header_h,
                                   facecolor=fill_clr, edgecolor=BORDER, linewidth=0.7))
        ax.text(x0 + w / 2, y + header_h / 2, label,
                ha="center", va="center", fontsize=9.4, color=TEXT_COLOR,
                fontweight="bold" if label != "Age" else None,
                style="italic" if label == "Age" else None)

    # Number of Youth row
    y -= cell_h
    n_labels = ["Number of Youth"] + [str(n_by_col[col]) for col in active_ages] + [str(n_total)]
    for x0, w, label in zip(x_starts, col_widths, n_labels):
        ax.add_patch(plt.Rectangle((x0, y), w, cell_h,
                                   facecolor="white", edgecolor=BORDER, linewidth=0.7))
        ax.text(x0 + w / 2, y + cell_h / 2, label,
                ha="center", va="center", fontsize=10, color=TEXT_COLOR,
                fontweight="bold" if label != "Number of Youth" else None,
                style="italic" if label == "Number of Youth" else None)

    # Data + total column descriptors for bar rendering
    bar_cols = [
        (col_left + reason_w + i * data_w, data_w, i) for i in range(n_data_cols)
    ] + [(col_left + reason_w + n_data_cols * data_w, total_w, "total")]

    # Data rows (variable height for wrapped labels)
    for row in wrapped_rows:
        rh = row["row_h"]
        y -= rh
        ax.add_patch(plt.Rectangle((col_left, y), reason_w, rh,
                                   facecolor="white", edgecolor=BORDER, linewidth=0.7))
        ax.text(col_left + 0.006, y + rh / 2, row["display_label"],
                ha="left", va="center", fontsize=9.5, color=TEXT_COLOR)
        rh = row["row_h"]
        for x0, w, idx in bar_cols:
            ax.add_patch(plt.Rectangle((x0, y), w, rh,
                                       facecolor="white", edgecolor=BORDER, linewidth=0.7))
            pct_val = row["total_pct"] if idx == "total" else row["pcts"][idx]
            if pct_val > 0:
                bar_margin = 0.003
                label_reserve = 0.028  # keep this gap clear for the pct label
                usable_w = w - 2 * bar_margin - label_reserve
                bar_w = usable_w * min(pct_val / 100.0, 1.0)
                bar_inner_h = min(cell_h - 0.012, rh * 0.65)
                ax.add_patch(plt.Rectangle(
                    (x0 + bar_margin, y + (rh - bar_inner_h) / 2), bar_w, bar_inner_h,
                    facecolor=BAR_FILL, edgecolor="none"))
            ax.text(x0 + w - 0.008, y + rh / 2, f"{int(round(pct_val))}%",
                    ha="right", va="center", fontsize=9.5, color=TEXT_COLOR)

    save_chart(fig, "chart_02_program_helped_by_age.png")


def make_environment_chart(df: pd.DataFrame):
    """100% stacked horizontal bar chart for program environment, matching 412YZ chart 14."""
    TEXT_COLOR = "#1F2933"
    GRID_COLOR = "#D9DDE3"
    AXIS_COLOR = "#9AA5B1"
    env_order = ["5", "4", "3", "2", "1"]
    env_labels = {"5": "All the time", "4": "Often", "3": "Sometimes", "2": "Rarely", "1": "Never"}
    env_colors = {
        "5": "#2E568F",
        "4": "#2E86AB",
        "3": "#D9DDE3",
        "2": "#F4A261",
        "1": "#C1440E",
    }

    env_counts = {}
    env_totals = {}
    for field, label in ENVIRONMENT_FIELDS:
        valid = df[df[field].astype(str).str.strip() != ""][field].astype(str)
        env_totals[label] = len(valid)
        env_counts[label] = {code: int((valid == code).sum()) for code in env_order}

    active_items = [(f, lbl) for f, lbl in ENVIRONMENT_FIELDS if env_totals.get(lbl, 0) > 0]
    if not active_items:
        return

    wrapped_labels = [fill(lbl, 22) for _, lbl in active_items]
    row_spacing = 0.80  # taller rows to fit two-line labels
    fig, ax = plt.subplots(figsize=(7.8, 1.0 + len(active_items) * row_spacing))
    fig.patch.set_facecolor("white")
    y_pos = np.arange(len(active_items)) * row_spacing
    lefts = np.zeros(len(active_items))
    top2_pcts = []

    for _, label in active_items:
        total = env_totals[label]
        top2_pcts.append(
            100 * (env_counts[label]["5"] + env_counts[label]["4"]) / total if total else 0
        )

    for code in env_order:
        pcts = []
        for _, label in active_items:
            total = env_totals[label]
            pct_val = 100 * env_counts[label][code] / total if total else 0
            pcts.append(pct_val)
        bars = ax.barh(y_pos, pcts, left=lefts, height=0.38,
                       color=env_colors[code], label=env_labels[code], zorder=3)
        for i, (bar, pval) in enumerate(zip(bars, pcts)):
            if pval >= 5:
                x_text = lefts[i] + pval / 2
                text_color = "white" if code in ("5", "4", "1") else TEXT_COLOR
                ax.text(x_text, bar.get_y() + bar.get_height() / 2,
                        f"{int(round(pval))}%", va="center", ha="center",
                        fontsize=10, color=text_color)
        lefts += np.array(pcts)

    for i, top2_pct in enumerate(top2_pcts):
        y_line = y_pos[i] + 0.30
        ax.plot([0, top2_pct], [y_line, y_line], color=TEXT_COLOR, linewidth=1)
        ax.plot([0, 0], [y_line - 0.04, y_line + 0.04], color=TEXT_COLOR, linewidth=1)
        ax.plot([top2_pct, top2_pct], [y_line - 0.04, y_line + 0.04], color=TEXT_COLOR, linewidth=1)
        ax.text(top2_pct + 1.3, y_line, f"Top-2 {int(round(top2_pct))}%",
                va="center", ha="left", fontsize=10, color=TEXT_COLOR)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(wrapped_labels, fontsize=10, color=TEXT_COLOR)
    ax.set_xlim(0, 100)
    ax.set_title("How Do Youth Rate the Program Environment?\n(1-5 scale; top-2 shown below each item)",
                 fontsize=10, fontweight="bold", color=TEXT_COLOR, pad=55)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["bottom"].set_color(AXIS_COLOR)
    ax.spines["bottom"].set_visible(False)
    ax.xaxis.grid(False)
    ax.xaxis.set_visible(False)
    ax.yaxis.grid(False)
    ax.set_axisbelow(True)
    ax.tick_params(axis="y", colors=TEXT_COLOR)
    ax.invert_yaxis()
    ax.legend(loc="lower center", bbox_to_anchor=(0.5, 1.04), ncol=5, fontsize=9, frameon=False,
              handlelength=0.8, handletextpad=0.3, columnspacing=1.2)
    save_chart(fig, "chart_04_environment_ratings.png")


def make_nps_chart(df: pd.DataFrame):
    if df.empty:
        return
    colors = []
    for score in df["Score"]:
        if score <= 6:
            colors.append("#B56576")
        elif score <= 8:
            colors.append("#A9B4C2")
        else:
            colors.append("#355070")
    fig, ax = plt.subplots(figsize=(7.5, 2.8))
    ax.bar(df["Score"].astype(str), df["Count"], color=colors)
    ax.set_title("Likelihood of recommending the IL program")
    ax.set_xlabel("0-10 rating")
    ax.set_ylabel("Number of youth")
    ax.grid(axis="y", color="#DDDDDD", linewidth=0.8)
    ax.set_axisbelow(True)
    save_chart(fig, "chart_03_nps.png")


def main():
    if not CSV_PATH.exists():
        print(f"CSV not found: {CSV_PATH}")
        sys.exit(1)

    df = pd.read_csv(CSV_PATH, dtype=str).fillna("")
    df["_age"] = df["age_range"].apply(age_label)
    df["_gender"] = df["gender"].apply(clean_gender)
    df["_orientation"] = df["sexual_orientation"].apply(clean_orientation)
    df["_race_once"] = df["race_ethnicity"].apply(race_once)

    sections = [
        ("01_demographics", "1. Survey Respondent Demographics", demographics_tables(df)),
        ("02_coach", "2. Relationships with Coach", coach_table(df)),
        ("03_communication", "3. Communication", communication_tables(df)),
        (
            "04_employment",
            "4. Employment and Education",
            {
                "Employment Status": employment_status_table(df),
                "Length of Employment for Youth Currently Employed": job_tenure_table(df),
                "Employment Status by School Enrollment": employment_by_school_table(df),
                "Reasons Youth Have Trouble Finding Jobs": job_barriers_table(df),
                "Reasons Youth Left a Job": left_job_table(df),
            },
        ),
        ("05_program_impact", "5. Program Impact", program_impact_tables(df)),
        ("06_respect_environment", "6. Respect and Environment", respect_environment_tables(df)),
        ("07_banking", "7. Banking", banking_tables(df)),
        ("08_nps", "8. Net Promoter Score", nps_tables(df)),
        ("09_comments", "9. Additional Comments", comments_tables(df)),
    ]

    CHARTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    for sheet_name, title, tables in sections:
        ws = wb.create_sheet(sheet_name)
        write_section(ws, title, tables)
        autofit_columns(ws)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)

    make_employment_chart(employment_by_school_table(df))
    make_program_helped_chart(df)
    make_environment_chart(df)
    nps_data = nps_tables(df)
    make_nps_chart(nps_data["NPS Distribution"])

    print(f"Saved workbook: {OUT_PATH}")
    for chart in sorted(CHARTS_DIR.glob("chart_*.png")):
        print(f"Saved chart: {chart}")


if __name__ == "__main__":
    main()
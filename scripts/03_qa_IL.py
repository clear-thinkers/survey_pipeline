"""
Run QA validation on output/IL/survey_data_IL.csv.

Produces two outputs:
  output/IL/flagged_IL.csv       - machine-readable issue log
  output/IL/qa_questions_IL.xlsx - reviewer workbook with QA questions,
                                   accepted items, and a summary tab.

Usage:
    python scripts/03_qa_IL.py
"""

import sys
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = Path(__file__).parent.parent

# ---------------------------------------------------------------------------
# Allowed values from extraction_prompt_IL.txt
# ---------------------------------------------------------------------------

SINGLE_SELECT = {
    "q2_communication_frequency": {
        "almost_every_day", "about_once_a_week", "1_2_times_per_month", "less_than_once_a_month"
    },
    "q3_communication_level": {"not_enough", "good_amount", "too_much"},
    "q4_program_duration": {"less_6mo", "6_12mo", "1_3yr", "3yr_plus"},
    "q5_school_status": {"high_school", "ged", "college_career", "graduate", "not_in_school"},
    "q5a_highest_education": {"some_hs", "hs_diploma_ged", "some_college", "college_degree", "graduate"},
    "q6_employment_status": {"yes_part_time", "yes_full_time", "no"},
    "q6a_job_tenure": {"less_3mo", "3_6mo", "more_6mo"},
    "q6b_job_seeking": {"yes", "no", "not_yet_in_training"},
    "q10_stay_focused": {"agree", "somewhat_agree", "disagree", "unsure"},
    "q13_staff_respect": {"never", "rarely", "sometimes", "often", "all_the_time"},
    "q14_peer_respect": {"never", "rarely", "sometimes", "often", "all_the_time"},
    "q16_gained_independence": {"agree", "somewhat", "disagree", "unsure"},
    "age_range": {"14_17", "18_20", "21_23"},
}

MULTI_SELECT = {
    "q6b_job_types": {
        "retail_customer_service", "food_service", "office_admin",
        "healthcare_childcare_helping", "warehouse_construction_handson",
        "technology_creative", "other"
    },
    "q7_barriers": {
        "childcare", "criminal_background", "no_references", "interview_skills",
        "no_diploma", "limited_experience", "mental_physical_health",
        "transportation", "drugs_alcohol", "not_getting_called", "something_else"
    },
    "q8_left_job_reasons": {
        "found_better", "quit", "fired_attendance", "fired_performance", "seasonal", "other"
    },
    "q8a_quit_reasons": {
        "low_pay_hours", "schedule_conflict", "lack_of_support", "poor_conditions",
        "mental_emotional_health", "transportation", "not_good_fit", "personal_family", "other"
    },
    "q9_bank_account": {"checking", "savings", "had_in_past", "never_had"},
    "q9a_no_account_reasons": {
        "dont_know_how", "fees", "bad_credit", "not_enough_money",
        "min_balance_requirements", "no_trusted_adult", "tried_and_failed", "other"
    },
    "q11_program_helped": {
        "health_counseling", "positive_relationships", "handle_problems", "housing",
        "education", "job", "drivers_license", "parenting", "everyday_skills",
        "decision_making", "vital_documents", "future", "something_else"
    },
}

LIKERT_FIELDS = [
    "q1_trustworthy", "q1_reliable", "q1_values_opinions", "q1_available", "q1_heard_understood",
    "q15_people_care", "q15_no_judgment", "q15_diversity_valued", "q15_treated_fairly", "q15_safe_sharing",
]

REQUIRED_FIELDS = [
    "q1_trustworthy", "q2_communication_frequency", "q3_communication_level",
    "q4_program_duration", "q5_school_status", "q6_employment_status",
    "q10_stay_focused", "q13_staff_respect", "q14_peer_respect",
    "q16_gained_independence", "q17_nps", "age_range", "gender",
]

CONDITIONAL = {
    "q5a_highest_education": ("q5_school_status", {"not_in_school"}),
    "q6a_job_tenure": ("q6_employment_status", {"yes_part_time", "yes_full_time"}),
    "q6b_job_seeking": ("q6_employment_status", {"no"}),
    "q10a_what_would_help": ("q10_stay_focused", {"somewhat_agree", "disagree"}),
}

KNOWN_GENDER_LABELS = {
    "Female", "Male", "Non-binary", "Transgender Male", "Transgender Female",
    "Genderqueer", "Gender Nonconforming", "Two-Spirit", "Prefer not to answer",
}

STANDARD_GENDER_VALUES = [
    "Female",
    "Male",
    "Non-binary",
    "Transgender Male",
    "Transgender Female",
    "Genderqueer",
    "Gender Nonconforming",
    "Two-Spirit",
    "Prefer not to answer",
]

KNOWN_RACE_LABELS = {
    "Black or of African or Caribbean Descent",
    "East Asian",
    "Hispanic or Latinx",
    "Native American or Indigenous peoples of America",
    "Native Hawaiian or Pacific Islander",
    "South Asian or Indian (Subcontinent)",
    "Southeast Asian",
    "Western Asian or Middle Eastern",
    "Other Asian",
    "White or of European Descent",
    "Multi-Racial",
    "Prefer not to answer",
}

STANDARD_RACE_VALUES = [
    "Black or of African or Caribbean Descent",
    "East Asian",
    "Hispanic or Latinx",
    "Native American or Indigenous peoples of America",
    "Native Hawaiian or Pacific Islander",
    "South Asian or Indian (Subcontinent)",
    "Southeast Asian",
    "Western Asian or Middle Eastern",
    "Other Asian",
    "White or of European Descent",
    "Multi-Racial",
    "Prefer not to answer",
]

KNOWN_ORIENTATION_LABELS = {
    "Heterosexual",
    "Gay or Lesbian",
    "Bisexual",
    "Asexual",
    "Pansexual",
    "Queer",
    "Demisexual",
    "Mostly heterosexual",
    "I am not sure yet",
    "I don't understand the question",
    "Prefer not to answer",
    "Same Gender Loving",
}

STANDARD_ORIENTATION_VALUES = [
    "Heterosexual",
    "Gay or Lesbian",
    "Bisexual",
    "Asexual",
    "Pansexual",
    "Queer",
    "Demisexual",
    "Mostly heterosexual",
    "I am not sure yet",
    "I don't understand the question",
    "Same Gender Loving",
    "Prefer not to answer",
]

# Demographic normalization defaults.
# Each entry becomes a prefilled QA row so the reviewer only needs to confirm it.
GENDER_NORMALIZATION = {
    "Prefer not to say": {
        "normalized": "Prefer not to answer",
        "scope": "all_surveys",
        "detail": "alternate wording for the printed nonresponse option",
    },
}

RACE_NORMALIZATION = {
    "White": {
        "normalized": "White or of European Descent",
        "scope": "all_surveys",
        "detail": "shorthand variant of the printed race label",
    },
    "White and Native American": {
        "normalized": "Multi-Racial",
        "scope": "this_survey",
        "detail": "free-text combination indicates multiple racial identities; mapping to the printed multiracial label",
    },
}

ORIENTATION_NORMALIZATION = {
    "Straight": {
        "normalized": "Heterosexual",
        "scope": "all_surveys",
        "detail": "common shorthand for the printed label",
    },
    "Demi-Sexual": {
        "normalized": "Demisexual",
        "scope": "all_surveys",
        "detail": "hyphenated spelling variant",
    },
    "Demi Sexye": {
        "normalized": "Demisexual",
        "scope": "all_surveys",
        "detail": "likely OCR misspelling of Demisexual",
    },
}

LOW_CONF_THRESHOLD = 0.75


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def split_pipe(val) -> list[str]:
    if val is None or str(val).strip() == "":
        return []
    return [token.strip() for token in str(val).split("|") if token.strip()]


def is_blank(val) -> bool:
    return val is None or str(val).strip() == ""


def has_value(row, field: str) -> bool:
    value = row.get(field, "")
    if is_blank(value):
        return False
    if field in MULTI_SELECT:
        return bool(split_pipe(value))
    return True


def is_online(row) -> bool:
    return str(row.get("source", "")).strip().lower() == "online"


def make_issue(
    survey_id,
    field,
    value,
    issue_type,
    detail,
    question,
    suggestion="",
    action="",
    corrected_value="",
    apply_to="",
):
    return {
        "survey_id": survey_id,
        "field": field,
        "current_value": "" if is_blank(value) else str(value),
        "issue_type": issue_type,
        "detail": detail,
        "question_for_reviewer": question,
        "suggested_correction": suggestion,
        "action": action,
        "corrected_value": corrected_value,
        "apply_to": apply_to,
        "notes": "",
    }


def make_recode_issue(survey_id, field, value, issue_type, detail, question, corrected_value, apply_to="this_survey"):
    return make_issue(
        survey_id,
        field,
        value,
        issue_type,
        detail,
        question,
        suggestion=corrected_value,
        action="recode",
        corrected_value=corrected_value,
        apply_to=apply_to,
    )


def make_accept_issue(survey_id, field, value, issue_type, detail, question, suggestion=""):
    return make_issue(
        survey_id,
        field,
        value,
        issue_type,
        detail,
        question,
        suggestion=suggestion,
        action="accept",
        apply_to="this_survey",
    )


def allowed_demographic_values(field: str) -> str:
    if field == "gender":
        return ", ".join(STANDARD_GENDER_VALUES)
    if field == "race_ethnicity":
        return ", ".join(STANDARD_RACE_VALUES)
    if field == "sexual_orientation":
        return ", ".join(STANDARD_ORIENTATION_VALUES)
    return ""


def make_demographic_recode_issue(survey_id, field, value, mapping):
    corrected = mapping["normalized"]
    return make_recode_issue(
        survey_id,
        field,
        value,
        "F_likely_typo",
        f"'{value}' is {mapping['detail']}",
        f"[{survey_id}] {field}='{value}'. Suggested standardization is '{corrected}'. Review the default and change it only if needed.",
        corrected,
        apply_to=mapping.get("scope", "this_survey"),
    )


def make_demographic_accept_issue(survey_id, field, value, detail):
    return make_accept_issue(
        survey_id,
        field,
        value,
        "F_self_describe" if field != "race_ethnicity" else "F_unrecognized_race_label",
        detail,
        f"[{survey_id}] {field}='{value}'. Default action is accept as-is. Review only if you want to override the captured wording for analysis.",
        suggestion="accept_as_is",
    )


def make_demographic_review_issue(survey_id, field, value, detail):
    allowed = allowed_demographic_values(field)
    issue_type = "F_self_describe" if field != "race_ethnicity" else "F_unrecognized_race_label"
    return make_issue(
        survey_id,
        field,
        value,
        issue_type,
        detail,
        f"[{survey_id}] {field}='{value}'. Reviewer approval required. Choose the best standard value in corrected_value, "
        f"or explicitly set action=accept only if the wording should be preserved.",
        suggestion=f"Allowed values: {allowed}",
        action="",
        corrected_value="",
        apply_to="this_survey",
    )


def add_auto_clear_issue(issues, sid, field, row, parent_desc):
    issues.append(make_issue(
        sid,
        field,
        row.get(field, ""),
        "D_auto_clear",
        f"{field} is non-blank but {parent_desc}",
        f"[{sid}] {field}='{row.get(field, '')}' but {parent_desc} - auto-clearing {field}.",
        "clear",
        action="clear",
        apply_to="this_survey",
    ))


def comma_parts(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


# ---------------------------------------------------------------------------
# Rule A - Type / range validation
# ---------------------------------------------------------------------------

def check_A(row, issues):
    sid = row["survey_id"]

    for field in LIKERT_FIELDS:
        value = row.get(field, "")
        if is_blank(value):
            continue
        try:
            numeric = int(float(value))
            assert 1 <= numeric <= 5
        except (ValueError, TypeError, AssertionError):
            issues.append(make_issue(
                sid,
                field,
                value,
                "A_invalid_range",
                f"Expected integer 1-5, got '{value}'",
                f"[{sid}] {field} = '{value}'. Valid range is 1-5. What is the correct value?",
            ))

    nps = row.get("q17_nps", "")
    if not is_blank(nps):
        try:
            numeric = int(float(nps))
            assert 0 <= numeric <= 10
        except (ValueError, TypeError, AssertionError):
            issues.append(make_issue(
                sid,
                "q17_nps",
                nps,
                "A_invalid_range",
                f"Expected integer 0-10, got '{nps}'",
                f"[{sid}] q17_nps = '{nps}'. Valid range is 0-10. What is the correct value?",
            ))


# ---------------------------------------------------------------------------
# Rule B - Categorical allowed-values
# ---------------------------------------------------------------------------

def check_B(row, issues):
    sid = row["survey_id"]
    for field, allowed in SINGLE_SELECT.items():
        value = str(row.get(field, "")).strip()
        if not value:
            continue
        if value not in allowed:
            issues.append(make_issue(
                sid,
                field,
                value,
                "B_invalid_category",
                f"'{value}' not in allowed set",
                f"[{sid}] {field} = '{value}'. Not a recognized code. What should this be?",
                f"Allowed: {', '.join(sorted(allowed))}",
            ))


# ---------------------------------------------------------------------------
# Rule C - Array token allowed-values
# ---------------------------------------------------------------------------

def check_C(row, issues):
    sid = row["survey_id"]
    for field, allowed in MULTI_SELECT.items():
        for token in split_pipe(row.get(field, "")):
            if token not in allowed:
                issues.append(make_issue(
                    sid,
                    field,
                    token,
                    "C_invalid_token",
                    f"Token '{token}' not in allowed codes for {field}",
                    f"[{sid}] {field} contains '{token}'. Not a recognized code. What should this be?",
                    f"Allowed: {', '.join(sorted(allowed))}",
                ))


# ---------------------------------------------------------------------------
# Rule D - Conditional logic
# ---------------------------------------------------------------------------

def check_D(row, issues):
    sid = row["survey_id"]

    for child, (parent, allowed_vals) in CONDITIONAL.items():
        parent_val = str(row.get(parent, "")).strip()
        condition_met = parent_val in allowed_vals
        child_filled = has_value(row, child)

        if not condition_met and child_filled:
            add_auto_clear_issue(issues, sid, child, row, f"{parent}='{parent_val}' (condition not met)")

        if condition_met and not child_filled and child != "q10a_what_would_help":
            issues.append(make_issue(
                sid,
                child,
                "",
                "D_conditional_missing",
                f"{child} is blank but {parent}='{parent_val}' (condition is met)",
                f"[{sid}] {parent}='{parent_val}' so {child} is expected, but it's blank. "
                f"Was this intentionally left blank on the survey?",
            ))

    q6b_yes = str(row.get("q6b_job_seeking", "")).strip() == "yes"
    q6b_types_filled = has_value(row, "q6b_job_types")
    q6b_other_filled = has_value(row, "q6b_job_types_other")
    if not q6b_yes:
        if q6b_types_filled:
            add_auto_clear_issue(issues, sid, "q6b_job_types", row, "q6b_job_seeking is not 'yes'")
        if q6b_other_filled:
            add_auto_clear_issue(issues, sid, "q6b_job_types_other", row, "q6b_job_seeking is not 'yes'")
    elif not q6b_types_filled:
        issues.append(make_issue(
            sid,
            "q6b_job_types",
            "",
            "D_conditional_missing",
            "q6b_job_types is blank but q6b_job_seeking='yes'",
            f"[{sid}] q6b_job_seeking='yes' so q6b_job_types is expected, but it's blank. "
            f"Was this intentionally left blank on the survey?",
        ))

    q7_tokens = set(split_pipe(row.get("q7_barriers", "")))
    if "something_else" not in q7_tokens and has_value(row, "q7_something_else_text"):
        add_auto_clear_issue(issues, sid, "q7_something_else_text", row, "'something_else' is not selected in q7_barriers")

    q8_tokens = set(split_pipe(row.get("q8_left_job_reasons", "")))
    quit_selected = "quit" in q8_tokens
    if "other" not in q8_tokens and has_value(row, "q8_other_text"):
        add_auto_clear_issue(issues, sid, "q8_other_text", row, "'other' is not selected in q8_left_job_reasons")

    q8a_codes_filled = has_value(row, "q8a_quit_reasons")
    q8a_text_filled = has_value(row, "q8a_other_text")
    if not quit_selected:
        if q8a_codes_filled:
            add_auto_clear_issue(issues, sid, "q8a_quit_reasons", row, "'quit' is not selected in q8_left_job_reasons")
        if q8a_text_filled:
            add_auto_clear_issue(issues, sid, "q8a_other_text", row, "'quit' is not selected in q8_left_job_reasons")
    else:
        if is_online(row):
            if not (q8a_codes_filled or q8a_text_filled):
                issues.append(make_issue(
                    sid,
                    "q8a_other_text",
                    "",
                    "D_conditional_missing",
                    "Online quit follow-up is blank even though 'quit' is selected",
                    f"[{sid}] 'quit' is selected in q8_left_job_reasons, but both q8a_quit_reasons "
                    f"and q8a_other_text are blank. Was the follow-up intentionally left blank?",
                ))
        elif not q8a_codes_filled:
            issues.append(make_issue(
                sid,
                "q8a_quit_reasons",
                "",
                "D_conditional_missing",
                "'quit' is selected in q8_left_job_reasons but q8a_quit_reasons is blank",
                f"[{sid}] 'quit' is selected in q8_left_job_reasons but q8a_quit_reasons is blank. "
                f"Was this intentionally left blank on the survey?",
            ))

    q9_tokens = set(split_pipe(row.get("q9_bank_account", "")))
    no_account = bool(q9_tokens & {"had_in_past", "never_had"})
    q9a_filled = has_value(row, "q9a_no_account_reasons")
    if not no_account and q9a_filled:
        add_auto_clear_issue(issues, sid, "q9a_no_account_reasons", row, "q9_bank_account has no had_in_past/never_had selection")
    if no_account and not q9a_filled:
        issues.append(make_issue(
            sid,
            "q9a_no_account_reasons",
            "",
            "D_conditional_missing",
            "q9a_no_account_reasons is blank but q9_bank_account shows no current account",
            f"[{sid}] q9_bank_account indicates no current account, but q9a_no_account_reasons is blank. "
            f"Was this intentionally left blank on the survey?",
        ))

    q9a_tokens = set(split_pipe(row.get("q9a_no_account_reasons", "")))
    if "tried_and_failed" not in q9a_tokens and has_value(row, "q9a_tried_failed_text"):
        add_auto_clear_issue(issues, sid, "q9a_tried_failed_text", row, "'tried_and_failed' is not selected in q9a_no_account_reasons")
    if "other" not in q9a_tokens and has_value(row, "q9a_other_text"):
        add_auto_clear_issue(issues, sid, "q9a_other_text", row, "'other' is not selected in q9a_no_account_reasons")

    q11_tokens = set(split_pipe(row.get("q11_program_helped", "")))
    if "something_else" not in q11_tokens and has_value(row, "q11_something_else_text"):
        add_auto_clear_issue(issues, sid, "q11_something_else_text", row, "'something_else' is not selected in q11_program_helped")


# ---------------------------------------------------------------------------
# Rule E - Missing required fields
# ---------------------------------------------------------------------------

def check_E(row, issues):
    sid = row["survey_id"]
    for field in REQUIRED_FIELDS:
        if is_blank(row.get(field, "")):
            issues.append(make_issue(
                sid,
                field,
                "",
                "E_missing_required",
                f"Required field '{field}' is blank",
                f"[{sid}] Required field '{field}' is blank. Was this left blank on the survey, "
                f"or is there a value that was missed during extraction?",
            ))


# ---------------------------------------------------------------------------
# Rule F - Free-text / demographic normalization
# ---------------------------------------------------------------------------

def check_F(row, issues):
    sid = row["survey_id"]

    dob = str(row.get("dob", "")).strip()
    if dob:
        try:
            pd.to_datetime(dob)
        except Exception:
            issues.append(make_issue(
                sid,
                "dob",
                dob,
                "F_invalid_date",
                f"'{dob}' cannot be parsed as a date",
                f"[{sid}] dob='{dob}' is not a valid date. What is the correct date of birth?",
            ))

    gender = str(row.get("gender", "")).strip()
    if gender:
        if gender in GENDER_NORMALIZATION:
            issues.append(make_demographic_recode_issue(sid, "gender", gender, GENDER_NORMALIZATION[gender]))
        elif gender not in KNOWN_GENDER_LABELS:
            issues.append(make_demographic_review_issue(
                sid,
                "gender",
                gender,
                f"'{gender}' is not a standard printed label",
            ))

    orientation = str(row.get("sexual_orientation", "")).strip()
    if orientation:
        if orientation in ORIENTATION_NORMALIZATION:
            issues.append(make_demographic_recode_issue(
                sid,
                "sexual_orientation",
                orientation,
                ORIENTATION_NORMALIZATION[orientation],
            ))
        elif orientation not in KNOWN_ORIENTATION_LABELS:
            parts = comma_parts(orientation)
            if len(parts) > 1 and all(part in KNOWN_ORIENTATION_LABELS for part in parts):
                issues.append(make_demographic_review_issue(
                    sid,
                    "sexual_orientation",
                    orientation,
                    "contains multiple recognizable orientation labels and needs reviewer grouping",
                ))
            else:
                issues.append(make_demographic_review_issue(
                    sid,
                    "sexual_orientation",
                    orientation,
                    f"'{orientation}' is not a standard printed label",
                ))

    for token in split_pipe(row.get("race_ethnicity", "")):
        if token in RACE_NORMALIZATION:
            issues.append(make_demographic_recode_issue(sid, "race_ethnicity", token, RACE_NORMALIZATION[token]))
        elif token not in KNOWN_RACE_LABELS:
            issues.append(make_demographic_review_issue(
                sid,
                "race_ethnicity",
                token,
                f"Race token '{token}' is not a recognized printed label",
            ))


# ---------------------------------------------------------------------------
# Rule G - Low-confidence flags
# ---------------------------------------------------------------------------

def check_G(row, issues):
    if is_online(row):
        return

    sid = row["survey_id"]
    for column in row.index:
        if not column.endswith("_conf"):
            continue
        value = row[column]
        if is_blank(value):
            continue
        try:
            conf = float(value)
        except (ValueError, TypeError):
            continue
        if conf < LOW_CONF_THRESHOLD:
            field = column[:-5]
            current_value = row.get(field, "")
            issues.append(make_issue(
                sid,
                field,
                current_value,
                "G_low_confidence",
                f"Confidence {conf:.2f} < threshold {LOW_CONF_THRESHOLD}",
                f"[{sid}] {field}='{current_value}' - confidence {conf:.2f}. Please verify this extraction is correct.",
            ))


# ---------------------------------------------------------------------------
# Reviewer workbook
# ---------------------------------------------------------------------------

ISSUE_CATEGORIES = {
    "A_invalid_range": ("A - Type / Range", "FF6B6B"),
    "B_invalid_category": ("B - Invalid Category", "FFA07A"),
    "C_invalid_token": ("C - Invalid Array Token", "FFD700"),
    "D_auto_clear": ("D - Auto-Clear Child", "ADD8E6"),
    "D_conditional_missing": ("D - Conditional Missing", "B0E0E6"),
    "E_missing_required": ("E - Missing Required", "FFB6C1"),
    "F_invalid_date": ("F - Invalid Date", "98FB98"),
    "F_likely_typo": ("F - Likely Typo", "90EE90"),
    "F_self_describe": ("F - Self-Describe", "ADFF2F"),
    "F_unrecognized_race_label": ("F - Unrecognized Race Label", "7FFF00"),
    "G_low_confidence": ("G - Low Confidence", "D3D3D3"),
}

ACCEPTED_TYPES = {
    "D_auto_clear",
    "D_conditional_missing",
    "E_missing_required",
    "G_low_confidence",
}

HEADERS = [
    "Issue Category", "survey_id", "field", "current_value",
    "detail", "question_for_reviewer", "suggested_correction",
    "action", "corrected_value", "apply_to", "notes",
]

COL_WIDTHS = [28, 12, 30, 35, 50, 80, 35, 14, 30, 16, 25]

INSTRUCTIONS = (
    "HOW TO USE THIS WORKBOOK:\n"
    "For each active QA row, fill in columns H-J:\n"
    "  Difficult self-described demographic rows stay on QA Questions with no default action; use column G for the allowed standard values before approving a recode or accept decision.\n"
    "  action          -> clear | recode | accept\n"
    "  corrected_value -> new value if action=recode (leave blank otherwise)\n"
    "  apply_to        -> this_survey | all_surveys\n"
    "Accepted items are listed on the separate tab for reference only."
)


def _write_issue_rows(ws, issues, header_fill, header_font, include_reviewer_cols):
    reviewer_fill = PatternFill("solid", fgColor="FFF9C4")

    ws.append(HEADERS)
    for index, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    if include_reviewer_cols:
        for column in (8, 9, 10, 11):
            ws.cell(1, column).fill = PatternFill("solid", fgColor="F9A825")
    ws.row_dimensions[1].height = 30

    for issue in issues:
        label, hex_color = ISSUE_CATEGORIES.get(issue["issue_type"], (issue["issue_type"], "FFFFFF"))
        ws.append([
            label,
            issue["survey_id"],
            issue["field"],
            issue["current_value"],
            issue["detail"],
            issue["question_for_reviewer"],
            issue["suggested_correction"],
            issue["action"],
            issue["corrected_value"],
            issue["apply_to"],
            issue["notes"],
        ])
        row_fill = PatternFill("solid", fgColor=hex_color)
        for index, cell in enumerate(ws[ws.max_row], 1):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = row_fill if index <= 7 else reviewer_fill

    for index, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def issue_needs_reviewer_input(issue: dict) -> bool:
    if issue["issue_type"] in ACCEPTED_TYPES:
        return False
    return str(issue.get("action", "")).strip().lower() != "accept"


def write_reviewer_workbook(issues: list[dict], out_path: Path):
    active = [issue for issue in issues if issue_needs_reviewer_input(issue)]
    accepted = [issue for issue in issues if not issue_needs_reviewer_input(issue)]

    workbook = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    ws_instructions = workbook.active
    ws_instructions.title = "Instructions"
    ws_instructions["A1"] = INSTRUCTIONS
    ws_instructions["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_instructions.row_dimensions[1].height = 120
    ws_instructions.column_dimensions["A"].width = 90

    ws_active = workbook.create_sheet("QA Questions")
    _write_issue_rows(ws_active, active, header_fill, header_font, include_reviewer_cols=True)

    end_row = len(active) + 1
    if end_row >= 2:
        dv_action = DataValidation(
            type="list",
            formula1='"clear,recode,accept"',
            allow_blank=True,
            showDropDown=False,
        )
        ws_active.add_data_validation(dv_action)
        dv_action.add(f"H2:H{end_row}")

        dv_scope = DataValidation(
            type="list",
            formula1='"this_survey,all_surveys"',
            allow_blank=True,
            showDropDown=False,
        )
        ws_active.add_data_validation(dv_scope)
        dv_scope.add(f"J2:J{end_row}")

    ws_accepted = workbook.create_sheet("Accepted - No Action")
    _write_issue_rows(ws_accepted, accepted, header_fill, header_font, include_reviewer_cols=False)

    ws_summary = workbook.create_sheet("Summary")
    ws_summary.append(["Issue Category", "Count", "Tab"])
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    counts_active = Counter(issue["issue_type"] for issue in active)
    counts_accepted = Counter(issue["issue_type"] for issue in accepted)
    for issue_type, (label, hex_color) in ISSUE_CATEGORIES.items():
        count = counts_active.get(issue_type, 0) + counts_accepted.get(issue_type, 0)
        if counts_active.get(issue_type, 0) and counts_accepted.get(issue_type, 0):
            tab_name = "Both"
        elif counts_active.get(issue_type, 0):
            tab_name = "QA Questions"
        else:
            tab_name = "Accepted - No Action"
        ws_summary.append([label, count, tab_name])
        row_fill = PatternFill("solid", fgColor=hex_color)
        for cell in ws_summary[ws_summary.max_row]:
            cell.fill = row_fill

    ws_summary.append(["TOTAL - needs review", len(active), "QA Questions"])
    ws_summary.append(["TOTAL - accepted", len(accepted), "Accepted - No Action"])
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 25

    workbook.save(str(out_path))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    csv_path = BASE_DIR / "output" / "IL" / "survey_data_IL.csv"
    if not csv_path.exists():
        print(f"CSV not found: {csv_path}")
        sys.exit(1)

    df = pd.read_csv(str(csv_path), encoding="utf-8-sig", dtype=str).fillna("")
    print(f"Loaded {len(df)} surveys, {len(df.columns)} columns.")

    all_issues: list[dict] = []
    for _, row in df.iterrows():
        check_A(row, all_issues)
        check_B(row, all_issues)
        check_C(row, all_issues)
        check_D(row, all_issues)
        check_E(row, all_issues)
        check_F(row, all_issues)
        check_G(row, all_issues)

    out_dir = BASE_DIR / "output" / "IL"
    flag_path = out_dir / "flagged_IL.csv"
    xlsx_path = out_dir / "qa_questions_IL.xlsx"

    pd.DataFrame(all_issues).to_csv(str(flag_path), index=False, encoding="utf-8-sig")
    write_reviewer_workbook(all_issues, xlsx_path)

    active = [issue for issue in all_issues if issue_needs_reviewer_input(issue)]
    accepted = [issue for issue in all_issues if not issue_needs_reviewer_input(issue)]
    counts_active = Counter(issue["issue_type"] for issue in active)
    counts_accepted = Counter(issue["issue_type"] for issue in accepted)

    print(f"\n{'=' * 62}")
    print(f"QA Summary -- {len(df)} surveys, {len(all_issues)} total issues")
    print(f"{'=' * 62}")
    print("  Needs review (QA Questions tab):")
    for issue_type, (label, _) in ISSUE_CATEGORIES.items():
        count = counts_active.get(issue_type, 0)
        if count:
            print(f"    {label:<38} {count:>4}")
    print("  Accepted -- no action (separate tab):")
    for issue_type, (label, _) in ISSUE_CATEGORIES.items():
        count = counts_accepted.get(issue_type, 0)
        if count:
            print(f"    {label:<38} {count:>4}")
    print(f"  {'-' * 46}")
    print(f"  {'Needs review':<40} {len(active):>4}")
    print(f"  {'Accepted':<40} {len(accepted):>4}")
    print("\nOutputs written:")
    print(f"  {flag_path}")
    print(f"  {xlsx_path}")


if __name__ == "__main__":
    main()